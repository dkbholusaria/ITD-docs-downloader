import os
import re
import sys
import csv
import json
import uuid
import datetime
from base64 import urlsafe_b64encode
from cryptography.fernet import Fernet
from cryptography.hazmat.primitives import hashes
from cryptography.hazmat.primitives.kdf.pbkdf2 import PBKDF2HMAC

_PAN_RE = re.compile(r'^[A-Z]{3}[PCHFATBLJG][A-Z][0-9]{4}[A-Z]$')
_DOB_RE = re.compile(r'^(\d{2})-(\d{2})-(\d{4})$')

# Accepted import formats → normalised DD-MM-YYYY
_DOB_FORMATS = [
    '%d-%m-%Y',   # 06-07-1974  (already correct)
    '%d/%m/%Y',   # 06/07/1974
    '%d.%m.%Y',   # 06.07.1974
    '%Y-%m-%d',   # 1974-07-06  (ISO)
    '%Y/%m/%d',   # 1974/07/06
    '%d-%m-%y',   # 06-07-74
    '%d/%m/%y',   # 06/07/74
    '%d.%m.%y',   # 06.07.74
    '%d %m %Y',   # 06 07 1974
    '%d %B %Y',   # 06 July 1974
    '%d-%B-%Y',   # 06-July-1974
]

def _normalise_dob(raw: str) -> str:
    """Convert any recognised date string to DD-MM-YYYY; return raw if unrecognised."""
    # Strip time component if present (e.g. "06/07/1974 00:00:00")
    s = raw.strip().split(' ')[0].split('T')[0]
    for fmt in _DOB_FORMATS:
        try:
            return datetime.datetime.strptime(s, fmt).strftime('%d-%m-%Y')
        except ValueError:
            continue
    return raw.strip()  # let _validate_fields report the error with the original value

def _validate_fields(name: str, pan: str, dob: str, password: str):
    name = name.strip()
    pan  = pan.strip().upper()
    dob  = dob.strip()

    if not name:
        raise ValueError("Full Name is required.")
    if len(name) < 2:
        raise ValueError("Full Name must be at least 2 characters.")

    if not pan:
        raise ValueError("PAN Number is required.")
    if not _PAN_RE.match(pan):
        raise ValueError(
            "Invalid PAN format.\n\n"
            "Format: AAA · T · N · 0001 · Z\n"
            "  · Characters 1–3 : any letters (A–Z)\n"
            "  · Character 4    : P/C/H/F/A/T/B/L/J/G (taxpayer type)\n"
            "  · Character 5    : any letter (A–Z)\n"
            "  · Characters 6–9 : 4 digits\n"
            "  · Character 10   : any letter (A–Z)\n\n"
            "Example: AAAPT0001A")

    if not dob:
        raise ValueError("Date of Birth is required.")
    m = _DOB_RE.match(dob)
    if not m:
        raise ValueError("Date of Birth must be in DD-MM-YYYY format.\nExample: 01-01-1980")
    day, month, year = int(m.group(1)), int(m.group(2)), int(m.group(3))
    try:
        dt = datetime.date(year, month, day)
    except ValueError:
        raise ValueError(f"'{dob}' is not a valid date. Please use DD-MM-YYYY.")
    today = datetime.date.today()
    if dt >= today:
        raise ValueError("Date of Birth cannot be today or a future date.")
    if year < 1900:
        raise ValueError("Date of Birth year seems incorrect (before 1900).")

    if not password:
        raise ValueError("Portal Password is required.")
    if len(password) < 4:
        raise ValueError("Password is too short (minimum 4 characters).")

class VaultManager:
    """
    Manages secure CRUD operations for Standalone Tax Downloader.
    Stores details in an encrypted local JSON vault.
    """
    def __init__(self, vault_path=None, master_password="automated_tax_app_key"):
        if vault_path is None:
            # When frozen by PyInstaller use folder next to .exe, not _MEIPASS
            if getattr(sys, "frozen", False):
                base_dir = os.path.dirname(sys.executable)
            else:
                base_dir = os.path.dirname(os.path.abspath(__file__))
            self.vault_file = os.path.join(base_dir, "tax_vault.json")
        else:
            self.vault_file = vault_path
        
        # Derive cryptographic key
        salt = b'secure_tax_salt'
        kdf = PBKDF2HMAC(
            algorithm=hashes.SHA256(),
            length=32,
            salt=salt,
            iterations=100000,
        )
        self._key = urlsafe_b64encode(kdf.derive(master_password.encode()))
        self._cipher = Fernet(self._key)
        
        # Initialize and migrate/ensure schema
        self._ensure_vault()

    def _ensure_vault(self):
        dir_name = os.path.dirname(self.vault_file)
        if dir_name:
            os.makedirs(dir_name, exist_ok=True)
        if not os.path.exists(self.vault_file):
            self._save_raw({"assessees": [], "settings": {}})
        else:
            data = self._get_raw()
            updated = False
            if "assessees" not in data:
                data["assessees"] = []
                updated = True
            if "settings" not in data:
                data["settings"] = {}
                updated = True

            # Migrate old entries if needed (add uuid, clean up keys)
            for entry in data["assessees"]:
                if "id" not in entry:
                    entry["id"] = str(uuid.uuid4())
                    updated = True
                if "pan" in entry:
                    entry["pan"] = entry["pan"].strip().upper()
                if "email" not in entry:
                    entry["email"] = ""
                    updated = True
                if "cc" not in entry:
                    entry["cc"] = ""
                    updated = True

            if updated:
                self._save_raw(data)

    def _get_raw(self):
        try:
            with open(self.vault_file, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception:
            return {"assessees": [], "settings": {}}

    def _save_raw(self, data):
        with open(self.vault_file, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=4)

    def encrypt_password(self, password: str) -> str:
        if not password:
            return ""
        return self._cipher.encrypt(password.encode('utf-8')).decode('utf-8')

    def decrypt_password(self, encrypted_pwd: str) -> str:
        if not encrypted_pwd:
            return ""
        try:
            return self._cipher.decrypt(encrypted_pwd.encode('utf-8')).decode('utf-8')
        except Exception:
            return ""

    # --- Assessee CRUD Operations ---

    def get_all_assessees(self):
        """Returns all assessees with decrypted passwords for utility."""
        raw_data = self._get_raw()
        assessees = []
        for a in raw_data.get("assessees", []):
            decrypted = a.copy()
            decrypted["password"] = self.decrypt_password(a.get("password_enc", ""))
            assessees.append(decrypted)
        return assessees

    def add_update_assessee(self, name: str, pan: str, dob: str, password: str,
                            assessee_id: str = None, email: str = "", cc: str = "",
                            group: str = "") -> str:
        """Adds or updates a single assessee."""
        raw_data = self._get_raw()
        _validate_fields(name, pan, dob, password)
        pan = pan.strip().upper()

        password_enc = self.encrypt_password(password)

        found = False
        if assessee_id:
            # Update by ID
            for i, a in enumerate(raw_data["assessees"]):
                if a.get("id") == assessee_id:
                    raw_data["assessees"][i] = {
                        "id": assessee_id,
                        "name": name.strip(),
                        "pan": pan,
                        "dob": dob.strip(),
                        "password_enc": password_enc,
                        "email": email.strip(),
                        "cc": cc.strip(),
                        "group": group.strip(),
                    }
                    found = True
                    break
        else:
            # Check if PAN already exists to update it, or treat as new
            for i, a in enumerate(raw_data["assessees"]):
                if a.get("pan") == pan:
                    assessee_id = a.get("id") or str(uuid.uuid4())
                    raw_data["assessees"][i] = {
                        "id": assessee_id,
                        "name": name.strip(),
                        "pan": pan,
                        "dob": dob.strip(),
                        "password_enc": password_enc,
                        "email": email.strip(),
                        "cc": cc.strip(),
                        "group": group.strip(),
                    }
                    found = True
                    break

        if not found:
            new_id = str(uuid.uuid4())
            raw_data["assessees"].append({
                "id": new_id,
                "name": name.strip(),
                "pan": pan,
                "dob": dob.strip(),
                "password_enc": password_enc,
                "email": email.strip(),
                "cc": cc.strip(),
                "group": group.strip(),
            })
            assessee_id = new_id

        self._save_raw(raw_data)
        return assessee_id

    # --- Client Groups (F-11) ---
    # A group is just a shared value of one client's own "group" field —
    # same lightweight model download_history already uses for AY labels
    # (a plain string, not a managed entity with its own storage), so no
    # migration is needed for vault files written before this feature.

    def get_all_groups(self) -> list:
        """Distinct non-empty group names currently in use, sorted — powers
        every group dropdown/combo so they never drift from what clients
        are actually tagged with."""
        raw_data = self._get_raw()
        groups = {a.get("group", "").strip() for a in raw_data.get("assessees", [])}
        groups.discard("")
        return sorted(groups)

    def set_client_group(self, assessee_id: str, group: str) -> bool:
        """Dedicated single-field setter for bulk-assign — avoids routing
        through add_update_assessee(), which needs every other field
        re-supplied. Returns True if the client was found."""
        raw_data = self._get_raw()
        for a in raw_data.get("assessees", []):
            if a.get("id") == assessee_id:
                a["group"] = group.strip()
                self._save_raw(raw_data)
                return True
        return False

    def rename_group(self, old: str, new: str) -> int:
        """Renames a group across every client currently in it. Returns
        the number of clients updated."""
        raw_data = self._get_raw()
        new = new.strip()
        count = 0
        for a in raw_data.get("assessees", []):
            if a.get("group", "").strip() == old:
                a["group"] = new
                count += 1
        if count:
            self._save_raw(raw_data)
        return count

    def clear_group(self, group: str) -> int:
        """"Deletes" a group by un-grouping every client in it — never
        deletes the clients themselves. Returns the number un-grouped."""
        return self.rename_group(group, "")

    def update_assessee_email(self, pan: str, email: str, cc: str = "") -> bool:
        """Update only the email and cc fields for a client by PAN. Returns True if found."""
        raw_data = self._get_raw()
        pan = pan.strip().upper()
        for entry in raw_data["assessees"]:
            if entry.get("pan") == pan:
                entry["email"] = email.strip()
                entry["cc"] = cc.strip()
                self._save_raw(raw_data)
                return True
        return False

    def delete_assessee(self, assessee_id: str):
        """Deletes an assessee by ID."""
        raw_data = self._get_raw()
        raw_data["assessees"] = [a for a in raw_data["assessees"] if a.get("id") != assessee_id]
        self._save_raw(raw_data)

    # --- Download History ---
    # Keyed (pan, ay_label, doc_type) so a batch touching several document
    # types (e.g. 26AS + Filed Returns) for the same client/AY doesn't have
    # one call's status silently overwrite another's.

    # Worst-status-wins ranking for collapsing several doc types into one
    # grid cell — higher rank displays when multiple doc types disagree.
    _STATUS_RANK = {"❌": 4, "⚠": 3, "🕐": 2, "⏹": 2, "⬜": 1, "✅": 0}

    @staticmethod
    def _status_rank(status: str) -> int:
        for prefix, rank in VaultManager._STATUS_RANK.items():
            if status.startswith(prefix):
                return rank
        return 0

    def record_download(self, pan: str, ay_label: str, doc_type: str, status: str, path: str):
        """Record the last download status + path for a client/AY/doc_type triple."""
        raw_data = self._get_raw()
        pan = pan.strip().upper()
        hist = raw_data.setdefault("download_history", {})
        ay_map = hist.setdefault(pan, {})
        existing = ay_map.get(ay_label)
        # Backward compatibility: older versions stored the leaf directly as
        # {"status", "path", "ts"} with no doc_type axis. Migrate any such
        # entry in place to {"legacy": <old leaf>} before adding doc_type,
        # so old data is re-homed rather than lost.
        if isinstance(existing, dict) and "status" in existing:
            ay_map[ay_label] = {"legacy": existing}
        doc_map = ay_map.setdefault(ay_label, {})
        doc_map[doc_type] = {
            "status": status,
            "path": path,
            "ts": datetime.datetime.now().strftime("%d-%b-%Y %H:%M:%S"),
        }
        self._save_raw(raw_data)

    def get_download_history(self, ay_label: str) -> dict:
        """Return {pan: {doc_type: {"status":..., "path":..., "ts":...}}} for the given AY label."""
        raw_data = self._get_raw()
        hist = raw_data.get("download_history", {})
        result = {}
        for pan, ay_map in hist.items():
            entry = ay_map.get(ay_label)
            if entry is None:
                continue
            # Backward compat for reads of never-migrated old-shape entries.
            if "status" in entry:
                entry = {"legacy": entry}
            result[pan] = entry
        return result

    def get_download_history_summary(self, ay_label: str) -> dict:
        """Return {pan: {"status": <worst doc_type's status text>, "ts": ..., "path": ...,
        "breakdown": [(doc_type, status_text), ...]}} — collapses multiple
        doc types into one display-ready summary per client for the grid."""
        history = self.get_download_history(ay_label)
        summary = {}
        for pan, doc_map in history.items():
            if not doc_map:
                continue
            breakdown = sorted(doc_map.items(), key=lambda kv: kv[0])
            worst_doc_type, worst_entry = max(
                doc_map.items(), key=lambda kv: self._status_rank(kv[1].get("status", ""))
            )
            summary[pan] = {
                "status": worst_entry.get("status", ""),
                "ts": worst_entry.get("ts", ""),
                "path": worst_entry.get("path", ""),
                "breakdown": [(dt, e.get("status", "")) for dt, e in breakdown],
            }
        return summary

    # --- ITR Processing Status (F-67) ---
    # Separate top-level key from download_history — this isn't a document
    # download record, it's the portal's own status-timeline text for a
    # client/AY, refreshed by an explicit "Update" action, not implied by
    # any download having happened.

    def record_return_status(self, pan: str, ay_label: str, status: str,
                              status_date: str = "", filing_date: str = "", ack_no: str = ""):
        """Record the latest known ITR processing status for a client/AY.
        `ts` is when THIS app last checked, distinct from both `status_date`
        (the date the PORTAL itself shows against that status step, e.g.
        "Aug 3, 2026") and `filing_date` (when the return was filed)."""
        raw_data = self._get_raw()
        pan = pan.strip().upper()
        hist = raw_data.setdefault("return_status_history", {})
        ay_map = hist.setdefault(pan, {})
        ay_map[ay_label] = {
            "status": status,
            "status_date": status_date,
            "filing_date": filing_date,
            "ack_no": ack_no,
            "ts": datetime.datetime.now().strftime("%d-%b-%Y %H:%M:%S"),
        }
        self._save_raw(raw_data)

    def get_return_status(self, ay_label: str) -> dict:
        """{pan: {"status", "status_date", "filing_date", "ack_no", "ts"}}
        for one AY — only clients that have actually been checked at least
        once."""
        raw_data = self._get_raw()
        hist = raw_data.get("return_status_history", {})
        return {
            pan: ay_map[ay_label]
            for pan, ay_map in hist.items()
            if ay_label in ay_map
        }

    def get_return_status_all(self) -> dict:
        """{pan: {ay_label: {"status", "status_date", "filing_date",
        "ack_no", "ts"}}} — every client/AY ever checked, for the status
        window's "All Years" view."""
        raw_data = self._get_raw()
        return raw_data.get("return_status_history", {})

    # --- Bulk Import / Export ---

    def import_bulk(self, file_path: str) -> tuple:
        """
        Imports assessees from an Excel (.xlsx) or CSV (.csv) file.
        Expects columns: Name, PAN, DOB, Password
        Returns: (success_count, error_messages_list)
        """
        if not os.path.exists(file_path):
            return 0, [f"File {file_path} does not exist."]
        
        try:
            if file_path.endswith('.xlsx'):
                from openpyxl import load_workbook
                wb = load_workbook(file_path, data_only=True)
                # Files this app writes now carry a leading "Instructions"
                # sheet (see _write_client_table_file) — pick "Clients"
                # explicitly rather than trusting wb.active, which just
                # reflects whichever sheet was active when last saved.
                ws = wb["Clients"] if "Clients" in wb.sheetnames else wb.active
                raw_rows = list(ws.iter_rows(values_only=True))
                if not raw_rows:
                    return 0, ["File is empty."]
                headers = [str(c).strip().lower() if c is not None else "" for c in raw_rows[0]]
                # BUG FIX (2026-09-03): padding the Clients table out to a
                # 50-row minimum for bulk entry means openpyxl reports those
                # blank padded rows as real rows on read-back — filter out
                # anything with no actual content before validating rows.
                data_rows = [
                    row for row in raw_rows[1:]
                    if any(c is not None and str(c).strip() != "" for c in row)
                ]
            elif file_path.endswith('.csv'):
                with open(file_path, newline='', encoding='utf-8-sig') as f:
                    reader = csv.reader(f)
                    raw_rows = list(reader)
                if not raw_rows:
                    return 0, ["File is empty."]
                headers = [c.strip().lower() for c in raw_rows[0]]
                data_rows = [row for row in raw_rows[1:] if any(c.strip() for c in row)]
            else:
                return 0, ["Unsupported file format. Please use Excel (.xlsx) or CSV (.csv)."]
        except Exception as e:
            return 0, [f"Failed to read file: {str(e)}"]

        required_cols = {'name', 'pan', 'dob', 'password'}
        missing = required_cols - set(headers)
        if missing:
            return 0, [f"Missing required columns: {', '.join(missing)}. Headers must include: Name, PAN, DOB, Password."]

        col = {h: i for i, h in enumerate(headers)}
        added_count = 0
        updated_count = 0
        errors = []

        existing_pans = {a.get("pan") for a in self.get_all_assessees()}

        for idx, row in enumerate(data_rows):
            row_num = idx + 2
            try:
                def _cell(key):
                    v = row[col[key]] if key in col and col[key] < len(row) else None
                    return str(v).strip() if v is not None and str(v).strip() not in ("", "None") else ""

                name = _cell('name')
                pan = _cell('pan').upper()

                dob_val = row[col['dob']] if col['dob'] < len(row) else None
                if dob_val is None or str(dob_val).strip() in ("", "None"):
                    dob = ""
                elif isinstance(dob_val, (datetime.datetime, datetime.date)):
                    dob = dob_val.strftime('%d-%m-%Y')
                else:
                    dob = _normalise_dob(str(dob_val).strip())

                password = _cell('password')
                email = _cell('email') if 'email' in col else ""
                cc = _cell('cc') if 'cc' in col else ""
                group = _cell('group') if 'group' in col else ""

                if not name or not pan or not dob or not password:
                    errors.append(f"Row {row_num}: Missing values in Name, PAN, DOB, or Password.")
                    continue

                if len(pan) != 10:
                    errors.append(f"Row {row_num}: Invalid PAN length (must be 10 characters).")
                    continue

                is_existing = pan in existing_pans
                self.add_update_assessee(name, pan, dob, password, email=email, cc=cc, group=group)
                if is_existing:
                    updated_count += 1
                else:
                    existing_pans.add(pan)
                    added_count += 1
            except Exception as e:
                errors.append(f"Row {row_num}: Error importing entry: {str(e)}")

        return added_count, updated_count, errors

    def generate_template(self, file_path: str):
        """Generates an Excel import template with sample columns."""
        headers = ["Name", "Group", "PAN", "DOB", "Password", "Email", "CC"]
        sample = ["John Doe", "", "AAAPT0001A", "01-01-1980", "YourPortalPassword",
                  "client@example.com", "spouse@example.com;accountant@firm.com"]
        target = file_path if (file_path.endswith('.csv') or file_path.endswith('.xlsx')) else file_path + ".xlsx"
        self._write_client_table_file(target, headers, [sample])

    def export_data(self, file_path: str):
        """Exports all saved assessees (with decrypted passwords) to Excel or CSV."""
        assessees = self.get_all_assessees()
        headers = ["Name", "Group", "PAN", "DOB", "Password", "Email", "CC"]
        rows = [[a.get("name", ""), a.get("group", ""), a.get("pan", ""), a.get("dob", ""),
                 a.get("password", ""), a.get("email", ""), a.get("cc", "")]
                for a in assessees]
        self._write_client_table_file(file_path, headers, rows)

    def _client_instructions_text(self) -> str:
        """Plain-text version of the Client Master Instructions sheet, for
        the CSV path (CSV has no second sheet to carry it) — same wording
        as the Excel version, kept in one place per column below."""
        lines = [
            "Client Master — Import Template",
            "AayDoc Capio™  ·  © 2026  ·  Developed by CA. Deepak Bhholusaria  ·  "
            "linkedin.com/in/bhholusaria  ·  deepak@ailearrning.guru",
            "",
            "HOW TO FILL IN THIS TEMPLATE",
            "=" * 29,
            "",
            "One row per client.",
            "",
            "Name",
            "  The client's name, just for you to recognise them by.",
            "",
            "Group",
            "  Optional — e.g. a family or firm name (\"Bholusaria Family\"). Leave "
            "blank if this client isn't part of a group. Importing the same group "
            "name for several clients groups them together in the app.",
            "",
            "PAN",
            "  The client's 10-character PAN, e.g. AAAPT0001A. If a PAN already "
            "exists in AayDocCapio, importing it again updates that client instead "
            "of adding a duplicate.",
            "",
            "DOB",
            "  Date of birth, as DD-MM-YYYY (e.g. 01-01-1980). A few other common "
            "date formats are also accepted.",
            "",
            "Password",
            "  The client's ITD e-Filing portal password.",
            "",
            "Email / CC",
            "  Optional. Email is who reports get sent to; CC can have more than "
            "one address, separated by semicolons.",
            "",
            "Note: Name, PAN, DOB, and Password are all required — a row missing "
            "any of these is skipped and listed as an error when you import it.",
            "",
            "Note: this file has real portal passwords in plain text once you fill "
            "it in (or once you export your saved clients) — keep it somewhere safe "
            "and delete it when you're done with it.",
        ]
        return "\n".join(lines)

    def _write_client_table_file(self, path, headers, rows):
        """Shared writer for both the import template and the full-data
        export — same branded/table-formatted layout the bulk tax challan
        template uses (see ui/dialogs.py's _write_table_file), kept
        consistent across the app's two Excel import flows."""
        if path.endswith(".csv"):
            with open(path, 'w', newline='', encoding='utf-8') as f:
                w = csv.writer(f)
                w.writerow(headers)
                w.writerows(rows)
            instructions_path = re.sub(r"\.csv$", "", path, flags=re.IGNORECASE) + "_Instructions.txt"
            with open(instructions_path, "w", encoding="utf-8") as f:
                f.write(self._client_instructions_text())
            return

        from openpyxl import Workbook
        from openpyxl.worksheet.table import Table, TableStyleInfo
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "Clients"
        ws.append(headers)
        for row in rows:
            ws.append(row)

        last_col_letter = get_column_letter(len(headers))
        MIN_TEMPLATE_ROWS = 50
        table_last_row = max(len(rows), MIN_TEMPLATE_ROWS) + 1
        tab = Table(displayName="Clients", ref=f"A1:{last_col_letter}{table_last_row}")
        tab.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False,
            showRowStripes=True, showColumnStripes=False,
        )
        ws.add_table(tab)
        ws.freeze_panes = "A2"
        col_widths = {"Name": 22, "Group": 20, "PAN": 14, "DOB": 12, "Password": 20, "Email": 26, "CC": 30}
        for i, label in enumerate(headers, start=1):
            ws.column_dimensions[get_column_letter(i)].width = col_widths.get(label, 16)
        for row_cells in ws.iter_rows(min_row=2, max_row=max(table_last_row, 2), max_col=len(headers)):
            for cell in row_cells:
                cell.alignment = Alignment(vertical="center")

        # ── "Instructions" sheet — same brand banner + Column/What to enter ─
        # table layout as the tax challan template (ui/dialogs.py), so both
        # of the app's import flows read as the same product.
        NAVY = "0A1628"
        GREY = "94A3B8"
        BANNER_FILL = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
        HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        BAND_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        NOTE_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        thin = Side(style="thin", color="BFBFBF")
        box_border = Border(left=thin, right=thin, top=thin, bottom=thin)

        ws_help = wb.create_sheet("Instructions", 0)
        ws_help.sheet_view.showGridLines = False
        ws_help.column_dimensions["A"].width = 22
        ws_help.column_dimensions["B"].width = 95

        r = 1
        ws_help.merge_cells(f"A{r}:B{r}")
        title_cell = ws_help.cell(row=r, column=1, value="Client Master — Import Template")
        title_cell.font = Font(bold=True, size=13, color="FFFFFF")
        title_cell.fill = BANNER_FILL
        title_cell.alignment = Alignment(vertical="center", horizontal="center")
        ws_help.row_dimensions[r].height = 28
        r += 1

        ws_help.merge_cells(f"A{r}:B{r}")
        credit_cell = ws_help.cell(
            row=r, column=1,
            value="AayDoc Capio™  ·  © 2026  ·  Developed by CA. Deepak Bhholusaria  ·  "
                  "linkedin.com/in/bhholusaria  ·  deepak@ailearrning.guru",
        )
        credit_cell.font = Font(size=8, color=GREY)
        credit_cell.fill = BANNER_FILL
        credit_cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
        ws_help.row_dimensions[r].height = 18
        r += 2

        ws_help.merge_cells(f"A{r}:B{r}")
        heading_cell = ws_help.cell(row=r, column=1, value="How to fill in this template")
        heading_cell.font = Font(bold=True, size=13, color=NAVY)
        heading_cell.alignment = Alignment(vertical="center", indent=1)
        ws_help.row_dimensions[r].height = 22
        r += 1

        ws_help.merge_cells(f"A{r}:B{r}")
        intro_cell = ws_help.cell(row=r, column=1, value="One row per client.")
        intro_cell.font = Font(italic=True, color="595959")
        intro_cell.alignment = Alignment(wrap_text=True, vertical="top", indent=1)
        r += 2

        header_row = r
        for col, text in ((1, "Column"), (2, "What to enter")):
            c = ws_help.cell(row=header_row, column=col, value=text)
            c.font = Font(bold=True, color="1F4E78")
            c.fill = HEADER_FILL
            c.border = box_border
            c.alignment = Alignment(vertical="center", indent=1)
        r += 1

        sections = [
            ("Name", "The client's name, just for you to recognise them by."),
            ("Group",
             'Optional — e.g. a family or firm name ("Bholusaria Family"). Leave '
             "blank if this client isn't part of a group. Importing the same group "
             "name for several clients groups them together in the app."),
            ("PAN",
             "The client's 10-character PAN, e.g. AAAPT0001A. If a PAN already exists "
             "in AayDocCapio, importing it again updates that client instead of adding "
             "a duplicate."),
            ("DOB",
             "Date of birth, as DD-MM-YYYY (e.g. 01-01-1980). A few other common date "
             "formats are also accepted."),
            ("Password", "The client's ITD e-Filing portal password."),
            ("Email / CC",
             "Optional. Email is who reports get sent to; CC can have more than one "
             "address, separated by semicolons."),
        ]
        for i, (field, desc) in enumerate(sections):
            field_cell = ws_help.cell(row=r, column=1, value=field)
            desc_cell = ws_help.cell(row=r, column=2, value=desc)
            fill = BAND_FILL if i % 2 == 1 else None
            for c in (field_cell, desc_cell):
                c.border = box_border
                if fill:
                    c.fill = fill
            field_cell.font = Font(bold=True)
            field_cell.alignment = Alignment(wrap_text=True, vertical="top", indent=1)
            desc_cell.alignment = Alignment(wrap_text=True, vertical="top", indent=1)
            ws_help.row_dimensions[r].height = 36
            r += 1

        r += 1
        ws_help.merge_cells(f"A{r}:B{r}")
        note_cell = ws_help.cell(
            row=r, column=1,
            value="Note: Name, PAN, DOB, and Password are all required — a row missing "
                  "any of these is skipped and listed as an error when you import it.",
        )
        note_cell.font = Font(bold=True, color="7F6000")
        note_cell.fill = NOTE_FILL
        note_cell.border = box_border
        note_cell.alignment = Alignment(wrap_text=True, vertical="center", indent=1)
        ws_help.row_dimensions[r].height = 32
        r += 1

        ws_help.merge_cells(f"A{r}:B{r}")
        security_note_cell = ws_help.cell(
            row=r, column=1,
            value="Note: this file has real portal passwords in plain text once you "
                  "fill it in (or once you export your saved clients) — keep it "
                  "somewhere safe and delete it when you're done with it.",
        )
        security_note_cell.font = Font(bold=True, color="7F6000")
        security_note_cell.fill = NOTE_FILL
        security_note_cell.border = box_border
        security_note_cell.alignment = Alignment(wrap_text=True, vertical="center", indent=1)
        ws_help.row_dimensions[r].height = 32

        wb.active = wb["Instructions"]
        wb.save(path)

    # --- Settings Management ---

    def get_setting(self, key: str, default=None):
        raw_data = self._get_raw()
        return raw_data.get("settings", {}).get(key, default)

    def update_setting(self, key: str, value):
        raw_data = self._get_raw()
        raw_data["settings"][key] = value
        self._save_raw(raw_data)

    # --- Email / SMTP Settings ---

    _EMAIL_DEFAULTS = {
        "smtp_host": "",
        "smtp_port": "587",
        "smtp_user": "",
        "smtp_from": "",            # optional Send-As address; uses smtp_user when blank
        "smtp_password_enc": "",
        "smtp_use_tls": True,       # legacy — kept for back-compat; use smtp_encryption
        "smtp_encryption": "STARTTLS",  # "STARTTLS" | "SSL/TLS" | "None"
        "firm_name": "",
        "bcc_addresses": "",
        "email_subject_tpl": "[Action Required] Your Annual Income Tax Documents for {ay} | {client_name}",
        "email_body_tpl": (
            "<p>Dear {client_name},</p>"
            "<p>We hope this message finds you well.</p>"
            "<p>Please find attached your Income Tax documents for <b>{ay}</b>, "
            "as downloaded from the Income Tax Department's e-Filing portal:</p>"
            "<p>{documents}</p>"
            "<p>Kindly review the attached documents at your earliest convenience. "
            "If you notice any discrepancies or have any queries, please do not "
            "hesitate to contact us and we will be happy to assist you.</p>"
            "<p>Please note that these documents are sourced directly from the "
            "IT Department portal and are provided for your reference and records.</p>"
            "<p>Warm regards,<br><b>{firm_name}</b></p>"
        ),
    }

    def get_email_settings(self) -> dict:
        raw = self._get_raw().get("settings", {})
        cfg = {}
        for k, default in self._EMAIL_DEFAULTS.items():
            cfg[k] = raw.get(k, default)
        # Migrate legacy smtp_use_tls → smtp_encryption
        if "smtp_encryption" not in raw and "smtp_use_tls" in raw:
            cfg["smtp_encryption"] = "STARTTLS" if raw["smtp_use_tls"] else "None"
        # Migrate plain-text body template → HTML
        body = cfg.get("email_body_tpl", "")
        if body and not body.lstrip().startswith("<"):
            import html as _html
            cfg["email_body_tpl"] = (
                "<p style='white-space:pre-wrap'>"
                + _html.escape(body).replace("\n", "<br>")
                + "</p>"
            )
        # Decrypt password for caller
        cfg["smtp_password"] = self.decrypt_password(cfg.get("smtp_password_enc", ""))
        return cfg

    def save_email_settings(self, cfg: dict):
        """Save SMTP settings. Pass plain-text 'smtp_password'; it will be encrypted."""
        raw_data = self._get_raw()
        s = raw_data.setdefault("settings", {})
        for k in self._EMAIL_DEFAULTS:
            if k == "smtp_password_enc":
                continue
            if k in cfg:
                s[k] = cfg[k]
        if "smtp_password" in cfg and cfg["smtp_password"]:
            s["smtp_password_enc"] = self.encrypt_password(cfg["smtp_password"])
        elif "smtp_password" in cfg and not cfg["smtp_password"]:
            # Blank password clears the stored encrypted value
            s["smtp_password_enc"] = ""
        self._save_raw(raw_data)

    _ALL_DOCS = ["26as_pdf", "26as_xlsx", "168_pdf", "168_xlsx", "ais_pdf", "ais_xlsx", "tis_pdf",
                 "itr_form", "itr_receipt", "itr_v", "intimation"]

    def _make_legacy_template(self, cfg: dict) -> dict:
        return {
            "name":    "Legacy Template",
            "subject": cfg.get("email_subject_tpl", self._EMAIL_DEFAULTS["email_subject_tpl"]),
            "body":    cfg.get("email_body_tpl",    self._EMAIL_DEFAULTS["email_body_tpl"]),
            "docs":    {k: True for k in self._ALL_DOCS},
        }

    def get_email_templates(self) -> list:
        """Return list of template dicts. Auto-creates 'Legacy Template' from existing settings if none exist."""
        raw_data = self._get_raw()
        s = raw_data.get("settings", {})
        templates = s.get("email_templates")
        if not templates:
            # First run — migrate existing subject/body into Legacy Template
            cfg = self.get_email_settings()
            templates = [self._make_legacy_template(cfg)]
            s = raw_data.setdefault("settings", {})
            s["email_templates"]  = templates
            s["active_template"]  = "Legacy Template"
            self._save_raw(raw_data)
        return templates

    def save_email_templates(self, templates: list):
        raw_data = self._get_raw()
        raw_data.setdefault("settings", {})["email_templates"] = templates
        self._save_raw(raw_data)

    def get_active_template_name(self) -> str:
        s = self._get_raw().get("settings", {})
        templates = s.get("email_templates", [])
        name = s.get("active_template", "")
        # Fall back to first template if stored name no longer exists
        if templates and not any(t["name"] == name for t in templates):
            name = templates[0]["name"]
        return name

    def set_active_template(self, name: str):
        raw_data = self._get_raw()
        raw_data.setdefault("settings", {})["active_template"] = name
        self._save_raw(raw_data)

    def get_active_template(self) -> dict | None:
        templates = self.get_email_templates()
        name = self.get_active_template_name()
        for t in templates:
            if t["name"] == name:
                return t
        return templates[0] if templates else None
