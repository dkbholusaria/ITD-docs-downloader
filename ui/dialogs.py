import os, json, datetime, threading
from PyQt6.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
    QCheckBox, QRadioButton, QLineEdit, QScrollArea, QWidget,
    QFrame, QTableWidget, QTableWidgetItem, QHeaderView,
    QAbstractItemView, QProgressBar, QSizePolicy, QFileDialog, QMessageBox,
    QTextEdit, QSpinBox, QComboBox, QFontComboBox, QTabWidget,
)
from PyQt6.QtCore import Qt, pyqtSignal
from PyQt6.QtGui import (
    QFont, QColor, QPixmap, QPainter, QBrush, QPen,
    QTextCharFormat, QIcon,
)

from ui._theme import _t
from ui.helpers import _btn, _lbl, _status_style, _UI_FONT, _icon_path
from automation.doc_types import match_doc_type
from config import _open_path, _log_open, _app_dir
from themes import MONO_FONT_NAME as _MONO_FONT


# ── Manage Years Dialog ───────────────────────────────────────────────────────

class ManageYearsDialog(QDialog):
    def __init__(self, parent, json_path: str, on_save):
        super().__init__(parent)
        self.setWindowTitle("Manage Assessment / Tax Years")
        self.setFixedSize(500, 560)
        self.setModal(True)
        self._json_path = json_path
        self._on_save = on_save
        self._checkboxes = []

        try:
            with open(json_path, "r", encoding="utf-8") as f:
                entries = json.load(f)
        except Exception:
            entries = []

        self._build_ui(entries)

    def _build_ui(self, entries):
        t = _t()
        self.setStyleSheet(
            f"QDialog{{background:{t.bg_window};}}"
            f"QLabel{{color:{t.text_primary};background:transparent;}}"
            f"QRadioButton{{color:{t.text_primary};background:transparent;spacing:6px;}}"
            f"QRadioButton::indicator{{width:14px;height:14px;border:1.5px solid {t.border};"
            f"border-radius:7px;background:{t.bg_checkbox};}}"
            f"QRadioButton::indicator:checked{{background:{t.accent};border-color:{t.accent};}}"
            f"QScrollArea{{background:{t.bg_input};border:1px solid {t.border};border-radius:6px;}}"
            f"QScrollArea > QWidget > QWidget{{background:{t.bg_input};}}"
        )
        main = QVBoxLayout(self)
        main.setContentsMargins(20, 16, 20, 16)
        main.setSpacing(8)

        # Header: logo left, title+subtitle stacked right
        from config import _bundled_dir
        hdr_row = QHBoxLayout()
        hdr_row.setSpacing(12)
        hdr_row.setAlignment(Qt.AlignmentFlag.AlignVCenter)
        _icon_lbl = QLabel()
        _icon_lbl.setStyleSheet("background:transparent;")
        _ip = os.path.join(_bundled_dir(), "resources", "app_icon.png")
        if os.path.isfile(_ip):
            _icon_lbl.setPixmap(QPixmap(_ip).scaled(48, 48, Qt.AspectRatioMode.KeepAspectRatio,
                                                     Qt.TransformationMode.SmoothTransformation))
        hdr_row.addWidget(_icon_lbl)
        txt_col = QVBoxLayout()
        txt_col.setSpacing(2)
        txt_col.addWidget(_lbl("Manage Assessment / Tax Years", 13, bold=True))
        txt_col.addWidget(_lbl("Toggle enabled/disabled or add new years.", 10, color=t.text_muted))
        hdr_row.addLayout(txt_col)
        hdr_row.addStretch()
        main.addLayout(hdr_row)
        main.addWidget(_lbl("Existing Entries", 11, bold=True, color=t.text_muted))

        scroll = QScrollArea()
        scroll.setFixedHeight(180)
        scroll.setWidgetResizable(True)
        inner = QWidget()
        inner.setAttribute(Qt.WidgetAttribute.WA_StyledBackground, True)
        inner.setStyleSheet(f"QWidget{{background:{t.bg_input};}}")
        self._list_layout = QVBoxLayout(inner)
        self._list_layout.setSpacing(2)
        self._list_layout.addStretch()
        scroll.setWidget(inner)
        main.addWidget(scroll)

        for e in entries:
            self._add_row(e)

        sep = QFrame(); sep.setFrameShape(QFrame.Shape.HLine)
        sep.setStyleSheet(f"background:{t.border};border:none;max-height:1px;")
        main.addWidget(sep)

        main.addWidget(_lbl("Add New Year", 11, bold=True, color=t.text_muted))

        type_row = QHBoxLayout()
        type_row.addWidget(_lbl("Type:", 11))
        self._type_ay = QRadioButton("AY (Assessment Year)"); self._type_ay.setChecked(True)
        self._type_ty = QRadioButton("TY (Tax Year)")
        self._type_ay.toggled.connect(self._auto_fy)
        type_row.addWidget(self._type_ay); type_row.addWidget(self._type_ty)
        type_row.addStretch(); main.addLayout(type_row)

        yr_row = QHBoxLayout()
        yr_row.addWidget(_lbl("Year:", 11))
        self._year_edit = QLineEdit(); self._year_edit.setPlaceholderText("e.g. 2027-28")
        self._year_edit.setFixedWidth(120); self._year_edit.textChanged.connect(self._auto_fy)
        yr_row.addWidget(self._year_edit)
        yr_row.addWidget(_lbl("FY:", 11))
        self._fy_edit = QLineEdit(); self._fy_edit.setPlaceholderText("auto-filled")
        self._fy_edit.setFixedWidth(120)
        yr_row.addWidget(self._fy_edit)
        yr_row.addWidget(_lbl("(editable)", 10, color=t.text_muted))
        yr_row.addStretch(); main.addLayout(yr_row)

        add_btn = _btn("Add to List", "outline", height=32, min_width=130, icon="btn_add_list.png")
        add_btn.clicked.connect(self._add_entry)
        main.addWidget(add_btn, alignment=Qt.AlignmentFlag.AlignLeft)

        sep2 = QFrame(); sep2.setFrameShape(QFrame.Shape.HLine)
        sep2.setStyleSheet(f"background:{t.border};border:none;max-height:1px;")
        main.addWidget(sep2)

        btns_row = QHBoxLayout()
        save_btn = _btn("Save & Close", "primary", height=36, icon="btn_save_close.png")
        save_btn.clicked.connect(self._save)
        cancel_btn = _btn("Cancel", "secondary", height=36, icon="btn_cancel.png")
        cancel_btn.clicked.connect(self.reject)
        btns_row.addWidget(save_btn); btns_row.addWidget(cancel_btn)
        main.addLayout(btns_row)

    def _add_row(self, entry):
        t = _t()
        cb = QCheckBox(entry["label"])
        cb.setChecked(entry.get("enabled", True))
        cb.setStyleSheet(
            f"QCheckBox{{color:{t.text_primary};background:transparent;spacing:6px;}}"
            f"QCheckBox::indicator{{width:15px;height:15px;border:1.5px solid {t.border};"
            f"border-radius:3px;background:{t.bg_checkbox};}}"
            f"QCheckBox::indicator:checked{{background:{t.accent};border-color:{t.accent};}}"
        )
        self._checkboxes.append((entry, cb))
        self._list_layout.insertWidget(self._list_layout.count() - 1, cb)

    def _auto_fy(self):
        import re
        year = self._year_edit.text().strip()
        m = re.match(r"(\d{4})-(\d{2}|\d{4})$", year)
        if not m:
            return
        y1 = int(m.group(1)); suffix = m.group(2)
        y2 = int(str(y1)[:2] + suffix) if len(suffix) == 2 else int(suffix)
        fy = f"{y1}-{str(y2)[-2:]}" if self._type_ty.isChecked() else f"{y1-1}-{str(y2-1)[-2:]}"
        self._fy_edit.setText(fy)

    def _add_entry(self):
        year_type = "TY" if self._type_ty.isChecked() else "AY"
        year = self._year_edit.text().strip()
        fy = self._fy_edit.text().strip()
        if not year or not fy:
            QMessageBox.warning(self, "Missing Fields", "Please fill in Year and FY.")
            return
        label = f"{year_type} {year} (FY {fy})"
        if any(e["label"] == label for e, _ in self._checkboxes):
            QMessageBox.warning(self, "Duplicate", f'"{label}" already exists.')
            return
        year_obj = {"TY": year, "FY": fy} if year_type == "TY" else {"AY": year, "FY": fy}
        self._add_row({"label": label, "enabled": True, "year": year_obj})
        self._year_edit.clear(); self._fy_edit.clear()

    def _save(self):
        final = [{**e, "enabled": cb.isChecked()} for e, cb in self._checkboxes]
        def _sort_key(e):
            y = e.get("year", {})
            label_year = y.get("AY") or y.get("TY") or y.get("FY") or "0000-00"
            try:
                return (0 if not e.get("enabled", True) else 1, -int(label_year[:4]))
            except ValueError:
                return (1, 0)
        final.sort(key=_sort_key)
        try:
            with open(self._json_path, "w", encoding="utf-8") as f:
                json.dump(final, f, indent=2, ensure_ascii=False)
            self._on_save()
            self.accept()
        except Exception as ex:
            QMessageBox.critical(self, "Save Error", str(ex))


# ── Manage Groups Dialog (F-11) ────────────────────────────────────────────────

class ManageGroupsDialog(QDialog):
    """Client Master > Manage Groups… — two decks, confirmed live per the
    user's own description: an upper deck listing every group (the
    "mother"), and a lower deck showing/editing which clients currently
    belong to whichever group is selected above (the "children") —
    replacing the earlier rename/delete-only list, which had no way to
    actually see or change a group's membership at all.

    Changes apply immediately to the vault (rename_group/clear_group/
    set_client_group) rather than being staged and written on a single
    Save & Close, since a group is just a shared value of each client's
    own "group" field, not a separate managed list needing its own save
    step."""

    def __init__(self, parent, vault):
        super().__init__(parent)
        self._vault = vault
        self._selected_group: str = ""
        self.setWindowTitle("Manage Groups")
        self.setMinimumSize(520, 620)
        self.resize(560, 700)
        self.setSizeGripEnabled(True)
        self.setModal(True)
        self._build_ui()
        self._refresh_groups()

    def _build_ui(self):
        t = _t()
        self.setStyleSheet(
            f"QDialog{{background:{t.bg_window};}}"
            f"QLabel{{color:{t.text_primary};background:transparent;}}")
        main = QVBoxLayout(self)
        main.setContentsMargins(20, 16, 20, 16)
        main.setSpacing(8)

        main.addWidget(_lbl("Manage Groups", 13, bold=True))
        main.addWidget(_lbl(
            "Select a group below to see and edit which clients belong to it. "
            "Deleting a group only un-groups its clients — it never deletes them.",
            10, color=t.text_muted))

        # ── Upper deck: Groups ───────────────────────────────────────────
        # A real QListWidget (not a manually-stacked QVBoxLayout of rows)
        # so Up/Down arrow keys move between groups for free — confirmed
        # live that the earlier plain-widget-row version had no keyboard
        # navigation at all. Rename/Delete stay as per-row buttons via
        # setItemWidget(); the list's own selection visuals are suppressed
        # (background: transparent) in favor of the same manual
        # accent-highlight the row widget already painted before, so the
        # look is unchanged — only the navigation mechanics are new.
        main.addWidget(_lbl("Groups", 11, bold=True, color=t.text_muted))
        from PyQt6.QtWidgets import QListWidget, QListWidgetItem
        from PyQt6.QtCore import QSize
        self._groups_list = QListWidget()
        self._groups_list.setFixedHeight(200)
        self._groups_list.setFocusPolicy(Qt.FocusPolicy.StrongFocus)
        self._groups_list.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
        self._groups_list.setStyleSheet(
            f"QListWidget{{background:{t.bg_input};border:1px solid {t.border};"
            f"border-radius:6px;outline:0;padding:4px;}}"
            f"QListWidget::item{{border:none;padding:0;margin:2px 0;}}"
            f"QListWidget::item:selected{{background:transparent;}}"
            f"QListWidget::item:hover{{background:transparent;}}")
        self._groups_list.currentItemChanged.connect(self._on_group_selection_changed)
        main.addWidget(self._groups_list)

        # Lets a group be created ahead of assigning any client to it —
        # otherwise the only way to create one was from inside a client's
        # own Group field, and this list would stay empty until the first
        # client was actually assigned.
        add_row = QHBoxLayout()
        add_row.setSpacing(8)
        self._new_group_edit = QLineEdit()
        self._new_group_edit.setPlaceholderText("New group name, e.g. Bholusaria Family")
        self._new_group_edit.setFixedHeight(32)
        self._new_group_edit.returnPressed.connect(self._add_group)
        add_row.addWidget(self._new_group_edit, 1)
        add_group_btn = _btn("Add Group", "outline", height=32)
        add_group_btn.clicked.connect(self._add_group)
        add_row.addWidget(add_group_btn)
        main.addLayout(add_row)

        sep = QFrame(); sep.setFrameShape(QFrame.Shape.HLine)
        sep.setStyleSheet(f"background:{t.border};border:none;max-height:1px;")
        main.addWidget(sep)

        # ── Lower deck: Clients in the selected group ────────────────────
        self._clients_header = _lbl("Clients", 11, bold=True, color=t.text_muted)
        main.addWidget(self._clients_header)

        self._client_search = QLineEdit()
        self._client_search.setPlaceholderText("🔍  Search to add a client to this group...")
        self._client_search.setClearButtonEnabled(True)
        self._client_search.setFixedHeight(28)
        self._client_search.textChanged.connect(self._refresh_clients)
        main.addWidget(self._client_search)

        clients_scroll = QScrollArea()
        clients_scroll.setWidgetResizable(True)
        clients_scroll.setStyleSheet(
            f"QScrollArea{{background:{t.bg_input};border:1px solid {t.border};border-radius:6px;}}"
            f"QScrollArea > QWidget > QWidget{{background:{t.bg_input};}}")
        clients_inner = QWidget()
        clients_inner.setAttribute(Qt.WidgetAttribute.WA_StyledBackground, True)
        clients_inner.setStyleSheet(f"QWidget{{background:{t.bg_input};}}")
        self._clients_layout = QVBoxLayout(clients_inner)
        self._clients_layout.setSpacing(2)
        self._clients_layout.addStretch()
        clients_scroll.setWidget(clients_inner)
        main.addWidget(clients_scroll, 1)

        btns_row = QHBoxLayout()
        btns_row.addStretch()
        close_btn = _btn("Close", "primary", height=36)
        close_btn.clicked.connect(self.accept)
        btns_row.addWidget(close_btn)
        main.addLayout(btns_row)

    # ── Upper deck: Groups ────────────────────────────────────────────────

    def _add_group(self):
        name = self._new_group_edit.text().strip()
        if not name:
            return
        if not self._vault.add_group(name):
            QMessageBox.warning(self, "Group Exists", f'"{name}" already exists.')
            return
        self._new_group_edit.clear()
        self._selected_group = name  # jump straight to assigning clients to it
        self._refresh_groups()
        self._refresh_clients()

    def _on_group_selection_changed(self, current, _previous):
        """Fires on a click OR an arrow-key move (QListWidget handles both
        natively) — never re-entrantly from _refresh_groups() itself,
        since that rebuild runs with signals blocked."""
        if current is None:
            return
        group = current.data(Qt.ItemDataRole.UserRole)
        if group == self._selected_group:
            return
        self._selected_group = group
        self._refresh_groups()   # redraw highlight
        self._refresh_clients()

    def _refresh_groups(self):
        from PyQt6.QtWidgets import QListWidgetItem
        from PyQt6.QtCore import QSize

        groups = self._vault.get_all_groups()
        if self._selected_group and self._selected_group not in groups:
            self._selected_group = ""  # e.g. the selected group was just deleted

        counts = {}
        for a in self._vault.get_all_assessees():
            g = a.get("group", "").strip()
            if g:
                counts[g] = counts.get(g, 0) + 1

        self._groups_list.blockSignals(True)
        self._groups_list.clear()

        if not groups:
            t = _t()
            item = QListWidgetItem(
                'No groups yet — type a name below and click "Add Group" to create one.')
            item.setFlags(Qt.ItemFlag.NoItemFlags)
            item.setForeground(QColor(t.text_muted))
            self._groups_list.addItem(item)
            self._groups_list.blockSignals(False)
            return

        selected_row = -1
        for i, g in enumerate(groups):
            item = QListWidgetItem()
            item.setData(Qt.ItemDataRole.UserRole, g)
            item.setSizeHint(QSize(0, 40))
            self._groups_list.addItem(item)
            self._groups_list.setItemWidget(item, self._make_group_row_widget(g, counts.get(g, 0)))
            if g == self._selected_group:
                selected_row = i
        if selected_row >= 0:
            self._groups_list.setCurrentRow(selected_row)
        self._groups_list.blockSignals(False)

    def _make_group_row_widget(self, group: str, count: int) -> QWidget:
        t = _t()
        is_selected = group == self._selected_group
        row = QWidget()
        row.setAttribute(Qt.WidgetAttribute.WA_StyledBackground, True)
        row.setStyleSheet(
            f"QWidget{{background:{t.accent if is_selected else 'transparent'};border-radius:4px;}}")
        hl = QHBoxLayout(row)
        hl.setContentsMargins(8, 2, 4, 2)
        hl.setSpacing(8)

        plural = "s" if count != 1 else ""
        lbl = QLabel(f"{group}   ({count} client{plural})")
        lbl.setStyleSheet(
            f"background:transparent;font-size:12px;"
            f"font-weight:{'700' if is_selected else '400'};"
            f"color:{t.accent_text if is_selected else t.text_primary};")
        hl.addWidget(lbl, 1)

        rename_btn = _btn("Rename", "outline", height=28)
        rename_btn.clicked.connect(lambda _, g=group: self._rename(g))
        hl.addWidget(rename_btn)
        delete_btn = _btn("Delete", "danger", height=28)
        delete_btn.clicked.connect(lambda _, g=group: self._delete(g))
        hl.addWidget(delete_btn)
        return row

    def _rename(self, old: str):
        from PyQt6.QtWidgets import QInputDialog
        new, ok = QInputDialog.getText(self, "Rename Group", f'New name for "{old}":', text=old)
        if not ok:
            return
        new = new.strip()
        if not new or new == old:
            return
        if new in self._vault.get_all_groups():
            if QMessageBox.question(
                    self, "Merge Groups",
                    f'"{new}" already exists. Merge "{old}" into it?',
                    QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
               ) != QMessageBox.StandardButton.Yes:
                return
        self._vault.rename_group(old, new)
        if self._selected_group == old:
            self._selected_group = new
        self._refresh_groups()
        self._refresh_clients()

    def _delete(self, group: str):
        count = sum(1 for a in self._vault.get_all_assessees()
                    if a.get("group", "").strip() == group)
        if QMessageBox.question(
                self, "Delete Group",
                f'Delete group "{group}"? {count} client(s) will be un-grouped, not deleted.',
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
           ) != QMessageBox.StandardButton.Yes:
            return
        self._vault.clear_group(group)
        if self._selected_group == group:
            self._selected_group = ""
        self._refresh_groups()
        self._refresh_clients()

    # ── Lower deck: Clients in the selected group ─────────────────────────

    def _clear_client_rows(self):
        while self._clients_layout.count() > 1:
            item = self._clients_layout.takeAt(0)
            w = item.widget()
            if w:
                w.deleteLater()

    def _refresh_clients(self, *_args):
        """Confirmed live, corrected: clicking a group must show ONLY that
        group's actual members by default — the earlier "every client,
        checked or not" list read as broken (a brand-new empty group
        looked identical to one with members, and it was hard to tell at
        a glance who was actually in a group). Search now has a distinct
        job: finding a client to ADD (from everyone, member or not) rather
        than filtering the always-everyone list."""
        self._clear_client_rows()
        t = _t()
        if not self._selected_group:
            self._clients_header.setText("Clients")
            placeholder = _lbl(
                "Select a group above to see and edit its clients.", 11, color=t.text_muted)
            placeholder.setWordWrap(True)
            self._clients_layout.insertWidget(0, placeholder)
            return

        self._clients_header.setText(f'Clients in "{self._selected_group}"')
        search = self._client_search.text().strip().lower()
        all_clients = self._vault.get_all_assessees()

        if search:
            # Search mode: candidates from EVERYONE matching the search,
            # each row showing Remove (already a member) or Add (isn't).
            clients = [c for c in all_clients
                       if search in c.get("name", "").lower() or search in c.get("pan", "").lower()]
            clients.sort(key=lambda c: c.get("name", "").lower())
            if not clients:
                empty_lbl = _lbl("No clients match your search.", 11, color=t.text_muted)
                self._clients_layout.insertWidget(0, empty_lbl)
                return
            for c in clients:
                self._add_client_row(c, in_group=(c.get("group", "").strip() == self._selected_group))
            return

        # Default: members of this group only.
        members = [c for c in all_clients if c.get("group", "").strip() == self._selected_group]
        members.sort(key=lambda c: c.get("name", "").lower())
        if not members:
            empty_lbl = _lbl(
                "No clients in this group yet — search above to add some.",
                11, color=t.text_muted)
            empty_lbl.setWordWrap(True)
            self._clients_layout.insertWidget(0, empty_lbl)
            return
        for c in members:
            self._add_client_row(c, in_group=True)

    def _add_client_row(self, client: dict, in_group: bool):
        t = _t()
        row = QWidget()
        row.setAttribute(Qt.WidgetAttribute.WA_StyledBackground, True)
        row.setStyleSheet(f"QWidget{{background:{t.bg_input};}}")
        hl = QHBoxLayout(row)
        hl.setContentsMargins(4, 2, 4, 2)
        hl.setSpacing(8)

        lbl = QLabel(f"{client.get('name', '')}   ({client.get('pan', '')})")
        hl.addWidget(lbl, 1)
        if in_group:
            action_btn = _btn("✕ Remove", "danger", height=26)
        else:
            action_btn = _btn("+ Add", "outline", height=26)
        action_btn.clicked.connect(
            lambda _, cid=client.get("id"), add=not in_group: self._set_membership(cid, add))
        hl.addWidget(action_btn)
        self._clients_layout.insertWidget(self._clients_layout.count() - 1, row)

    def _set_membership(self, client_id: str, add: bool):
        self._vault.set_client_group(client_id, self._selected_group if add else "")
        self._refresh_groups()
        self._refresh_clients()


# ── Batch Progress Dialog ─────────────────────────────────────────────────────

class BatchProgressDialog(QDialog):
    """
    Live progress popup shown during any batch run.
    Columns: Name | PAN | Status | Save Path (clickable link)
    Status and path updates arrive from the worker thread via Qt signals.

    F-14 (multi-year): one row per (client, year) pair — every client
    row is repeated once per selected Assessment/Tax Year, so `year_specs`
    (list of {"ay_label": ..., ...} dicts, see batch_handlers.py) drives how
    many rows each client gets. All row-keyed state uses the composite key
    (pan, ay_label) instead of pan alone.
    """
    _update_signal = pyqtSignal(str, str, str)         # pan, ay_label, status
    _path_signal   = pyqtSignal(str, str, str)         # pan, ay_label, folder
    _resume_signal = pyqtSignal(list)
    _client_done_signal = pyqtSignal(str, str)   # pan, ay_label — ALL selected doc types finished for this (client, year)

    _COL_NAME   = 0
    _COL_YEAR   = 1
    _COL_PAN    = 2
    _COL_STATUS = 3
    _COL_PATH   = 4

    def __init__(self, targets: list, selected_docs, year_specs: list,
                 ay: str = "", stop_callback=None, resume_callback=None, skip_callback=None,
                 tray_callback=None, output_dir: str = "", parent=None):
        super().__init__(parent)
        self._stop_callback   = stop_callback
        self._resume_callback = resume_callback
        self._skip_callback   = skip_callback
        self._tray_callback   = tray_callback
        self._output_dir      = output_dir
        self._selected_docs   = selected_docs
        self._ay              = ay
        self._targets         = targets
        self._year_specs      = year_specs
        self._multi_year      = len(year_specs) > 1
        # Client × Year, client-major order — matches the execution order in
        # app.py's _execute_batch (all doc types/years for client A, then B).
        self._client_year_pairs = [(tgt, spec) for tgt in targets for spec in year_specs]
        self._path_by_key    = {}   # (pan, ay_label) → folder

        _doc_labels = {
            "26as":           "26AS",
            "request_ais":    "Requesting AIS Generation",
            "ais_tis":        "AIS / TIS",
            "filed_returns":  "Filed Returns",
            "challans":       "Tax Challans",
        }
        mode_label = " + ".join(
            _doc_labels.get(d, d) for d in sorted(selected_docs)
        ) or "Batch Run"
        mode_label = f"Downloading {mode_label}" if mode_label != "Batch Run" else mode_label
        self._mode_label = mode_label

        self.setWindowTitle(f"{mode_label} — Batch Progress")
        self.setMinimumSize(900, 500)
        self.resize(1060, min(160 + len(self._client_year_pairs) * 42, 720))
        self.setSizeGripEnabled(True)
        self.setWindowFlags(
            Qt.WindowType.Dialog |
            Qt.WindowType.WindowTitleHint |
            Qt.WindowType.WindowCloseButtonHint |
            Qt.WindowType.WindowMaximizeButtonHint)
        _bt = _t()
        self.setStyleSheet(f"QDialog{{background:{_bt.bg_window};}}")

        self._row_key_to_row = {}   # (pan, ay_label) → row
        self._counted_keys: set = set()   # (pan, ay_label) already counted toward done
        self._done_count   = 0
        self._total        = len(self._client_year_pairs)

        layout = QVBoxLayout(self)
        layout.setContentsMargins(16, 14, 16, 12)
        layout.setSpacing(8)

        # ── Title bar ─────────────────────────────────────────────────────────
        ay_tag = (f" &nbsp;·&nbsp; <span style='color:{_bt.accent}'>{ay}</span>") if ay else ""
        clients_tag = (f"{len(targets)} client(s)" if not self._multi_year
                       else f"{len(targets)} client(s) × {len(year_specs)} year(s)")
        title = QLabel(f"<b>{mode_label}</b> — {clients_tag}{ay_tag}")
        title.setStyleSheet(f"font-size:14px; color:{_bt.text_primary}; background:transparent;")
        layout.addWidget(title)

        # ── Table ─────────────────────────────────────────────────────────────
        self._table = QTableWidget(len(self._client_year_pairs), 5)
        self._table.setHorizontalHeaderLabels(["Name", "Year", "PAN", "Status", "Save Path"])
        if not self._multi_year:
            self._table.setColumnHidden(self._COL_YEAR, True)

        hdr = self._table.horizontalHeader()
        hdr.setSectionResizeMode(self._COL_NAME,   QHeaderView.ResizeMode.Interactive)
        hdr.setSectionResizeMode(self._COL_YEAR,   QHeaderView.ResizeMode.Interactive)
        hdr.setSectionResizeMode(self._COL_PAN,    QHeaderView.ResizeMode.Interactive)
        hdr.setSectionResizeMode(self._COL_STATUS, QHeaderView.ResizeMode.Interactive)
        hdr.setSectionResizeMode(self._COL_PATH,   QHeaderView.ResizeMode.Stretch)
        self._table.setColumnWidth(self._COL_NAME,   180)
        self._table.setColumnWidth(self._COL_YEAR,   150)
        self._table.setColumnWidth(self._COL_PAN,    120)
        self._table.setColumnWidth(self._COL_STATUS, 260)

        hdr.setStyleSheet(
            f"QHeaderView::section{{"
            f"background-color:{_bt.bg_header};"
            f"border:none;"
            f"border-right:1px solid {_bt.border};"
            f"border-bottom:1px solid {_bt.border};"
            f"font-weight:bold;color:{_bt.text_muted};"
            f"font-size:11px;height:34px;"
            f"padding:0 8px;}}")
        self._table.verticalHeader().setVisible(False)
        self._table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self._table.setSelectionMode(QAbstractItemView.SelectionMode.NoSelection)
        self._table.setShowGrid(True)
        self._table.setAlternatingRowColors(False)
        self._table.setWordWrap(False)
        self._table.setStyleSheet(
            f"QTableWidget{{border:1.5px solid {_bt.border};border-radius:8px;"
            f"background:{_bt.bg_table};outline:0;gridline-color:{_bt.grid};}}"
            f"QTableWidget::item{{border-bottom:1px solid {_bt.grid};padding:0 8px;}}"
            f"QPushButton{{border:none;background:transparent;font-size:14px;}}"
            f"QPushButton:hover{{background:{_bt.bg_table_alt};border-radius:4px;}}")
        self._table.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)

        for row, (tgt, spec) in enumerate(self._client_year_pairs):
            pan  = tgt.get("pan", "")
            name = tgt.get("name", "—")
            ay_label = spec["ay_label"]
            self._row_key_to_row[(pan, ay_label)] = row
            self._table.setRowHeight(row, 40)

            name_item = QTableWidgetItem(name)
            name_item.setForeground(QColor(_bt.text_primary))
            self._table.setItem(row, self._COL_NAME, name_item)

            year_item = QTableWidgetItem(ay_label)
            year_item.setForeground(QColor(_bt.text_muted))
            year_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter | Qt.AlignmentFlag.AlignVCenter)
            self._table.setItem(row, self._COL_YEAR, year_item)

            pan_item = QTableWidgetItem(pan)
            pan_item.setForeground(QColor(_bt.text_muted))
            pan_item.setFont(QFont(_MONO_FONT, 10))
            pan_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter | Qt.AlignmentFlag.AlignVCenter)
            self._table.setItem(row, self._COL_PAN, pan_item)

            self._set_status_item(row, "⬜ Waiting")

            path_lbl = QLabel("—")
            path_lbl.setStyleSheet(
                f"color:{_bt.text_muted};font-size:11px;padding:0 8px;background:transparent;")
            path_lbl.setAlignment(Qt.AlignmentFlag.AlignVCenter | Qt.AlignmentFlag.AlignLeft)
            path_lbl.setWordWrap(False)
            path_lbl.setTextInteractionFlags(Qt.TextInteractionFlag.TextBrowserInteraction)
            path_lbl.setOpenExternalLinks(False)
            path_lbl.linkActivated.connect(self._open_row_path)
            self._table.setCellWidget(row, self._COL_PATH, path_lbl)

        layout.addWidget(self._table, stretch=1)

        # ── Progress bar ──────────────────────────────────────────────────────
        self._progress_bar = QProgressBar()
        self._progress_bar.setRange(0, len(self._client_year_pairs))
        self._progress_bar.setValue(0)
        self._progress_bar.setFixedHeight(18)
        self._progress_bar.setTextVisible(True)
        self._progress_bar.setFormat(f"0 / {len(self._client_year_pairs)} done")
        self._progress_bar.setStyleSheet(
            f"QProgressBar{{border:1px solid {_bt.border};border-radius:9px;"
            f"background:{_bt.scrollbar_handle};text-align:center;font-size:11px;"
            f"font-weight:600;color:{_bt.accent_text};}}"
            f"QProgressBar::chunk{{background:#16A34A;border-radius:9px;}}")
        layout.addWidget(self._progress_bar)

        # ── Footer ────────────────────────────────────────────────────────────
        footer = QHBoxLayout()
        footer.setSpacing(8)
        footer.setContentsMargins(0, 0, 0, 0)

        loc_cap = QLabel("📁")
        loc_cap.setStyleSheet("font-size:13px;background:transparent;")
        loc_cap.setFixedWidth(18)
        footer.addWidget(loc_cap)

        self._loc_val = QLabel(output_dir or "—")
        self._loc_val.setStyleSheet(f"color:{_bt.text_muted};font-size:11px;background:transparent;")
        self._loc_val.setWordWrap(False)
        self._loc_val.setMinimumWidth(0)
        self._loc_val.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Preferred)
        self._loc_val.setTextInteractionFlags(Qt.TextInteractionFlag.TextSelectableByMouse)
        footer.addWidget(self._loc_val, stretch=1)

        self._open_folder_btn = QPushButton("📂  Open Folder")
        self._open_folder_btn.setFixedHeight(32)
        self._open_folder_btn.setStyleSheet(
            f"QPushButton{{background:{_bt.bg_table_alt};color:{_bt.text_primary};border:1px solid {_bt.border};"
            f"border-radius:6px;font-size:12px;padding:0 12px;}}"
            f"QPushButton:hover{{background:{_bt.bg_input};}}"
            f"QPushButton:disabled{{color:{_bt.text_muted};border-color:{_bt.border};}}")
        self._open_folder_btn.clicked.connect(self._open_output_dir)
        footer.addWidget(self._open_folder_btn)

        self._report_btn = QPushButton("📂  Open Report")
        self._report_btn.setFixedHeight(32)
        self._report_btn.setEnabled(False)
        self._report_btn.setStyleSheet(
            f"QPushButton{{background:{_bt.bg_table_alt};color:{_bt.text_primary};border:1px solid {_bt.border};"
            f"border-radius:6px;font-size:12px;padding:0 12px;}}"
            f"QPushButton:enabled:hover{{background:{_bt.bg_input};}}"
            f"QPushButton:disabled{{color:{_bt.text_muted};border-color:{_bt.border};}}")
        self._report_btn.clicked.connect(self._open_last_report)
        self._last_report_path = ""
        footer.addWidget(self._report_btn)

        self._skip_btn = QPushButton("⏭  Skip")
        self._skip_btn.setFixedHeight(32)
        self._skip_btn.setMinimumWidth(80)
        self._skip_btn.setToolTip("Skip the currently-downloading client and move to the next")
        self._skip_btn.setStyleSheet(
            f"QPushButton{{background:{_bt.bg_table_alt};color:{_bt.text_primary};border:1px solid {_bt.border};"
            f"border-radius:6px;font-size:12px;font-weight:600;padding:0 12px;}}"
            f"QPushButton:hover{{background:{_bt.bg_input};}}"
            f"QPushButton:disabled{{color:{_bt.text_muted};border-color:{_bt.border};}}")
        self._skip_btn.clicked.connect(self._on_skip_clicked)
        footer.addWidget(self._skip_btn)

        self._tray_btn = QPushButton("⬇  Tray")
        self._tray_btn.setFixedHeight(32)
        self._tray_btn.setMinimumWidth(80)
        self._tray_btn.setToolTip("Hide to system tray — click the tray icon to restore")
        self._tray_btn.setStyleSheet(
            f"QPushButton{{background:{_bt.bg_table_alt};color:{_bt.text_primary};border:1px solid {_bt.border};"
            f"border-radius:6px;font-size:12px;font-weight:600;padding:0 12px;}}"
            f"QPushButton:hover{{background:{_bt.bg_input};}}"
            f"QPushButton:disabled{{color:{_bt.text_muted};border-color:{_bt.border};}}")
        self._tray_btn.setVisible(bool(self._tray_callback))
        self._tray_btn.clicked.connect(self._on_tray_clicked)
        footer.addWidget(self._tray_btn)

        self._stop_btn = QPushButton("⏹  Stop")
        self._stop_btn.setFixedHeight(32)
        self._stop_btn.setMinimumWidth(90)
        self._stop_btn.setStyleSheet(
            "QPushButton{background:#EF4444;color:#FFFFFF;border:none;"
            "border-radius:6px;font-size:12px;font-weight:600;padding:0 12px;}"
            "QPushButton:hover{background:#DC2626;}"
            "QPushButton:disabled{background:#E2E8F0;color:#94A3B8;}")
        self._stop_btn.clicked.connect(self._on_stop_clicked)
        footer.addWidget(self._stop_btn)

        self._resume_btn = QPushButton("▶  Resume")
        self._resume_btn.setFixedHeight(32)
        self._resume_btn.setMinimumWidth(100)
        self._resume_btn.setVisible(False)
        self._resume_btn.setStyleSheet(
            "QPushButton{background:#16A34A;color:#FFFFFF;border:none;"
            "border-radius:6px;font-size:12px;font-weight:600;padding:0 12px;}"
            "QPushButton:hover{background:#15803D;}")
        self._resume_btn.clicked.connect(self._on_resume_clicked)
        footer.addWidget(self._resume_btn)

        self._close_btn = QPushButton("Close")
        self._close_btn.setFixedSize(80, 32)
        self._close_btn.setEnabled(False)
        self._close_btn.setStyleSheet(
            f"QPushButton{{background:{_bt.border};color:{_bt.text_muted};border:none;"
            f"border-radius:6px;font-size:12px;}}"
            f"QPushButton:enabled{{background:{_bt.accent};color:{_bt.accent_text};}}"
            f"QPushButton:enabled:hover{{background:{_bt.accent_hover};}}")
        self._close_btn.clicked.connect(self.accept)
        footer.addWidget(self._close_btn)

        layout.addLayout(footer)

        self._rows_data  = {}

        for tgt, spec in self._client_year_pairs:
            self._rows_data[(tgt.get("pan", ""), spec["ay_label"])] = {
                "name": tgt.get("name", ""), "ay_label": spec["ay_label"],
                "path": "", "status": "Waiting", "ts": ""}

        self._update_signal.connect(self._on_update)
        self._path_signal.connect(self._on_path_update)
        self._client_done_signal.connect(self._on_client_done)
        self._table.cellDoubleClicked.connect(self._on_row_double_clicked)

    # ── internal helpers ──────────────────────────────────────────────────────

    def _row_key_for_row(self, row: int):
        for key, r in self._row_key_to_row.items():
            if r == row:
                return key
        return None

    def _on_row_double_clicked(self, row: int, _col: int):
        """Show full status text for the double-clicked row."""
        status_item = self._table.item(row, self._COL_STATUS)
        name_item   = self._table.item(row, self._COL_NAME)
        if not status_item:
            return
        status = status_item.text()
        name   = name_item.text() if name_item else ""
        pan_item = self._table.item(row, self._COL_PAN)
        pan = pan_item.text() if pan_item else ""
        key = self._row_key_for_row(row)
        ts  = self._rows_data.get(key, {}).get("ts", "")
        msg = f"{name}  ({pan})\n"
        if ts:
            msg += f"{ts}\n"
        msg += f"\n{status}"
        QMessageBox.information(self, "Status Detail", msg)

    def _open_output_dir(self):
        _log_open(f"[OpenFolder] Button clicked: {self._output_dir!r}")
        _open_path(self._output_dir)

    def _open_row_path(self, url: str):
        _log_open(f"[OpenFolder] Row link clicked: {url!r}")
        _open_path(url)

    def _set_status_item(self, row: int, text: str):
        _bt = _t()
        _, light_fg, dark_fg = _status_style(text)
        fg = dark_fg if getattr(_bt, "name", "").lower() != "light" else light_fg
        item = QTableWidgetItem(text)
        item.setForeground(QColor(fg))
        item.setFont(QFont(_UI_FONT, 10))
        item.setToolTip(text)
        self._table.setItem(row, self._COL_STATUS, item)

    def _on_update(self, pan: str, ay_label: str, status: str):
        """Live status text update only — does NOT count the (client, year)
        row as done. A multi-select batch runs several doc types per client
        in sequence, each ending in its own terminal-looking status (e.g.
        "✅ 26AS Downloaded" while Filed Returns/AIS are still queued for
        the same client/year), so "saw a terminal glyph" can no longer mean
        "this row is finished" — only the explicit client_finished() call
        means that."""
        key = (pan, ay_label)
        row = self._row_key_to_row.get(key)
        if row is None:
            return
        self._set_status_item(row, status)
        if key in self._rows_data:
            self._rows_data[key]["status"] = status
            self._rows_data[key]["ts"] = datetime.datetime.now().strftime("%d-%b-%Y %H:%M:%S")

    def client_finished(self, pan: str, ay_label: str):
        """Thread-safe: call once per (client, year) row, after ALL selected
        doc types have finished (success or failure) for that year — the
        only correct signal that this row is truly done."""
        self._client_done_signal.emit(pan, ay_label)

    def _on_client_done(self, pan: str, ay_label: str):
        key = (pan, ay_label)
        if key not in self._counted_keys:
            self._counted_keys.add(key)
            self._done_count += 1
            self._progress_bar.setValue(self._done_count)
            self._progress_bar.setFormat(f"{self._done_count} / {self._total} done")
        if self._done_count >= self._total:
            self._close_btn.setEnabled(True)
            self._report_btn.setEnabled(True)
            self._progress_bar.setFormat(f"All {self._total} done")

    def _on_path_update(self, pan: str, ay_label: str, folder: str):
        key = (pan, ay_label)
        row = self._row_key_to_row.get(key)
        if row is None:
            return
        self._path_by_key[key] = folder
        if key in self._rows_data:
            self._rows_data[key]["path"] = folder
        lbl = self._table.cellWidget(row, self._COL_PATH)
        if isinstance(lbl, QLabel):
            lbl.setText(
                f'<a href="{folder}" style="color:#2563EB;text-decoration:underline;">'
                f'{folder}</a>')
            lbl.setToolTip(folder)
            lbl.setStyleSheet("font-size:11px;padding:0 8px;background:transparent;")

    def _on_skip_clicked(self):
        if self._skip_callback:
            self._skip_callback()
        self._skip_btn.setEnabled(False)
        self._skip_btn.setText("⏭  Skipping...")

    def _on_tray_clicked(self):
        if self._tray_callback:
            self._tray_callback()

    def _on_stop_clicked(self):
        if self._stop_callback:
            self._stop_callback()
        self._stop_btn.setEnabled(False)
        self._stop_btn.setText("⏹  Stopping...")

    # ── Excel report ──────────────────────────────────────────────────────────

    def _export_report(self):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        default_name = f"BatchReport_{self._ay.replace(' ','_')}_{timestamp}.xlsx"
        default_path = os.path.join(self._output_dir or os.path.expanduser("~"), default_name)

        path, _ = QFileDialog.getSaveFileName(
            self, "Save Download Report", default_path, "Excel Files (*.xlsx)")
        if not path:
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Download Report"

        hdr_fill  = PatternFill("solid", fgColor="0F172A")
        hdr_font  = Font(bold=True, color="FFFFFF", size=11)
        link_font = Font(color="2563EB", underline="single", size=10)
        body_font = Font(size=10)
        center    = Alignment(horizontal="center", vertical="center")
        left      = Alignment(horizontal="left",   vertical="center", wrap_text=False)
        thin      = Side(style="thin", color="CBD5E1")
        border    = Border(left=thin, right=thin, top=thin, bottom=thin)

        headers    = ["#", "Client Name", "Year", "Save Folder", "Status", "Timestamp"]
        col_widths = [5, 30, 14, 60, 40, 22]

        for col_idx, (h, w) in enumerate(zip(headers, col_widths), start=1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font      = hdr_font
            cell.fill      = hdr_fill
            cell.alignment = center
            cell.border    = border
            ws.column_dimensions[get_column_letter(col_idx)].width = w

        ws.row_dimensions[1].height = 22

        for seq, (tgt, spec) in enumerate(self._client_year_pairs, start=1):
            pan     = tgt.get("pan", "")
            key     = (pan, spec["ay_label"])
            data    = self._rows_data.get(key, {})
            row_num = seq + 1
            folder  = data.get("path", "")
            status  = data.get("status", "—")
            name    = data.get("name", tgt.get("name", ""))
            row_ts  = data.get("ts", "")

            ws.cell(row=row_num, column=1, value=seq).alignment = center
            ws.cell(row=row_num, column=2, value=name).alignment = left
            ws.cell(row=row_num, column=3, value=spec["ay_label"]).alignment = center

            if folder and os.path.exists(folder):
                cell = ws.cell(row=row_num, column=4, value=folder)
                cell.hyperlink = folder
                cell.font      = link_font
                cell.alignment = left
            else:
                ws.cell(row=row_num, column=4, value=folder or "—").alignment = left

            import re as _re
            status_clean = _re.sub(r'[^\x00-\x7F✅❌🕐⬜⏹⏳]+', '', status).strip()
            ws.cell(row=row_num, column=5, value=status_clean).alignment = left
            ws.cell(row=row_num, column=6, value=row_ts or "—").alignment = center

            for col_idx in range(1, 7):
                cell = ws.cell(row=row_num, column=col_idx)
                cell.border = border
                if not cell.font or cell.font == Font():
                    cell.font = body_font
            ws.row_dimensions[row_num].height = 18

        ws.freeze_panes = "A2"

        try:
            wb.save(path)
            _open_path(path)
        except Exception as e:
            QMessageBox.warning(self, "Export Failed", str(e))

    # ── public API ────────────────────────────────────────────────────────────

    def set_status(self, pan: str, ay_label: str, status: str):
        """Thread-safe status update."""
        self._update_signal.emit(pan, ay_label, status)

    def set_client_path(self, pan: str, ay_label: str, folder: str):
        """Thread-safe path update — call once the (client, year) folder is known."""
        self._path_signal.emit(pan, ay_label, folder)

    def batch_finished(self, aborted: bool = False):
        """Enable Close/Report and hide Stop. If aborted, sweeps non-terminal rows to ⏹ Stopped."""
        terminal = ("✅", "❌", "🕐", "⏹", "⬜ Skipped")
        if aborted:
            for key, row in self._row_key_to_row.items():
                item = self._table.item(row, self._COL_STATUS)
                current = item.text() if item else ""
                if not any(current.startswith(t) for t in terminal):
                    self._set_status_item(row, "⏹ Stopped")
                    if key in self._rows_data:
                        self._rows_data[key]["status"] = "⏹ Stopped"
        self._skip_btn.setVisible(False)
        self._tray_btn.setVisible(False)
        self._stop_btn.setVisible(False)
        self._resume_btn.setVisible(aborted)
        self._close_btn.setEnabled(True)
        self._report_btn.setEnabled(True)
        n = self._done_count
        self._progress_bar.setValue(n)
        label = "Stopped" if aborted else "All done"
        self._progress_bar.setFormat(f"{label} — {n} / {self._total} processed")
        # F-37: auto-save report to output folder
        self._auto_save_report()

    def _open_last_report(self):
        if self._last_report_path and os.path.exists(self._last_report_path):
            _open_path(self._last_report_path)
        else:
            from PyQt6.QtWidgets import QMessageBox
            QMessageBox.information(self, "No Report", "No report has been generated yet.")

    def _auto_save_report(self):
        """F-37: silently save a timestamped report to _app_dir()/BatchReports/ after every run."""
        try:
            from config import _app_dir
            reports_dir = os.path.join(_app_dir(), "BatchReports")
            os.makedirs(reports_dir, exist_ok=True)
            import openpyxl, re as _re
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            timestamp    = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            report_name  = f"BatchReport_{self._ay.replace(' ', '_')}_{timestamp}.xlsx"
            path         = os.path.join(reports_dir, report_name)
            hdr_fill  = PatternFill("solid", fgColor="0F172A")
            hdr_font  = Font(bold=True, color="FFFFFF", size=11)
            link_font = Font(color="2563EB", underline="single", size=10)
            body_font = Font(size=10)
            center    = Alignment(horizontal="center", vertical="center")
            left      = Alignment(horizontal="left",   vertical="center", wrap_text=False)
            thin      = Side(style="thin", color="CBD5E1")
            border    = Border(left=thin, right=thin, top=thin, bottom=thin)
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Download Report"
            headers    = ["#", "Client Name", "Year", "Save Folder", "Status", "Timestamp"]
            col_widths = [5, 30, 14, 60, 40, 22]
            for col_idx, (h, w) in enumerate(zip(headers, col_widths), start=1):
                cell = ws.cell(row=1, column=col_idx, value=h)
                cell.font = hdr_font; cell.fill = hdr_fill
                cell.alignment = center; cell.border = border
                ws.column_dimensions[get_column_letter(col_idx)].width = w
            ws.row_dimensions[1].height = 22
            for seq, (tgt, spec) in enumerate(self._client_year_pairs, start=1):
                pan     = tgt.get("pan", "")
                key     = (pan, spec["ay_label"])
                data    = self._rows_data.get(key, {})
                row_num = seq + 1
                folder  = data.get("path", "")
                status  = data.get("status", "—")
                name    = data.get("name", tgt.get("name", ""))
                row_ts  = data.get("ts", "")
                ws.cell(row=row_num, column=1, value=seq).alignment = center
                ws.cell(row=row_num, column=2, value=name).alignment = left
                ws.cell(row=row_num, column=3, value=spec["ay_label"]).alignment = center
                if folder and os.path.exists(folder):
                    cell = ws.cell(row=row_num, column=4, value=folder)
                    cell.hyperlink = folder; cell.font = link_font; cell.alignment = left
                else:
                    ws.cell(row=row_num, column=4, value=folder or "—").alignment = left
                status_clean = _re.sub(r'[^\x00-\x7F✅❌🕐⬜⏹⏳]+', '', status).strip()
                ws.cell(row=row_num, column=5, value=status_clean).alignment = left
                ws.cell(row=row_num, column=6, value=row_ts or "—").alignment = center
                for col_idx in range(1, 7):
                    cell = ws.cell(row=row_num, column=col_idx)
                    cell.border = border
                    if not cell.font or cell.font == Font():
                        cell.font = body_font
                ws.row_dimensions[row_num].height = 18
            ws.freeze_panes = "A2"
            wb.save(path)
            self._last_report_path = path
        except Exception:
            pass   # Never crash the UI due to report saving

    def batch_resumed(self):
        """Reset UI back to running state after Resume is clicked."""
        self._resume_btn.setVisible(False)
        self._skip_btn.setText("⏭  Skip")
        self._skip_btn.setEnabled(True)
        self._skip_btn.setVisible(True)
        self._tray_btn.setVisible(bool(self._tray_callback))
        self._stop_btn.setText("⏹  Stop")
        self._stop_btn.setEnabled(True)
        self._stop_btn.setVisible(True)
        self._close_btn.setEnabled(False)
        self._report_btn.setEnabled(False)
        self._progress_bar.setFormat(f"{self._done_count} / {self._total} done")

    def client_started(self):
        """Re-enable the Skip button when a new client begins downloading."""
        self._skip_btn.setText("⏭  Skip")
        self._skip_btn.setEnabled(True)

    def _on_resume_clicked(self):
        """A client is retried as a whole (login happens once, then every
        selected year re-runs), so a client is 'remaining' if ANY of its
        year-rows was left ⏹ Stopped — and ALL its rows reset to Waiting."""
        remaining = []
        for tgt in self._targets:
            pan = tgt.get("pan", "")
            client_keys = [k for k in self._row_key_to_row if k[0] == pan]
            any_stopped = any(
                (self._rows_data.get(k) or {}).get("status", "").startswith("⏹")
                for k in client_keys)
            if any_stopped:
                remaining.append(tgt)
                for k in client_keys:
                    row = self._row_key_to_row.get(k)
                    if row is not None:
                        self._set_status_item(row, "⬜ Waiting")
                        self._rows_data[k]["status"] = "⬜ Waiting"
        if remaining and self._resume_callback:
            self._resume_callback(remaining)


# ── SMTP provider presets ─────────────────────────────────────────────────────

_SMTP_PRESETS = [
    {
        "name": "Gmail",
        "icon": "G", "icon_color": "#EA4335", "icon_file": "email_gmail.png",
        "host": "smtp.gmail.com",
        "port": 587,
        "encryption": "STARTTLS",
        "help": (
            "Gmail requires an App Password — not your Google account password.<br>"
            "Go to: <a href='https://myaccount.google.com/apppasswords' style='color:#2563EB;'>myaccount.google.com → Security → 2-Step Verification → App Passwords</a> → Mail."
        ),
    },
    {
        "name": "Outlook.com",
        "icon": "O", "icon_color": "#0078D4", "icon_file": "email_outlook.png",
        "host": "smtp-mail.outlook.com",
        "port": 587,
        "encryption": "STARTTLS",
        "help": (
            "Use your Outlook.com / Hotmail password.<br>"
            "If MFA is on, create an App Password at <a href='https://account.microsoft.com/security' style='color:#2563EB;'>account.microsoft.com → Security</a>."
        ),
    },
    {
        "name": "Office 365",
        "icon": "365", "icon_color": "#D83B01", "icon_file": "email_office365.png",
        "host": "smtp.office365.com",
        "port": 587,
        "encryption": "STARTTLS",
        "help": (
            "⚠ MFA enabled? Your regular password will NOT work — use an App Password instead.<br><br>"
            "To create an App Password:<br>"
            "1. Go to <a href='https://mysignins.microsoft.com/security-info' style='color:#2563EB;'>mysignins.microsoft.com/security-info</a><br>"
            "2. Click '+ Add sign-in method' → choose 'App password' → Next<br>"
            "3. Enter a name (e.g. AayDocCapio) → copy the generated password → paste it here<br><br>"
            "No MFA? Enable Authenticated SMTP in <a href='https://admin.microsoft.com' style='color:#2563EB;'>Microsoft 365 Admin Centre</a>:<br>"
            "Users → [your user] → Mail → Manage email apps → tick Authenticated SMTP."
        ),
    },
    {
        "name": "Exchange",
        "icon": "Ex", "icon_color": "#0F6CBD", "icon_file": "email_exchange.png",
        "host": "",
        "port": 587,
        "encryption": "STARTTLS",
        "help": (
            "Enter your organisation's Exchange SMTP server address.<br>"
            "Typical format: <b>mail.yourcompany.com</b> — ask your IT admin if unsure."
        ),
    },
    {
        "name": "Yahoo",
        "icon": "Y!", "icon_color": "#6001D2", "icon_file": "email_yahoo.png",
        "host": "smtp.mail.yahoo.com",
        "port": 587,
        "encryption": "STARTTLS",
        "help": (
            "Yahoo requires an App Password.<br>"
            "Go to: <a href='https://login.yahoo.com/account/security' style='color:#2563EB;'>Yahoo Account Security</a> → Generate app password → select 'Other app'."
        ),
    },
    {
        "name": "iCloud",
        "icon": "iC", "icon_color": "#3478F6", "icon_file": "email_icloud.png",
        "host": "smtp.mail.me.com",
        "port": 587,
        "encryption": "STARTTLS",
        "help": (
            "Use an App-Specific Password from <a href='https://appleid.apple.com' style='color:#2563EB;'>appleid.apple.com</a><br>"
            "→ Sign-In and Security → App-Specific Passwords → Generate."
        ),
    },
    {
        "name": "Custom",
        "icon": "⚙", "icon_color": "#64748B", "icon_file": "email_custom.png",
        "host": None,
        "port": None,
        "encryption": None,
        "help": "",
    },
]

# Map known SMTP hosts → preset name for auto-highlight on load
_HOST_TO_PRESET = {p["host"]: p["name"] for p in _SMTP_PRESETS if p["host"]}

# Pixmap cache — built once per process, reused across dialog opens
_TILE_PIXMAP_CACHE: dict[str, QPixmap] = {}


# ── Selective Export / Import Dialogs ────────────────────────────────────────

class _SelectiveExportDialog(QDialog):
    """Choose what to include in an email-settings export."""

    def __init__(self, parent, tpl_names: list[str]):
        super().__init__(parent)
        self.setWindowTitle("Export Email Settings")
        self.setModal(True)
        self.setMinimumWidth(380)
        self.setSizePolicy(QSizePolicy.Policy.Preferred, QSizePolicy.Policy.Maximum)
        lay = QVBoxLayout(self)
        lay.setSpacing(10)
        lay.setContentsMargins(20, 16, 20, 16)

        lay.addWidget(_lbl("Select what to export:", bold=True))

        # SMTP section
        self._chk_smtp = QCheckBox("SMTP / Sender Settings")
        self._chk_smtp.setChecked(True)
        lay.addWidget(self._chk_smtp)

        # Templates section
        self._chk_tpls = QCheckBox("Email Templates")
        self._chk_tpls.setChecked(True)
        lay.addWidget(self._chk_tpls)

        # Template list — always visible, disabled when checkbox unchecked
        from PyQt6.QtWidgets import QListWidget, QListWidgetItem
        self._tpl_list = QListWidget()
        self._tpl_list.setSelectionMode(QAbstractItemView.SelectionMode.MultiSelection)
        for name in tpl_names:
            item = QListWidgetItem(name)
            self._tpl_list.addItem(item)
            item.setSelected(True)
        self._tpl_list.setFixedHeight(min(max(len(tpl_names), 2), 8) * 28 + 4)
        lay.addWidget(self._tpl_list)

        def _toggle_tpl_list(enabled: bool):
            self._tpl_list.setEnabled(enabled)
            if enabled:
                for i in range(self._tpl_list.count()):
                    self._tpl_list.item(i).setSelected(True)
            else:
                self._tpl_list.clearSelection()
        self._chk_tpls.toggled.connect(_toggle_tpl_list)

        sep = QFrame(); sep.setFrameShape(QFrame.Shape.HLine)
        lay.addWidget(sep)

        btn_row = QHBoxLayout()
        btn_row.setSpacing(8)
        cancel = _btn("Cancel", "outline", height=34)
        cancel.clicked.connect(self.reject)
        ok = _btn("Export", "primary", height=34)
        ok.clicked.connect(self._on_ok)
        btn_row.addWidget(cancel)
        btn_row.addStretch()
        btn_row.addWidget(ok)
        lay.addLayout(btn_row)
        self.adjustSize()

    def _on_ok(self):
        if not self._chk_smtp.isChecked() and not self._chk_tpls.isChecked():
            QMessageBox.warning(self, "Nothing Selected", "Please select at least one item to export.")
            return
        if self._chk_tpls.isChecked() and not self._tpl_list.selectedItems():
            QMessageBox.warning(self, "No Templates", "Please select at least one template.")
            return
        self.accept()

    def result_choices(self) -> tuple[bool, bool, set[str]]:
        include_smtp = self._chk_smtp.isChecked()
        include_tpls = self._chk_tpls.isChecked()
        selected = {item.text() for item in self._tpl_list.selectedItems()}
        return include_smtp, include_tpls, selected


class _SelectiveImportDialog(QDialog):
    """Choose what to apply from an email-settings import file."""

    def __init__(self, parent, has_smtp: bool, tpl_names: list[str]):
        super().__init__(parent)
        self.setWindowTitle("Import Email Settings")
        self.setModal(True)
        self.setMinimumWidth(380)
        self.setSizePolicy(QSizePolicy.Policy.Preferred, QSizePolicy.Policy.Maximum)
        lay = QVBoxLayout(self)
        lay.setSpacing(10)
        lay.setContentsMargins(20, 16, 20, 16)

        lay.addWidget(_lbl("Select what to import:", bold=True))

        self._chk_smtp = QCheckBox("SMTP / Sender Settings")
        self._chk_smtp.setChecked(has_smtp)
        self._chk_smtp.setEnabled(has_smtp)
        if not has_smtp:
            self._chk_smtp.setText("SMTP / Sender Settings (not in file)")
        lay.addWidget(self._chk_smtp)

        has_tpls = bool(tpl_names)
        self._chk_tpls = QCheckBox("Email Templates")
        self._chk_tpls.setChecked(has_tpls)
        self._chk_tpls.setEnabled(has_tpls)
        if not has_tpls:
            self._chk_tpls.setText("Email Templates (not in file)")
        lay.addWidget(self._chk_tpls)

        from PyQt6.QtWidgets import QListWidget, QListWidgetItem
        self._tpl_list = QListWidget()
        self._tpl_list.setSelectionMode(QAbstractItemView.SelectionMode.MultiSelection)
        for name in tpl_names:
            item = QListWidgetItem(name)
            self._tpl_list.addItem(item)
            item.setSelected(True)
        self._tpl_list.setFixedHeight(min(max(len(tpl_names), 2), 8) * 28 + 4)
        self._tpl_list.setEnabled(has_tpls)
        lay.addWidget(self._tpl_list)

        if has_tpls:
            def _toggle_tpl_list(enabled: bool):
                self._tpl_list.setEnabled(enabled)
                if enabled:
                    for i in range(self._tpl_list.count()):
                        self._tpl_list.item(i).setSelected(True)
                else:
                    self._tpl_list.clearSelection()
            self._chk_tpls.toggled.connect(_toggle_tpl_list)

        note = _lbl("Existing templates with the same name will be overwritten.")
        note.setWordWrap(True)
        lay.addWidget(note)

        sep = QFrame(); sep.setFrameShape(QFrame.Shape.HLine)
        lay.addWidget(sep)

        btn_row = QHBoxLayout()
        btn_row.setSpacing(8)
        cancel = _btn("Cancel", "outline", height=34)
        cancel.clicked.connect(self.reject)
        ok = _btn("Import", "primary", height=34)
        ok.clicked.connect(self._on_ok)
        btn_row.addWidget(cancel)
        btn_row.addStretch()
        btn_row.addWidget(ok)
        lay.addLayout(btn_row)
        self.adjustSize()

    def _on_ok(self):
        if not self._chk_smtp.isChecked() and not self._chk_tpls.isChecked():
            QMessageBox.warning(self, "Nothing Selected", "Please select at least one item to import.")
            return
        if self._chk_tpls.isChecked() and self._tpl_list.isEnabled() and not self._tpl_list.selectedItems():
            QMessageBox.warning(self, "No Templates", "Please select at least one template.")
            return
        self.accept()

    def result_choices(self) -> tuple[bool, bool, set[str]]:
        import_smtp = self._chk_smtp.isChecked()
        import_tpls = self._chk_tpls.isChecked()
        selected = {item.text() for item in self._tpl_list.selectedItems()}
        return import_smtp, import_tpls, selected


# ── SMTP Settings Dialog ──────────────────────────────────────────────────────

class SmtpSettingsDialog(QDialog):
    """Configure SMTP credentials, firm name, BCC addresses, and email templates."""

    _test_result = pyqtSignal(bool, str)   # (success, message)

    def __init__(self, parent, vault):
        super().__init__(parent)
        self._vault = vault
        self._tile_btns: dict[str, QPushButton] = {}
        self._selected_preset: str | None = None
        self._test_result.connect(self._on_test_result)
        self.setWindowTitle("Email Settings")
        self.setMinimumSize(1020, 620)
        self.setModal(True)
        self.setWindowFlags(
            Qt.WindowType.Dialog |
            Qt.WindowType.WindowTitleHint |
            Qt.WindowType.WindowSystemMenuHint |
            Qt.WindowType.WindowCloseButtonHint |
            Qt.WindowType.WindowMaximizeButtonHint)
        self._build_ui()

    # ── build ─────────────────────────────────────────────────────────────────

    def _build_ui(self):
        t = _t()
        field_ss = (
            f"QLineEdit{{border:1px solid {t.border};border-radius:6px;padding:6px 10px;"
            f"font-size:12px;background:{t.bg_input};color:{t.text_primary};}}"
            f"QLineEdit:focus{{border-color:{t.border_focus};background:{t.bg_input_focus};}}"
        )
        self.setStyleSheet(
            f"QDialog{{background:{t.bg_window};}}"
            f"QLabel{{color:{t.text_primary};background:transparent;font-size:12px;}}"
            + field_ss +
            f"QTextEdit{{border:1px solid {t.border};border-radius:6px;padding:8px 10px;"
            f"font-size:12px;background:{t.bg_input};color:{t.text_primary};}}"
            f"QTextEdit:focus{{border-color:{t.border_focus};}}"
            f"QSpinBox{{border:1px solid {t.border};border-radius:6px;padding:4px 8px;"
            f"font-size:12px;background:{t.bg_input};color:{t.text_primary};}}"
            f"QComboBox{{border:1px solid {t.border};border-radius:6px;padding:4px 10px;"
            f"font-size:12px;background:{t.bg_input};color:{t.text_primary};}}"
            f"QComboBox::drop-down{{border:none;width:20px;}}"
            f"QComboBox::down-arrow{{image:url(none);width:0;height:0;"
            f"border-left:4px solid transparent;border-right:4px solid transparent;"
            f"border-top:5px solid {t.text_primary};}}"
            f"QComboBox QAbstractItemView{{background:{t.bg_input};color:{t.text_primary};"
            f"selection-background-color:{t.accent};}}"
        )

        cfg = self._vault.get_email_settings()

        # ── outer: header bar (fixed) + tab widget (stretchy) + footer bar (fixed)
        outer = QVBoxLayout(self)
        outer.setContentsMargins(0, 0, 0, 0)
        outer.setSpacing(0)

        # ── Header bar ────────────────────────────────────────────────────────
        hdr_widget = QWidget()
        hdr_widget.setStyleSheet(
            f"QWidget{{background:{t.bg_window};}}"
            f"QLabel{{color:{t.text_primary};}}"
        )
        hdr_lay = QVBoxLayout(hdr_widget)
        hdr_lay.setContentsMargins(16, 14, 16, 12)
        hdr_lay.setSpacing(0)

        title_row = QHBoxLayout()
        title_row.addWidget(_lbl("Email Settings", 15, bold=True))
        title_row.addStretch()
        help_btn = QPushButton("? Help")
        help_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        help_btn.setStyleSheet(
            f"QPushButton{{background:transparent;color:{t.accent};border:none;"
            f"font-size:12px;font-weight:bold;padding:2px 6px;}}"
            f"QPushButton:hover{{text-decoration:underline;}}"
        )
        help_btn.clicked.connect(self._show_help)
        title_row.addWidget(help_btn)
        hdr_lay.addLayout(title_row)
        hdr_lay.addSpacing(3)
        hdr_lay.addWidget(_lbl(
            "Configure SMTP server and email template for mailing tax documents to clients.",
            10, color=t.text_muted))

        sep_top = QFrame()
        sep_top.setFrameShape(QFrame.Shape.HLine)
        sep_top.setStyleSheet(f"background:{t.border};border:none;max-height:1px;")

        outer.addWidget(hdr_widget)
        outer.addWidget(sep_top)

        # ── Shared scroll area stylesheet ─────────────────────────────────────
        _scroll_ss = (
            f"QScrollArea{{background:{t.bg_window};border:none;}}"
            f"QScrollBar:vertical{{width:8px;background:{t.bg_table_alt};}}"
            f"QScrollBar::handle:vertical{{background:{t.border};border-radius:4px;min-height:24px;}}"
            f"QScrollBar::add-line:vertical,QScrollBar::sub-line:vertical{{height:0px;}}"
        )

        # ── Tab widget ────────────────────────────────────────────────────────
        tab = QTabWidget()
        tab.setStyleSheet(
            f"QTabWidget::pane{{border:1px solid {t.border};border-top:none;}}"
            f"QTabBar::tab{{background:{t.bg_table_alt};color:{t.text_muted};"
            f"padding:8px 20px;border:1px solid {t.border};border-bottom:none;"
            f"border-radius:6px 6px 0 0;margin-right:2px;font-size:12px;}}"
            f"QTabBar::tab:selected{{background:{t.accent};color:white;"
            f"border-color:{t.accent};font-weight:600;}}"
            f"QTabBar::tab:hover:!selected{{background:{t.bg_input};}}"
        )
        outer.addWidget(tab, stretch=1)

        # ── Tab 1 — SMTP / Sender ─────────────────────────────────────────────
        tab1_scroll = QScrollArea()
        tab1_scroll.setWidgetResizable(True)
        tab1_scroll.setFrameShape(QFrame.Shape.NoFrame)
        tab1_scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        tab1_scroll.setStyleSheet(_scroll_ss)
        tab1_content = QWidget()
        tab1_content.setStyleSheet(f"QWidget{{background:{t.bg_window};}}")
        tab1_main = QVBoxLayout(tab1_content)
        tab1_main.setContentsMargins(16, 16, 16, 16)
        tab1_main.setSpacing(0)
        tab1_scroll.setWidget(tab1_content)
        # ── Tab 1 — Templates ─────────────────────────────────────────────────
        tab2_widget = QWidget()
        tab2_widget.setStyleSheet(f"QWidget{{background:{t.bg_window};}}")
        tab2_outer = QHBoxLayout(tab2_widget)
        tab2_outer.setContentsMargins(0, 0, 0, 0)
        tab2_outer.setSpacing(0)
        _tpl_icon_p = _icon_path("menu_template.png")
        _tpl_icon = QIcon(QPixmap(_tpl_icon_p).scaled(
            16, 16, Qt.AspectRatioMode.KeepAspectRatio,
            Qt.TransformationMode.SmoothTransformation)) if _tpl_icon_p else QIcon()
        tab.addTab(tab2_widget, _tpl_icon, "Templates")

        # ── Tab 2 — SMTP / Sender ─────────────────────────────────────────────
        tab.addTab(tab1_scroll, "SMTP / Sender")

        def _flbl(text):
            l = QLabel(text)
            l.setStyleSheet(
                f"font-size:11px;font-weight:600;color:{t.text_muted};margin-bottom:3px;")
            return l

        # ─────────────────────────────────────────────────────────────────────
        # Tab 1 content
        # ─────────────────────────────────────────────────────────────────────

        # Provider picker
        tab1_main.addWidget(_lbl("Select your email provider", 12, bold=True))
        tab1_main.addSpacing(10)

        tile_row = QHBoxLayout()
        tile_row.setSpacing(8)
        for preset in _SMTP_PRESETS:
            btn = self._make_tile(preset)
            self._tile_btns[preset["name"]] = btn
            tile_row.addWidget(btn)
        tile_row.addStretch()
        tab1_main.addLayout(tile_row)
        tab1_main.addSpacing(10)

        self._help_note = QLabel("")
        self._help_note.setWordWrap(True)
        self._help_note.setOpenExternalLinks(True)
        self._help_note.setTextFormat(Qt.TextFormat.RichText)
        self._help_note.setStyleSheet(
            f"background:#EFF6FF;color:#1E3A5F;"
            f"border-left:4px solid #2563EB;border-top:1px solid #BFDBFE;"
            f"border-right:1px solid #BFDBFE;border-bottom:1px solid #BFDBFE;"
            f"border-radius:0 6px 6px 0;padding:10px 14px;"
            f"font-size:11px;line-height:160%;")
        self._help_note.hide()
        tab1_main.addWidget(self._help_note)
        tab1_main.addSpacing(20)

        # SMTP Server / Port / Encryption
        row_host = QHBoxLayout(); row_host.setSpacing(10)

        host_col = QVBoxLayout(); host_col.setSpacing(4)
        host_col.addWidget(_flbl("SMTP Server"))
        self._host = QLineEdit(cfg.get("smtp_host", ""))
        self._host.setPlaceholderText("smtp.gmail.com")
        self._host.setFixedHeight(34)
        host_col.addWidget(self._host)
        row_host.addLayout(host_col, stretch=1)

        port_col = QVBoxLayout(); port_col.setSpacing(4)
        port_col.addWidget(_flbl("Port"))
        self._port = QSpinBox()
        self._port.setRange(1, 65535)
        self._port.setValue(int(cfg.get("smtp_port", 587)))
        self._port.setFixedHeight(34)
        self._port.setFixedWidth(90)
        port_col.addWidget(self._port)
        row_host.addLayout(port_col)

        enc_col = QVBoxLayout(); enc_col.setSpacing(4)
        enc_col.addWidget(_flbl("Encryption"))
        self._enc = QComboBox()
        self._enc.addItems(["STARTTLS", "SSL/TLS", "None"])
        saved_enc = cfg.get("smtp_encryption", "STARTTLS")
        self._enc.setCurrentIndex(max(0, self._enc.findText(saved_enc)))
        self._enc.setFixedHeight(34)
        self._enc.setFixedWidth(110)
        enc_col.addWidget(self._enc)
        row_host.addLayout(enc_col)
        tab1_main.addLayout(row_host)
        tab1_main.addSpacing(12)

        # Username / Password
        user_pwd_row = QHBoxLayout(); user_pwd_row.setSpacing(12)

        user_col = QVBoxLayout(); user_col.setSpacing(4)
        user_col.addWidget(_flbl("Username / Email"))
        self._user = QLineEdit(cfg.get("smtp_user", ""))
        self._user.setPlaceholderText("you@gmail.com")
        self._user.setFixedHeight(34)
        user_col.addWidget(self._user)
        user_pwd_row.addLayout(user_col, stretch=35)

        pwd_col = QVBoxLayout(); pwd_col.setSpacing(4)
        pwd_col.addWidget(_flbl("Password / App Password"))
        pwd_field_row = QHBoxLayout(); pwd_field_row.setSpacing(8)
        self._pwd = QLineEdit(cfg.get("smtp_password", ""))
        self._pwd.setEchoMode(QLineEdit.EchoMode.Password)
        self._pwd.setPlaceholderText("Enter password")
        self._pwd.setFixedHeight(34)
        pwd_field_row.addWidget(self._pwd)
        show_pwd_cb = QCheckBox("Show password")
        show_pwd_cb.setStyleSheet(
            f"QCheckBox{{font-size:11px;color:{t.text_muted};}}"
            f"QCheckBox::indicator{{width:14px;height:14px;}}"
        )
        show_pwd_cb.toggled.connect(
            lambda on: self._pwd.setEchoMode(
                QLineEdit.EchoMode.Normal if on else QLineEdit.EchoMode.Password))
        pwd_field_row.addWidget(show_pwd_cb)
        pwd_col.addLayout(pwd_field_row)
        user_pwd_row.addLayout(pwd_col, stretch=25)

        tab1_main.addLayout(user_pwd_row)
        tab1_main.addSpacing(10)

        # Send As / From + BCC
        from_bcc_row = QHBoxLayout(); from_bcc_row.setSpacing(12)

        from_col = QVBoxLayout(); from_col.setSpacing(4)
        from_col.addWidget(_flbl("Send As / From Address (optional — leave blank to use username)"))
        self._from = QLineEdit(cfg.get("smtp_from", ""))
        self._from.setPlaceholderText("income-tax@daksm.com")
        self._from.setFixedHeight(34)
        from_col.addWidget(self._from)
        from_bcc_row.addLayout(from_col, stretch=1)

        bcc_col = QVBoxLayout(); bcc_col.setSpacing(4)
        bcc_col.addWidget(_flbl("BCC Addresses  (separate multiple with ;)"))
        self._bcc = QLineEdit(cfg.get("bcc_addresses", ""))
        self._bcc.setPlaceholderText("partner@firm.com;team@firm.com")
        self._bcc.setFixedHeight(34)
        bcc_col.addWidget(self._bcc)
        from_bcc_row.addLayout(bcc_col, stretch=1)

        tab1_main.addLayout(from_bcc_row)
        tab1_main.addSpacing(14)

        tab1_main.addWidget(_flbl("Firm Name  (used in {firm_name} placeholder)"))
        tab1_main.addSpacing(3)
        self._firm = QLineEdit(cfg.get("firm_name", ""))
        self._firm.setPlaceholderText("AI Learrning Guru")
        self._firm.setFixedHeight(34)
        tab1_main.addWidget(self._firm)
        tab1_main.addStretch()

        # ─────────────────────────────────────────────────────────────────────
        # Tab 2 content
        # ─────────────────────────────────────────────────────────────────────

        _ph_chip_ss = (
            f"QPushButton{{background:{t.bg_table_alt};color:{t.accent};"
            f"border:1px solid {t.border};border-radius:5px;"
            f"font-size:10px;padding:0 6px;font-family:monospace;}}"
            f"QPushButton:hover{{background:{t.bg_input};border-color:{t.accent};}}"
        )
        _combo_ss = (
            f"QComboBox{{border:1px solid {t.border};border-radius:5px;padding:2px 6px;"
            f"font-size:11px;background:{t.bg_input};color:{t.text_primary};}}"
            f"QComboBox::drop-down{{border:none;width:16px;}}"
            f"QComboBox::down-arrow{{image:url(none);width:0;height:0;"
            f"border-left:4px solid transparent;border-right:4px solid transparent;"
            f"border-top:5px solid {t.text_primary};}}"
            f"QComboBox QAbstractItemView{{background:{t.bg_input};color:{t.text_primary};"
            f"selection-background-color:{t.accent};}}"
        )

        # ── Left panel — template list ────────────────────────────────────────
        left_panel = QWidget()
        left_panel.setFixedWidth(190)
        left_panel.setStyleSheet(
            f"QWidget{{background:{t.bg_panel};border-right:1px solid {t.border};}}")
        left_v = QVBoxLayout(left_panel)
        left_v.setContentsMargins(10, 12, 10, 10)
        left_v.setSpacing(6)

        left_v.addWidget(_lbl("Templates", 12, bold=True))
        left_v.addSpacing(4)

        from PyQt6.QtWidgets import QListWidget, QListWidgetItem
        self._tpl_list = QListWidget()
        self._tpl_list.setStyleSheet(
            f"QListWidget{{background:{t.bg_panel};border:none;outline:none;font-size:12px;}}"
            f"QListWidget::item{{padding:8px 10px;border-radius:5px;color:{t.text_primary};}}"
            f"QListWidget::item:selected{{background:{t.accent};color:white;}}"
            f"QListWidget::item:hover:!selected{{background:{t.bg_table_alt};}}"
        )
        self._tpl_list.setEditTriggers(QAbstractItemView.EditTrigger.DoubleClicked)
        self._tpl_list.itemChanged.connect(self._on_tpl_renamed)
        self._tpl_list.currentRowChanged.connect(self._on_tpl_select)
        left_v.addWidget(self._tpl_list, stretch=1)

        hint_lbl = QLabel("Double-click to rename")
        hint_lbl.setStyleSheet(f"color:{t.text_muted};font-size:10px;background:transparent;")
        left_v.addWidget(hint_lbl)

        self._tpl_default_btn = _btn("★ Set as Default", "outline", height=28)
        self._tpl_default_btn.clicked.connect(self._tpl_set_default)
        left_v.addWidget(self._tpl_default_btn)

        btn_row = QHBoxLayout(); btn_row.setSpacing(6)
        self._tpl_add_btn = _btn("+ New", "outline", height=28, icon="btn_add_list.png")
        self._tpl_add_btn.clicked.connect(self._tpl_add)
        self._tpl_del_btn = _btn("Delete", "danger", height=28, icon="btn_delete.png")
        self._tpl_del_btn.clicked.connect(self._tpl_delete)
        btn_row.addWidget(self._tpl_add_btn)
        btn_row.addWidget(self._tpl_del_btn)
        left_v.addLayout(btn_row)

        tab2_outer.addWidget(left_panel)

        # ── Right panel — header + editor ─────────────────────────────────────
        right_container = QWidget()
        right_container.setStyleSheet(f"QWidget{{background:{t.bg_window};}}")
        right_container_v = QVBoxLayout(right_container)
        right_container_v.setContentsMargins(0, 0, 0, 0)
        right_container_v.setSpacing(0)

        # Header strip showing current template name + default badge
        self._tpl_header = QWidget()
        self._tpl_header.setStyleSheet(
            f"QWidget{{background:{t.bg_table_alt};border-bottom:1px solid {t.border};}}")
        _hdr_lay = QHBoxLayout(self._tpl_header)
        _hdr_lay.setContentsMargins(16, 10, 16, 10)
        self._tpl_header_lbl = QLabel("")
        self._tpl_header_lbl.setStyleSheet(
            f"font-size:13px;font-weight:bold;color:{t.text_primary};background:transparent;")
        self._tpl_default_lbl = QLabel("Default")
        self._tpl_default_lbl.setStyleSheet(
            f"font-size:10px;color:{t.accent};background:transparent;"
            f"border:1px solid {t.accent};border-radius:4px;padding:1px 6px;")
        self._tpl_default_lbl.hide()
        _hdr_lay.addWidget(self._tpl_header_lbl)
        _hdr_lay.addSpacing(8)
        _hdr_lay.addWidget(self._tpl_default_lbl)
        _hdr_lay.addStretch()
        right_container_v.addWidget(self._tpl_header)

        right_scroll = QScrollArea()
        right_scroll.setWidgetResizable(True)
        right_scroll.setFrameShape(QFrame.Shape.NoFrame)
        right_scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        right_scroll.setStyleSheet(_scroll_ss)
        right_content = QWidget()
        right_content.setStyleSheet(f"QWidget{{background:{t.bg_window};}}")
        right_v = QVBoxLayout(right_content)
        right_v.setContentsMargins(16, 14, 16, 14)
        right_v.setSpacing(0)
        right_scroll.setWidget(right_content)
        right_container_v.addWidget(right_scroll, stretch=1)
        tab2_outer.addWidget(right_container, stretch=1)

        # Template name — stored internally, edited via list item directly
        self._tpl_name = QLineEdit()
        self._tpl_name.hide()

        # Subject
        subj_lbl_row = QHBoxLayout(); subj_lbl_row.setSpacing(6)
        subj_lbl_row.addWidget(_flbl("Subject"))
        subj_lbl_row.addStretch()
        for ph in ["{client_name}", "{ay}", "{firm_name}"]:
            pb = QPushButton(ph)
            pb.setFixedHeight(22)
            pb.setCursor(Qt.CursorShape.PointingHandCursor)
            pb.setStyleSheet(_ph_chip_ss)
            pb.clicked.connect(lambda _, p=ph: self._subj.insert(p))
            subj_lbl_row.addWidget(pb)
        right_v.addLayout(subj_lbl_row)
        right_v.addSpacing(3)
        self._subj = QLineEdit()
        self._subj.setFixedHeight(34)
        right_v.addWidget(self._subj)
        right_v.addSpacing(14)

        # Body
        right_v.addWidget(_flbl("Body"))
        right_v.addSpacing(4)

        _fmt_btn_ss = (
            f"QPushButton{{background:{t.bg_table_alt};color:{t.text_primary};"
            f"border:1px solid {t.border};border-radius:5px;"
            f"font-size:12px;padding:0 8px;min-width:28px;height:28px;}}"
            f"QPushButton:hover{{background:{t.bg_input};border-color:{t.accent};}}"
            f"QPushButton:checked{{background:{t.accent};color:#fff;border-color:{t.accent};}}"
        )

        def _fmt_toggle(label: str, prop: str) -> QPushButton:
            btn = QPushButton(label)
            btn.setCheckable(True)
            btn.setFixedHeight(28)
            btn.setStyleSheet(_fmt_btn_ss)
            btn.setCursor(Qt.CursorShape.PointingHandCursor)
            def _apply(checked, p=prop):
                fmt = QTextCharFormat()
                if p == "bold":
                    fmt.setFontWeight(QFont.Weight.Bold if checked else QFont.Weight.Normal)
                elif p == "italic":
                    fmt.setFontItalic(checked)
                elif p == "underline":
                    fmt.setFontUnderline(checked)
                self._body.mergeCurrentCharFormat(fmt)
                self._body.setFocus()
            btn.toggled.connect(_apply)
            return btn

        fmt_bar = QHBoxLayout(); fmt_bar.setSpacing(6)
        self._font_combo = QFontComboBox()
        self._font_combo.setEditable(False)
        self._font_combo.setFixedHeight(26)
        self._font_combo.setFixedWidth(140)
        self._font_combo.setStyleSheet(_combo_ss)
        self._font_combo.setCurrentFont(QFont(_UI_FONT))
        self._font_combo.currentFontChanged.connect(lambda f: (
            self._body.setFocus(),
            self._body.mergeCurrentCharFormat(
                (lambda fmt: (fmt.setFontFamilies([f.family()]), fmt))
                (QTextCharFormat())[1]
            )
        ))
        fmt_bar.addWidget(self._font_combo)

        self._size_combo = QComboBox()
        self._size_combo.setFixedHeight(26)
        self._size_combo.setFixedWidth(48)
        self._size_combo.setStyleSheet(_combo_ss)
        for sz in ["8", "9", "10", "11", "12", "14", "16", "18"]:
            self._size_combo.addItem(sz)
        self._size_combo.setCurrentText("12")
        self._size_combo.currentTextChanged.connect(lambda sz: (
            self._body.setFocus(),
            self._body.mergeCurrentCharFormat(
                (lambda fmt: (fmt.setFontPointSize(float(sz)), fmt))
                (QTextCharFormat())[1]
            ) if sz.isdigit() else None
        ))
        fmt_bar.addWidget(self._size_combo)

        fmt_bar.addSpacing(2)
        self._fmt_bold      = _fmt_toggle("B", "bold")
        self._fmt_bold.setFont(QFont(_UI_FONT, 10, QFont.Weight.Bold))
        self._fmt_italic    = _fmt_toggle("I", "italic")
        self._fmt_italic.setFont(QFont(_UI_FONT, 10, -1, True))
        self._fmt_underline = _fmt_toggle("U", "underline")
        fmt_bar.addWidget(self._fmt_bold)
        fmt_bar.addWidget(self._fmt_italic)
        fmt_bar.addWidget(self._fmt_underline)
        fmt_bar.addSpacing(6)
        for ph in ["{client_name}", "{pan}", "{ay}", "{firm_name}", "{documents}"]:
            pb = QPushButton(ph)
            pb.setFixedHeight(26)
            pb.setCursor(Qt.CursorShape.PointingHandCursor)
            pb.setStyleSheet(_ph_chip_ss)
            pb.clicked.connect(lambda _, p=ph: self._body.insertPlainText(p))
            fmt_bar.addWidget(pb)
        fmt_bar.addStretch()
        right_v.addLayout(fmt_bar)
        right_v.addSpacing(4)

        self._body = QTextEdit()
        self._body.setAcceptRichText(True)
        self._body.setMinimumHeight(160)
        self._body.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        self._body.currentCharFormatChanged.connect(self._sync_fmt_buttons)
        right_v.addWidget(self._body, stretch=1)
        right_v.addSpacing(14)

        # Document type checkboxes
        right_v.addWidget(_flbl("Attach these document types"))
        right_v.addSpacing(6)
        _cb_ss = (
            f"QCheckBox{{color:{t.text_primary};font-size:12px;spacing:6px;}}"
            f"QCheckBox::indicator{{width:15px;height:15px;border:1.5px solid {t.border};"
            f"border-radius:3px;background:{t.bg_checkbox};}}"
            f"QCheckBox::indicator:checked{{background:{t.accent};border-color:{t.accent};}}"
        )
        self._doc_cbs = {}
        docs_rows = QVBoxLayout(); docs_rows.setSpacing(6)
        docs_row1 = QHBoxLayout(); docs_row1.setSpacing(20)
        docs_row2 = QHBoxLayout(); docs_row2.setSpacing(20)
        for label, key in [
            ("26AS PDF",   "26as_pdf"),
            ("26AS Excel", "26as_xlsx"),
            ("168 PDF",    "168_pdf"),
            ("168 Excel",  "168_xlsx"),
            ("AIS PDF",    "ais_pdf"),
            ("AIS Excel",  "ais_xlsx"),
            ("TIS",        "tis_pdf"),
        ]:
            cb = QCheckBox(label)
            cb.setChecked(True)
            cb.setStyleSheet(_cb_ss)
            self._doc_cbs[key] = cb
            docs_row1.addWidget(cb)
        for label, key in [
            ("ITR Form",    "itr_form"),
            ("ITR Receipt", "itr_receipt"),
            ("ITR-V",       "itr_v"),
            ("Intimation",  "intimation"),
            ("Challan",     "challan_pdf"),
        ]:
            cb = QCheckBox(label)
            cb.setChecked(True)
            cb.setStyleSheet(_cb_ss)
            self._doc_cbs[key] = cb
            docs_row2.addWidget(cb)
        docs_row1.addStretch()
        docs_row2.addStretch()
        docs_rows.addLayout(docs_row1)
        docs_rows.addLayout(docs_row2)
        right_v.addLayout(docs_rows)

        # Load templates into list
        self._templates = self._vault.get_email_templates()
        self._loading_tpl = False
        for tpl in self._templates:
            item = QListWidgetItem(tpl["name"])
            item.setFlags(item.flags() | Qt.ItemFlag.ItemIsEditable)
            self._tpl_list.addItem(item)

        # Select active template
        active_name = self._vault.get_active_template_name()
        for i, tpl in enumerate(self._templates):
            if tpl["name"] == active_name:
                self._tpl_list.setCurrentRow(i)
                break
        else:
            if self._templates:
                self._tpl_list.setCurrentRow(0)

        # Update delete button state and header
        self._tpl_del_btn.setEnabled(len(self._templates) > 1)
        self._refresh_tpl_header()

        # Auto-highlight tile if saved host matches a known preset
        saved_host = cfg.get("smtp_host", "")
        if saved_host in _HOST_TO_PRESET:
            self._highlight_tile(_HOST_TO_PRESET[saved_host])

        # ── Footer bar (fixed, outside tabs) ──────────────────────────────────
        sep_bot = QFrame()
        sep_bot.setFrameShape(QFrame.Shape.HLine)
        sep_bot.setStyleSheet(f"background:{t.border};border:none;max-height:1px;")
        outer.addWidget(sep_bot)

        footer_widget = QWidget()
        footer_widget.setStyleSheet(f"QWidget{{background:{t.bg_window};}}")
        footer_lay = QHBoxLayout(footer_widget)
        footer_lay.setContentsMargins(16, 10, 16, 12)
        footer_lay.setSpacing(8)
        self._test_btn = _btn("Send Test Email", "outline", height=36, icon="btn_send_test.png")
        self._test_btn.clicked.connect(self._send_test)
        footer_lay.addWidget(self._test_btn)
        log_btn = _btn("View Log", "outline", height=36, icon="btn_view_log.png")
        export_btn = _btn("Export Settings", "outline", height=36, icon="menu_export.png")
        export_btn.setToolTip("Export SMTP settings and templates to a JSON file")
        export_btn.clicked.connect(self._export_settings)
        footer_lay.addWidget(export_btn)
        import_btn = _btn("Import Settings", "outline", height=36, icon="menu_import.png")
        import_btn.setToolTip("Import SMTP settings and templates from a JSON file")
        import_btn.clicked.connect(self._import_settings)
        footer_lay.addWidget(import_btn)
        log_btn.clicked.connect(self._open_log)
        footer_lay.addWidget(log_btn)
        footer_lay.addStretch()
        self._ctx_save_btn = _btn("Save SMTP Settings", "outline", height=36, icon="btn_save.png")
        footer_lay.addWidget(self._ctx_save_btn)
        self._save_close_btn = _btn("Cancel", "secondary", height=36, icon="btn_cancel.png")
        self._save_close_btn.clicked.connect(self._on_save_close_clicked)
        footer_lay.addWidget(self._save_close_btn)

        # Wire all fields to mark dirty (switches button to Save & Close)
        for field in (self._host, self._user, self._pwd, self._from, self._bcc, self._firm):
            field.textChanged.connect(self._mark_dirty)
        self._port.valueChanged.connect(self._mark_dirty)
        self._enc.currentIndexChanged.connect(self._mark_dirty)
        self._tpl_list.itemChanged.connect(self._mark_dirty)
        self._subj.textChanged.connect(self._mark_dirty)
        self._body.textChanged.connect(self._mark_dirty)

        def _on_tab_changed(i):
            try:
                self._ctx_save_btn.clicked.disconnect()
            except (RuntimeError, TypeError):
                pass
            if i == 0:
                self._ctx_save_btn.setText("Save Templates")
                self._ctx_save_btn.clicked.connect(self._save_templates_only)
            else:
                self._ctx_save_btn.setText("Save SMTP Settings")
                self._ctx_save_btn.clicked.connect(self._save_smtp_only)

        tab.currentChanged.connect(_on_tab_changed)
        _on_tab_changed(0)   # initialise for Templates tab
        outer.addWidget(footer_widget)

        self.resize(1100, 720)
        # Open on SMTP tab if not yet configured, else Templates
        if not cfg.get("smtp_host"):
            tab.setCurrentIndex(1)
        from PyQt6.QtCore import QTimer
        QTimer.singleShot(0, lambda: tab1_scroll.verticalScrollBar().setValue(0))

    # ── tile factory ──────────────────────────────────────────────────────────

    def _make_tile(self, preset: dict) -> QPushButton:
        """Return an icon-only provider button with tooltip."""
        name = preset["name"]
        icon_text = preset["icon"]
        icon_color = preset["icon_color"]

        # Build pixmap once, cache for reuse
        if name not in _TILE_PIXMAP_CACHE:
            # Try loading PNG from resources/
            from config import _bundled_dir
            png_path = os.path.join(_bundled_dir(), "resources", "icons", preset.get("icon_file", ""))
            loaded = False
            if preset.get("icon_file") and os.path.isfile(png_path):
                src = QPixmap(png_path)
                if not src.isNull():
                    px = src.scaled(40, 40, Qt.AspectRatioMode.KeepAspectRatio,
                                    Qt.TransformationMode.SmoothTransformation)
                    _TILE_PIXMAP_CACHE[name] = px
                    loaded = True
            if not loaded:
                px = QPixmap(40, 40)
                px.fill(Qt.GlobalColor.transparent)
                painter = QPainter(px)
                painter.setRenderHint(QPainter.RenderHint.Antialiasing)
                painter.setBrush(QBrush(QColor(icon_color)))
                painter.setPen(QPen(Qt.GlobalColor.transparent))
                painter.drawRoundedRect(0, 0, 40, 40, 8, 8)
                painter.setPen(QPen(QColor("white")))
                f = painter.font()
                f.setBold(True)
                f.setPixelSize(13 if len(icon_text) > 1 else 18)
                painter.setFont(f)
                painter.drawText(px.rect(), Qt.AlignmentFlag.AlignCenter, icon_text)
                painter.end()
                _TILE_PIXMAP_CACHE[name] = px
        px = _TILE_PIXMAP_CACHE[name]

        from PyQt6.QtGui import QIcon
        btn = QPushButton()
        btn.setFixedSize(52, 52)
        btn.setIcon(QIcon(px))
        btn.setIconSize(px.size())
        btn.setToolTip(name)
        btn.setCursor(Qt.CursorShape.PointingHandCursor)

        self._set_tile_style(btn, selected=False)
        btn.clicked.connect(lambda _checked, n=name: self._apply_preset(n))
        return btn

    def _set_tile_style(self, btn: QPushButton, selected: bool):
        t = _t()
        if selected:
            btn.setStyleSheet(
                f"QPushButton{{background:{t.bg_input_focus};border:2px solid {t.accent};"
                f"border-radius:8px;color:{t.text_primary};}}"
                f"QPushButton:hover{{background:{t.bg_input_focus};}}"
            )
        else:
            btn.setStyleSheet(
                f"QPushButton{{background:{t.bg_table_alt};border:1px solid {t.border};"
                f"border-radius:8px;color:{t.text_primary};}}"
                f"QPushButton:hover{{background:{t.bg_input};border-color:{t.border_focus};}}"
            )

    def _highlight_tile(self, name: str):
        for n, btn in self._tile_btns.items():
            self._set_tile_style(btn, selected=(n == name))
        self._selected_preset = name

    # ── preset apply ─────────────────────────────────────────────────────────

    def _apply_preset(self, name: str):
        self._highlight_tile(name)
        preset = next((p for p in _SMTP_PRESETS if p["name"] == name), None)
        if not preset:
            return

        if preset["host"] is not None:    # None = Custom, leave fields alone
            self._host.setText(preset["host"])
        if preset["port"] is not None:
            self._port.setValue(preset["port"])
        if preset["encryption"] is not None:
            idx = self._enc.findText(preset["encryption"])
            if idx >= 0:
                self._enc.setCurrentIndex(idx)

        help_text = preset.get("help", "")
        if help_text:
            self._help_note.setText(help_text)
            self._help_note.show()
        else:
            self._help_note.hide()

    # ── format button sync ───────────────────────────────────────────────────

    def _sync_fmt_buttons(self, fmt: QTextCharFormat):
        for btn, attr in (
            (self._fmt_bold,      fmt.fontWeight() == QFont.Weight.Bold),
            (self._fmt_italic,    fmt.fontItalic()),
            (self._fmt_underline, fmt.fontUnderline()),
        ):
            btn.blockSignals(True)
            btn.setChecked(attr)
            btn.blockSignals(False)
        families = fmt.fontFamilies()
        if families:
            self._font_combo.blockSignals(True)
            self._font_combo.setCurrentFont(QFont(families[0]))
            self._font_combo.blockSignals(False)
        sz = fmt.fontPointSize()
        if sz > 0:
            self._size_combo.blockSignals(True)
            self._size_combo.setCurrentText(str(int(sz)))
            self._size_combo.blockSignals(False)

    # ── template list interactions ────────────────────────────────────────────

    def _current_tpl_dict(self) -> dict:
        """Read the right-panel editor into a template dict."""
        row = self._tpl_list.currentRow()
        item = self._tpl_list.item(row)
        name = item.text().strip() if item else "Untitled"
        return {
            "name":    name or "Untitled",
            "subject": self._subj.text().strip(),
            "body":    self._body.toHtml(),
            "docs":    {k: cb.isChecked() for k, cb in self._doc_cbs.items()},
        }

    def _on_tpl_renamed(self, item):
        """Sync list item rename back into self._templates."""
        row = self._tpl_list.row(item)
        if 0 <= row < len(self._templates):
            self._templates[row]["name"] = item.text().strip() or "Untitled"

    def _flush_current_tpl(self):
        """Save right-panel edits back into self._templates at the current row."""
        row = self._tpl_list.currentRow()
        if row < 0 or row >= len(self._templates):
            return
        self._templates[row] = self._current_tpl_dict()
        self._tpl_list.item(row).setText(self._templates[row]["name"])

    def _load_tpl_into_editor(self, tpl: dict):
        """Populate right-panel fields from a template dict."""
        self._loading_tpl = True
        self._subj.setText(tpl.get("subject", ""))
        raw_body = tpl.get("body", "")
        if raw_body.lstrip().startswith("<"):
            self._body.setHtml(raw_body)
        else:
            self._body.setPlainText(raw_body)
        docs = tpl.get("docs", {})
        for key, cb in self._doc_cbs.items():
            cb.setChecked(docs.get(key, True))
        self._loading_tpl = False

    def _on_tpl_select(self, row: int):
        if self._loading_tpl or row < 0 or row >= len(self._templates):
            return
        self._load_tpl_into_editor(self._templates[row])
        self._tpl_del_btn.setEnabled(len(self._templates) > 1)
        self._refresh_tpl_header()

    def _refresh_tpl_header(self):
        row = self._tpl_list.currentRow()
        if row < 0 or row >= len(self._templates):
            return
        name = self._templates[row]["name"]
        active_name = self._vault.get_active_template_name()
        self._tpl_header_lbl.setText(name)
        is_default = (name == active_name)
        self._tpl_default_lbl.setVisible(is_default)
        self._tpl_default_btn.setEnabled(not is_default)

    def _tpl_set_default(self):
        self._flush_current_tpl()
        row = self._tpl_list.currentRow()
        if row < 0 or row >= len(self._templates):
            return
        self._vault.set_active_template(self._templates[row]["name"])
        self._refresh_tpl_header()

    def _tpl_add(self):
        self._flush_current_tpl()

        # Ask: blank or copy from existing?
        dlg = QDialog(self)
        dlg.setWindowTitle("New Template")
        dlg.setFixedWidth(360)
        dlg.setModal(True)
        t = _t()
        dlg.setStyleSheet(f"QDialog{{background:{t.bg_window};}}"
                          f"QLabel{{color:{t.text_primary};background:transparent;}}"
                          f"QRadioButton{{color:{t.text_primary};font-size:12px;spacing:6px;}}"
                          f"QRadioButton::indicator{{width:14px;height:14px;border:1.5px solid {t.border};"
                          f"border-radius:7px;background:{t.bg_checkbox};}}"
                          f"QRadioButton::indicator:checked{{background:{t.accent};border-color:{t.accent};}}"
                          f"QComboBox{{border:1px solid {t.border};border-radius:5px;padding:4px 8px;"
                          f"font-size:12px;background:{t.bg_input};color:{t.text_primary};}}"
                          f"QComboBox::drop-down{{border:none;width:18px;}}"
                          f"QComboBox QAbstractItemView{{background:{t.bg_input};color:{t.text_primary};"
                          f"selection-background-color:{t.accent};}}")
        v = QVBoxLayout(dlg)
        v.setContentsMargins(20, 16, 20, 16)
        v.setSpacing(10)
        v.addWidget(_lbl("Create new template from:", 12, bold=True))

        rb_blank = QRadioButton("Start blank")
        rb_blank.setChecked(True)
        rb_copy  = QRadioButton("Copy from existing template:")
        v.addWidget(rb_blank)

        copy_row = QHBoxLayout()
        copy_row.addWidget(rb_copy)
        copy_combo = QComboBox()
        for tpl in self._templates:
            copy_combo.addItem(tpl["name"])
        copy_combo.setEnabled(False)
        rb_blank.toggled.connect(lambda checked: copy_combo.setEnabled(not checked))
        copy_row.addWidget(copy_combo, stretch=1)
        v.addLayout(copy_row)

        v.addSpacing(4)
        btn_row = QHBoxLayout()
        btn_row.addStretch()
        cancel_btn = _btn("Cancel", "secondary", height=32)
        cancel_btn.clicked.connect(dlg.reject)
        ok_btn = _btn("Create", "primary", height=32)
        ok_btn.clicked.connect(dlg.accept)
        btn_row.addWidget(cancel_btn)
        btn_row.addWidget(ok_btn)
        v.addLayout(btn_row)

        if dlg.exec() != QDialog.DialogCode.Accepted:
            return

        if rb_copy.isChecked():
            src_name = copy_combo.currentText()
            src = next((t for t in self._templates if t["name"] == src_name), None)
            new_tpl = {
                "name":    f"Copy of {src_name}",
                "subject": src.get("subject", "") if src else "",
                "body":    src.get("body", "") if src else "",
                "docs":    dict(src.get("docs", {})) if src else {k: True for k in self._doc_cbs},
            }
        else:
            new_tpl = {
                "name":    "New Template",
                "subject": "",
                "body":    "",
                "docs":    {k: True for k in self._doc_cbs},
            }

        self._templates.append(new_tpl)
        from PyQt6.QtWidgets import QListWidgetItem
        item = QListWidgetItem(new_tpl["name"])
        item.setFlags(item.flags() | Qt.ItemFlag.ItemIsEditable)
        self._tpl_list.addItem(item)
        new_row = len(self._templates) - 1
        self._tpl_list.setCurrentRow(new_row)
        self._tpl_list.editItem(self._tpl_list.item(new_row))
        self._tpl_del_btn.setEnabled(True)

    def _tpl_delete(self):
        row = self._tpl_list.currentRow()
        if row < 0 or len(self._templates) <= 1:
            return
        self._templates.pop(row)
        self._tpl_list.takeItem(row)
        new_row = min(row, len(self._templates) - 1)
        self._tpl_list.setCurrentRow(new_row)
        self._tpl_del_btn.setEnabled(len(self._templates) > 1)

    def _export_settings(self):
        """Export SMTP settings and/or selected templates to a JSON file."""
        tpl_names = [t["name"] for t in self._templates]
        dlg = _SelectiveExportDialog(self, tpl_names)
        if dlg.exec() != QDialog.DialogCode.Accepted:
            return
        include_smtp, include_tpls, selected_tpls = dlg.result_choices()

        data: dict = {}

        if include_smtp:
            cfg = self._collect()
            raw_pwd = cfg.pop("smtp_password", "") or ""
            if raw_pwd:
                try:
                    from cryptography.fernet import Fernet
                    from cryptography.hazmat.primitives.kdf.pbkdf2 import PBKDF2HMAC
                    from cryptography.hazmat.primitives import hashes
                    from base64 import urlsafe_b64encode
                    kdf = PBKDF2HMAC(algorithm=hashes.SHA256(), length=32, salt=b"AayDocCapio-export-v1", iterations=100000)
                    key = urlsafe_b64encode(kdf.derive(b"AayDocCapio-portable-export-key"))
                    cfg["smtp_password_exported"] = Fernet(key).encrypt(raw_pwd.encode()).decode()
                except Exception:
                    pass
            data["smtp"] = cfg

        if include_tpls:
            tpls_to_export = [t for t in self._templates if t["name"] in selected_tpls]
            data["templates"] = tpls_to_export
            active = self._vault.get_active_template_name()
            if active in selected_tpls:
                data["active_template"] = active

        path, _ = QFileDialog.getSaveFileName(
            self, "Export Email Settings", "AayDocCapio_EmailSettings.json",
            "JSON Files (*.json)")
        if not path:
            return
        try:
            with open(path, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
            QMessageBox.information(self, "Exported", "Settings exported successfully.")
        except Exception as e:
            QMessageBox.warning(self, "Export Failed", str(e))

    def _import_settings(self):
        """Import SMTP settings and/or templates from a JSON file (selective)."""
        path, _ = QFileDialog.getOpenFileName(
            self, "Import Email Settings", "", "JSON Files (*.json)")
        if not path:
            return
        try:
            with open(path, "r", encoding="utf-8") as f:
                data = json.load(f)
        except Exception as e:
            QMessageBox.warning(self, "Import Failed", f"Could not read file:\n{e}")
            return

        has_smtp = bool(data.get("smtp"))
        file_tpls = data.get("templates", [])
        file_tpl_names = [t["name"] for t in file_tpls]

        dlg = _SelectiveImportDialog(self, has_smtp, file_tpl_names)
        if dlg.exec() != QDialog.DialogCode.Accepted:
            return
        import_smtp, import_tpls, selected_tpls = dlg.result_choices()

        if import_smtp and has_smtp:
            smtp = data["smtp"]
            if smtp.get("smtp_host"): self._host.setText(smtp["smtp_host"])
            if smtp.get("smtp_port"): self._port.setValue(int(smtp["smtp_port"]))
            if smtp.get("smtp_user"): self._user.setText(smtp["smtp_user"])
            if smtp.get("smtp_from"): self._from.setText(smtp["smtp_from"])
            if smtp.get("bcc_addresses"): self._bcc.setText(smtp["bcc_addresses"])
            if smtp.get("firm_name"): self._firm.setText(smtp["firm_name"])
            enc = smtp.get("smtp_encryption", "")
            if enc:
                idx = self._enc.findText(enc)
                if idx >= 0: self._enc.setCurrentIndex(idx)
            if smtp.get("smtp_password_exported"):
                try:
                    from cryptography.fernet import Fernet
                    from cryptography.hazmat.primitives.kdf.pbkdf2 import PBKDF2HMAC
                    from cryptography.hazmat.primitives import hashes
                    from base64 import urlsafe_b64encode
                    kdf = PBKDF2HMAC(algorithm=hashes.SHA256(), length=32, salt=b"AayDocCapio-export-v1", iterations=100000)
                    key = urlsafe_b64encode(kdf.derive(b"AayDocCapio-portable-export-key"))
                    pwd = Fernet(key).decrypt(smtp["smtp_password_exported"].encode()).decode()
                    self._pwd.setText(pwd)
                except Exception:
                    pass

        if import_tpls and file_tpls:
            from PyQt6.QtWidgets import QListWidgetItem
            tpls_to_import = [t for t in file_tpls if t["name"] in selected_tpls]
            existing_names = {t["name"] for t in self._templates}
            for tpl in tpls_to_import:
                if tpl["name"] in existing_names:
                    # overwrite existing template with same name
                    for i, et in enumerate(self._templates):
                        if et["name"] == tpl["name"]:
                            self._templates[i] = tpl
                            break
                else:
                    self._templates.append(tpl)
            self._tpl_list.clear()
            for tpl in self._templates:
                item = QListWidgetItem(tpl["name"])
                item.setFlags(item.flags() | Qt.ItemFlag.ItemIsEditable)
                self._tpl_list.addItem(item)
            active = data.get("active_template", "")
            matched = False
            for i, tpl in enumerate(self._templates):
                if tpl["name"] == active:
                    self._tpl_list.setCurrentRow(i)
                    matched = True
                    break
            if not matched:
                self._tpl_list.setCurrentRow(0)

        QMessageBox.information(self, "Imported",
            "Settings imported successfully. Please verify and click Save.")

    def _mark_dirty(self):
        """Switch the Cancel button to Save & Close once any change is made."""
        self._save_close_btn.setText("Save && Close")
        self._save_close_btn.setProperty("style_variant", "primary")
        # Re-apply stylesheet via _btn helper logic — simplest is to swap icon/style directly
        t = _t()
        self._save_close_btn.setStyleSheet(
            f"QPushButton{{background:{t.accent};color:#FFFFFF;border:none;"
            f"border-radius:8px;font-size:13px;font-weight:600;padding:0 18px;}}"
            f"QPushButton:hover{{background:{t.accent_hover};}}")

    def _on_save_close_clicked(self):
        if self._save_close_btn.text() == "Cancel":
            self.reject()
        else:
            self._save()

    def _save_smtp_only(self):
        """Save SMTP/Sender settings without closing."""
        cfg = self._collect()
        if not cfg["smtp_host"]:
            QMessageBox.warning(self, "Missing Field", "Please enter an SMTP server address.")
            return
        if not cfg["smtp_user"]:
            QMessageBox.warning(self, "Missing Field", "Please enter a username / email.")
            return
        self._vault.save_email_settings(cfg)
        QMessageBox.information(self, "Saved", "SMTP settings saved.")

    def _save_templates_only(self):
        """Flush and persist templates without closing the dialog."""
        self._flush_current_tpl()
        names = [t["name"] for t in self._templates]
        if len(names) != len(set(names)):
            QMessageBox.warning(self, "Duplicate Template Name",
                                "Each template must have a unique name.")
            return
        self._vault.save_email_templates(self._templates)
        row = self._tpl_list.currentRow()
        if 0 <= row < len(self._templates):
            self._vault.set_active_template(self._templates[row]["name"])
        QMessageBox.information(self, "Saved", "Templates saved successfully.")

    # ── collect / save / test ─────────────────────────────────────────────────

    def _collect(self) -> dict:
        enc = self._enc.currentText()
        return {
            "smtp_host":       self._host.text().strip(),
            "smtp_port":       str(self._port.value()),
            "smtp_user":       self._user.text().strip(),
            "smtp_from":       self._from.text().strip(),
            "smtp_password":   self._pwd.text(),
            "smtp_encryption": enc,
            "smtp_use_tls":    enc == "STARTTLS",
            "firm_name":       self._firm.text().strip(),
            "bcc_addresses":   self._bcc.text().strip(),
        }

    def _save(self):
        cfg = self._collect()
        if not cfg["smtp_host"]:
            QMessageBox.warning(self, "Missing Field", "Please enter an SMTP server address.")
            return
        if not cfg["smtp_user"]:
            QMessageBox.warning(self, "Missing Field", "Please enter a username / email.")
            return
        # Flush current template editor into list
        self._flush_current_tpl()
        # Validate template names are unique and non-empty
        names = [t["name"] for t in self._templates]
        if len(names) != len(set(names)):
            QMessageBox.warning(self, "Duplicate Template Name",
                                "Each template must have a unique name.")
            return
        # Save SMTP settings
        # Also sync active template's subject/body into legacy keys for emailer compat
        active = self._templates[self._tpl_list.currentRow()] if self._templates else {}
        cfg["email_subject_tpl"] = active.get("subject", "")
        cfg["email_body_tpl"]    = active.get("body", "")
        self._vault.save_email_settings(cfg)
        self._vault.save_email_templates(self._templates)
        active_name = active.get("name", "")
        if active_name:
            self._vault.set_active_template(active_name)
        self.accept()

    def _send_test(self):
        cfg = self._collect()
        to = cfg["smtp_user"]
        if not cfg["smtp_host"] or not to:
            QMessageBox.warning(self, "Incomplete",
                                "Fill in the server address and username before sending a test.")
            return

        # Disable button and show spinner while connecting
        self._test_btn.setEnabled(False)
        self._test_btn.setText("⏳ Connecting…")

        def _worker():
            from automation.emailer import send_email
            try:
                send_email(cfg, to,
                           subject="AayDocCapio — Test Email",
                           body="This is a test email from AayDocCapio.\n\n"
                                "If you received this, your SMTP settings are working correctly.",
                           attachments=[])
                self._test_result.emit(True, f"Test email sent to {to}.\nCheck your inbox.")
            except Exception as e:
                from automation.emailer import friendly_smtp_error
                self._test_result.emit(False, friendly_smtp_error(e))

        threading.Thread(target=_worker, daemon=True).start()

    def _on_test_result(self, success: bool, message: str):
        self._test_btn.setEnabled(True)
        self._test_btn.setText("Send Test Email")
        if success:
            QMessageBox.information(self, "Test Sent", message)
        else:
            QMessageBox.critical(self, "Send Failed", message)

    def _show_help(self):
        from ui.smtp_help import _write_smtp_help_html
        import webbrowser
        path = _write_smtp_help_html()
        webbrowser.open("file:///" + path.replace(os.sep, "/"))

    def _open_log(self):
        from automation.emailer import _email_log_path
        path = _email_log_path()
        if not os.path.exists(path):
            QMessageBox.information(self, "No Log Yet",
                "No email activity has been logged yet.\n"
                "Send a test email first, then check the log.")
            return
        EmailLogDialog(self, path).exec()



# ── Email Log Dialog ──────────────────────────────────────────────────────────

class EmailLogDialog(QDialog):
    """Shows email_log.txt with a refresh button and option to clear."""

    def __init__(self, parent, log_path: str):
        super().__init__(parent)
        self._path = log_path
        self.setWindowTitle("Email Activity Log")
        self.setModal(True)
        self.resize(700, 500)
        self._build_ui()

    def _build_ui(self):
        t = _t()
        self.setStyleSheet(f"QDialog{{background:{t.bg_window};}}")

        outer = QVBoxLayout(self)
        outer.setContentsMargins(0, 0, 0, 0)
        outer.setSpacing(0)

        # Title bar
        title_bar = QWidget()
        title_bar.setStyleSheet(
            f"QWidget{{background:{t.bg_table_alt};border-bottom:1px solid {t.border};}}")
        tb = QHBoxLayout(title_bar)
        tb.setContentsMargins(20, 10, 16, 10)
        tb.addWidget(_lbl("Email Activity Log", 13, bold=True))
        tb.addSpacing(8)
        tb.addWidget(_lbl(self._path, 10, color=t.text_muted))
        tb.addStretch()
        outer.addWidget(title_bar)

        # Log text area
        self._text = QTextEdit()
        self._text.setReadOnly(True)
        self._text.setFont(QFont(_MONO_FONT, 10))
        self._text.setStyleSheet(
            f"QTextEdit{{background:{t.bg_input};color:{t.text_primary};"
            f"border:none;padding:12px 16px;}}")
        outer.addWidget(self._text, stretch=1)
        self._reload()

        # Footer
        sep = QFrame(); sep.setFrameShape(QFrame.Shape.HLine)
        sep.setStyleSheet(f"background:{t.border};border:none;max-height:1px;")
        outer.addWidget(sep)

        footer = QWidget()
        footer.setStyleSheet(f"QWidget{{background:{t.bg_table_alt};}}")
        ft = QHBoxLayout(footer)
        ft.setContentsMargins(16, 10, 16, 10)
        ft.setSpacing(8)
        refresh_btn = _btn("Refresh", "outline", height=34, icon="btn_refresh.png")
        refresh_btn.clicked.connect(self._reload)
        ft.addWidget(refresh_btn)
        clear_btn = _btn("Clear Log", "danger", height=34, icon="btn_clear_log.png")
        clear_btn.clicked.connect(self._clear)
        ft.addWidget(clear_btn)
        ft.addStretch()
        close_btn = _btn("Close", "primary", height=34, icon="btn_close.png")
        close_btn.clicked.connect(self.accept)
        ft.addWidget(close_btn)
        outer.addWidget(footer)

    def _reload(self):
        try:
            with open(self._path, "r", encoding="utf-8", errors="replace") as f:
                content = f.read()
        except Exception as e:
            content = f"Could not read log: {e}"
        self._text.setPlainText(content)
        # Scroll to bottom so latest entry is visible
        self._text.moveCursor(self._text.textCursor().MoveOperation.End)

    def _clear(self):
        from PyQt6.QtWidgets import QMessageBox
        if QMessageBox.question(self, "Clear Log",
                "Delete all email log entries?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
                ) == QMessageBox.StandardButton.Yes:
            try:
                open(self._path, "w").close()
            except Exception:
                pass
            self._text.clear()


# ── Generate Tax Challans Dialog (F-64) ────────────────────────────────────────

class ChallanRowDetailDialog(QDialog):
    """
    Full-detail editor for ONE GenerateChallansDialog row — PAN, Payment
    Mode, Bank/Sub-Mode (options depend on Mode), and the Tax/Surcharge/
    Cess/Interest/Penalty/Others breakup. Opened by double-clicking a
    summary-table row, or by "+ Add Row" for a new one — the summary table
    itself only ever shows PAN/Name/Total/Mode, per the user's request to
    keep the main view scannable and push full editing into its own dialog.
    """
    def __init__(self, parent, vault, row_data=None):
        super().__init__(parent)
        from automation.challan_generator import PAYMENT_MODES, DEFAULT_PAYMENT_MODE, cash_limit_exceeded
        from automation.challan_fields import CHALLAN_AMOUNT_FIELDS

        self._vault = vault
        self._payment_modes = PAYMENT_MODES
        self._cash_limit_exceeded = cash_limit_exceeded
        self._amount_keys = CHALLAN_AMOUNT_FIELDS
        row_data = row_data or {}
        self.result_data = None

        self.setWindowTitle("Challan Details")
        self.setMinimumWidth(440)
        _bt = _t()
        self.setStyleSheet(f"QDialog{{background:{_bt.bg_window};}}")

        layout = QVBoxLayout(self)
        layout.setContentsMargins(18, 16, 18, 14)
        layout.setSpacing(10)

        # ── PAN (pick from client list, or type a new one) ─────────────
        layout.addWidget(_lbl("PAN"))
        self._pan_combo = QComboBox()
        self._pan_combo.setEditable(True)
        self._pan_combo.setInsertPolicy(QComboBox.InsertPolicy.NoInsert)
        for a in sorted(vault.get_all_assessees(), key=lambda a: a.get("name", "")):
            self._pan_combo.addItem(f"{a.get('pan','')} — {a.get('name','')}", a.get("pan", ""))
        # Qt's auto-created completer for an editable combo defaults to
        # "starts with" matching against the item text — since items read
        # "PAN — Name", typing a name (which sits after the PAN prefix)
        # never matched. Switch it to "contains", case-insensitive, so
        # searching by either PAN or name works.
        completer = self._pan_combo.completer()
        if completer:
            completer.setFilterMode(Qt.MatchFlag.MatchContains)
            completer.setCaseSensitivity(Qt.CaseSensitivity.CaseInsensitive)
        # Qt's default popup only shows 10 items before scrolling, which on
        # a 33-bank list reads as "incomplete" even though every item is
        # actually there — raise the visible count so most lists fit
        # without scrolling.
        self._pan_combo.setMaxVisibleItems(18)
        self._pan_combo.setCurrentText(row_data.get("pan", ""))
        # setCurrentText()/clicking a dropdown item fires currentTextChanged,
        # not editTextChanged (that one's typing-only) — connect both, since
        # relying on editTextChanged alone silently misses item selection.
        self._pan_combo.editTextChanged.connect(self._on_pan_changed)
        self._pan_combo.currentTextChanged.connect(self._on_pan_changed)
        layout.addWidget(self._pan_combo)

        self._name_label = QLabel("")
        self._name_label.setStyleSheet(f"color:{_bt.text_muted};font-size:11px;background:transparent;")
        layout.addWidget(self._name_label)

        # ── Payment Mode + Bank/Sub-Mode ─────────────────────────────────
        mode_row = QHBoxLayout()
        mode_col = QVBoxLayout()
        mode_col.addWidget(_lbl("Payment Mode"))
        self._mode_combo = QComboBox()
        self._mode_combo.addItems(list(self._payment_modes.keys()))
        self._mode_combo.setCurrentText(row_data.get("payment_mode") or DEFAULT_PAYMENT_MODE)
        self._mode_combo.currentTextChanged.connect(self._on_mode_changed)
        mode_col.addWidget(self._mode_combo)
        mode_row.addLayout(mode_col)

        # One flat, searchable picklist per the user's direction — no
        # visible "Other Bank" tier. Populated from
        # automation.challan_generator.all_bank_options(mode) (primary
        # on-screen tiles + a best-effort extended list); the field stays
        # editable so a bank not in that list can still be typed directly.
        # generate_challan() decides on its own, per bank name, whether it's
        # a primary tile (clicked directly) or needs the portal's own
        # "Other Bank" nested search — the user never needs to know which.
        bank_col = QVBoxLayout()
        bank_col.addWidget(_lbl("Bank / Sub-Mode"))
        self._bank_combo = QComboBox()
        self._bank_combo.setEditable(True)
        self._bank_combo.setInsertPolicy(QComboBox.InsertPolicy.NoInsert)
        bank_completer = self._bank_combo.completer()
        if bank_completer:
            bank_completer.setFilterMode(Qt.MatchFlag.MatchContains)
            bank_completer.setCaseSensitivity(Qt.CaseSensitivity.CaseInsensitive)
        self._bank_combo.setMaxVisibleItems(18)
        # Same reasoning as the PAN combo above — connect both signals so
        # picking a bank from the dropdown (not just typing one) reliably
        # updates the cash-cap warning and the Drawn on Bank field's
        # visibility.
        self._bank_combo.editTextChanged.connect(self._on_bank_changed)
        self._bank_combo.currentTextChanged.connect(self._on_bank_changed)
        bank_col.addWidget(self._bank_combo)
        mode_row.addLayout(bank_col)
        layout.addLayout(mode_row)

        # Confirmed against a real sample PDF ("Drawn on Bank: Kotak
        # Mahindra Bank"): Pay at Bank Counter's Cheque/Demand Draft
        # sub-modes carry their own separate bank field for which bank the
        # cheque/DD itself is drawn on — Cash has no such field, and no
        # other Payment Mode has one either. Reuses the same full bank list
        # as Net Banking (any real bank could plausibly issue a cheque),
        # editable/searchable the same way.
        drawee_row = QHBoxLayout()
        drawee_label = QLabel("Drawn on Bank")
        drawee_label.setFixedWidth(120)
        drawee_label.setStyleSheet(f"color:{_bt.text_primary};background:transparent;")
        drawee_row.addWidget(drawee_label)
        self._drawee_combo = QComboBox()
        self._drawee_combo.setEditable(True)
        self._drawee_combo.setInsertPolicy(QComboBox.InsertPolicy.NoInsert)
        from automation.challan_generator import all_bank_options as _all_bank_options
        self._drawee_combo.addItems(_all_bank_options("Net Banking"))
        drawee_completer = self._drawee_combo.completer()
        if drawee_completer:
            drawee_completer.setFilterMode(Qt.MatchFlag.MatchContains)
            drawee_completer.setCaseSensitivity(Qt.CaseSensitivity.CaseInsensitive)
        self._drawee_combo.setMaxVisibleItems(18)
        self._drawee_combo.setCurrentText(row_data.get("drawee_bank", ""))
        drawee_row.addWidget(self._drawee_combo)
        self._drawee_widget = QWidget()
        self._drawee_widget.setLayout(drawee_row)
        self._drawee_widget.setVisible(False)
        layout.addWidget(self._drawee_widget)

        self._cash_warning = QLabel("")
        self._cash_warning.setStyleSheet(f"color:{getattr(_bt, 'warning', '#D97706')};font-size:11px;background:transparent;")
        self._cash_warning.setWordWrap(True)
        layout.addWidget(self._cash_warning)

        # ── Amount breakup ───────────────────────────────────────────────
        layout.addWidget(_lbl("Amount Breakup (₹)"))
        self._amount_edits = {}
        amounts_grid = QVBoxLayout()
        amounts_grid.setSpacing(4)
        for key, label, kind in [
            ("tax", "Tax", "amount"), ("surcharge", "Surcharge", "amount"),
            ("cess", "Cess", "amount"), ("interest", "Interest", "amount"),
            ("penalty", "Penalty", "amount"), ("others", "Others", "amount"),
        ]:
            row = QHBoxLayout()
            lbl = QLabel(label)
            lbl.setFixedWidth(90)
            lbl.setStyleSheet(f"color:{_bt.text_primary};background:transparent;")
            row.addWidget(lbl)
            edit = QLineEdit(str(row_data.get(key, "") or ""))
            edit.textChanged.connect(self._update_total)
            self._amount_edits[key] = edit
            row.addWidget(edit)
            amounts_grid.addLayout(row)
        layout.addLayout(amounts_grid)

        total_row = QHBoxLayout()
        total_row.addWidget(_lbl("Total", bold=True))
        self._total_label = QLabel("0")
        self._total_label.setStyleSheet(f"color:{_bt.text_primary};font-weight:bold;background:transparent;")
        total_row.addWidget(self._total_label)
        total_row.addStretch(1)
        layout.addLayout(total_row)

        # ── OK / Cancel ──────────────────────────────────────────────────
        btn_row = QHBoxLayout()
        btn_row.addStretch(1)
        btn_cancel = _btn("Cancel", "outline")
        btn_cancel.clicked.connect(self.reject)
        btn_row.addWidget(btn_cancel)
        self._btn_ok = _btn("OK", "primary")
        self._btn_ok.clicked.connect(self._on_ok)
        btn_row.addWidget(self._btn_ok)
        layout.addLayout(btn_row)

        self._on_pan_changed()
        self._on_mode_changed(select_bank=row_data.get("bank", ""))
        self._update_total()

    def _current_pan(self) -> str:
        idx = self._pan_combo.currentIndex()
        text = self._pan_combo.currentText().strip()
        if idx >= 0 and self._pan_combo.itemText(idx) == text:
            return (self._pan_combo.itemData(idx) or "").upper()
        if " — " in text:
            return text.split(" — ")[0].strip().upper()
        return text.upper()

    def _on_pan_changed(self):
        pan = self._current_pan()
        matched = next((a for a in self._vault.get_all_assessees() if a.get("pan", "").upper() == pan), None)
        _bt = _t()
        if pan and not matched:
            self._name_label.setText("⚠ Unknown PAN — not in Client Master")
            self._name_label.setStyleSheet(f"color:{getattr(_bt, 'warning', '#D97706')};font-size:11px;background:transparent;")
        elif matched:
            self._name_label.setText(matched.get("name", ""))
            self._name_label.setStyleSheet(f"color:{_bt.text_muted};font-size:11px;background:transparent;")
        else:
            self._name_label.setText("")

    def _on_mode_changed(self, *_args, select_bank=""):
        from automation.challan_generator import all_bank_options
        mode = self._mode_combo.currentText()
        options = all_bank_options(mode)
        self._bank_combo.blockSignals(True)
        self._bank_combo.clear()
        if options:
            self._bank_combo.addItems(options)
            self._bank_combo.setEnabled(True)
            # BUG FIX (2026-09-03): confirmed live — since this combo is
            # editable, setCurrentText() will happily display ANY string
            # here, valid option or not. Opening an existing row whose
            # stored Bank / Sub-Mode no longer matches its Payment Mode
            # (e.g. imported before a validation fix, or edited to a
            # different mode) silently pre-filled the stale value instead
            # of showing it was actually invalid. Only pre-fill it when
            # it's genuinely one of this mode's own options.
            self._bank_combo.setCurrentText(select_bank if select_bank in options else "")
        else:
            self._bank_combo.addItem("(not required for this mode)")
            self._bank_combo.setEnabled(False)
        self._bank_combo.blockSignals(False)
        self._on_bank_changed()

    def _on_bank_changed(self):
        mode = self._mode_combo.currentText()
        bank = self._current_bank()
        # BUG FIX (2026-09-02): confirmed live — the portal's "Select Bank
        # (authorised Banks only)" dropdown is mandatory for EVERY Pay at
        # Bank Counter sub-mode, Cash included, not just Cheque/Demand
        # Draft as originally assumed. A live run with Cash selected had
        # nowhere in this dialog to enter that bank, so the automation had
        # no value to fill and Continue stayed disabled on the portal.
        self._drawee_widget.setVisible(mode == "Pay at Bank Counter")
        self._update_cash_warning()

    def _current_bank(self) -> str:
        if not self._bank_combo.isEnabled():
            return ""
        return self._bank_combo.currentText().strip()

    def _current_total(self) -> float:
        total = 0.0
        for key, edit in self._amount_edits.items():
            try:
                total += float(edit.text()) if edit.text().strip() else 0.0
            except ValueError:
                pass
        return total

    def _update_total(self):
        self._total_label.setText(f"{self._current_total():g}")
        self._update_cash_warning()

    def _update_cash_warning(self):
        mode = self._mode_combo.currentText()
        bank = self._current_bank()
        total = self._current_total()
        if self._cash_limit_exceeded(mode, bank, total):
            self._cash_warning.setText(
                f"⚠ Pay at Bank Counter / Cash is capped at ₹10,000 (RBI rule) — this total is "
                f"₹{total:,.0f}. Use Cheque or Demand Draft instead."
            )
            self._btn_ok.setEnabled(False)
        else:
            self._cash_warning.setText("")
            self._btn_ok.setEnabled(True)

    def _on_ok(self):
        pan = self._current_pan()
        if not pan:
            QMessageBox.warning(self, "PAN Required", "Please enter or select a PAN.")
            return
        # BUG FIX (2026-09-03): confirmed live — the Bank / Sub-Mode combo
        # is editable, so it will happily accept and save a value that
        # isn't actually valid for the selected Payment Mode (e.g. "Cash"
        # under Payment Gateway) without any complaint. Block on the same
        # check used everywhere else (table warnings, import) instead of
        # only catching this in the table after the fact.
        from automation.challan_generator import bank_problem, drawee_bank_problem
        mode = self._mode_combo.currentText()
        problem = bank_problem(mode, self._current_bank())
        if problem:
            QMessageBox.warning(self, "Bank / Sub-Mode", problem.capitalize() + ".")
            return
        # BUG FIX (2026-09-03): confirmed live — Drawn on Bank is mandatory
        # for every Pay at Bank Counter sub-mode (Cash included), but
        # nothing stopped OK from being clicked while it was left blank.
        drawee_bank = self._drawee_combo.currentText().strip() if self._drawee_widget.isVisible() else ""
        drawee_problem = drawee_bank_problem(mode, drawee_bank)
        if drawee_problem:
            QMessageBox.warning(self, "Drawn on Bank", drawee_problem.capitalize() + ".")
            return
        amounts = {}
        for key, edit in self._amount_edits.items():
            text = edit.text().strip()
            try:
                amounts[key] = float(text) if text else 0
            except ValueError:
                QMessageBox.warning(self, "Invalid Amount", f"'{text}' is not a valid amount.")
                return
        self.result_data = {
            "pan": pan,
            "payment_mode": mode,
            "bank": self._current_bank(),
            "drawee_bank": drawee_bank,
            **amounts,
        }
        self.accept()



def challan_instructions_text() -> str:
    """Plain-text version of the Instructions sheet, for the CSV export
    path (CSV has no second sheet to carry it)."""
    lines = [
        "Bulk Tax Challan — Import Template",
        "AayDoc Capio™  ·  © 2026  ·  Developed by CA. Deepak Bhholusaria  ·  "
        "linkedin.com/in/bhholusaria  ·  deepak@ailearrning.guru",
        "",
        "HOW TO FILL IN THIS TEMPLATE",
        "=" * 29,
        "",
        "One row per client. If you're using the Excel version instead, don't "
        'edit the hidden "Lists" sheet — it just powers the dropdowns on the '
        '"Challans" sheet.',
        "",
        "PAN",
        "  Must already be saved in AayDocCapio's Client Master. If it isn't, "
        "you'll see a warning when you import this file — the row still comes "
        "in, but it won't run until you fix the PAN.",
        "",
        "Name",
        "  Just for you to see whose row is whose. The app ignores this column "
        "when importing — it only looks at the PAN.",
        "",
        "Payment Mode",
        "  Choose one: Net Banking, Debit Card, Pay at Bank Counter, RTGS/NEFT, "
        "or Payment Gateway including UPI and Credit Card.",
        "",
        "Bank / Sub-Mode",
        "  What you can pick here depends on the Payment Mode you chose. "
        "RTGS/NEFT doesn't need a bank at all — leave this blank for those rows. "
        "If you change the Payment Mode after already picking a bank, clear and "
        "re-pick this cell too — it won't clear itself.",
        "",
        "Drawn on Bank",
        "  Only needed for Pay at Bank Counter (Cash, Cheque, or Demand Draft) — "
        "the bank the payment is made at. Leave this blank for every other "
        "Payment Mode.",
        "",
        "Tax / Surcharge / Cess / Interest / Penalty / Others",
        "  Most of the time you'll only have one total amount — put it all in "
        "Tax and leave the rest as 0. Only split it across the other columns if "
        "you actually have a breakup, e.g. interest calculated separately.",
        "",
        "Note: if a row's Payment Mode doesn't need a Bank / Sub-Mode or Drawn "
        "on Bank, leave that cell blank. Don't put anything else in it.",
    ]
    return "\n".join(lines)



def write_challan_table_file(path, headers, rows):
    if path.endswith(".csv"):
        import csv
        import re
        with open(path, "w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            w.writerow(headers)
            w.writerows(rows)
        # CSV has no second sheet to carry the fill-in instructions, so
        # they'd otherwise vanish entirely for anyone who picks CSV over
        # Excel — write them as a plain-text sibling file instead.
        instructions_path = re.sub(r"\.csv$", "", path, flags=re.IGNORECASE) + "_Instructions.txt"
        with open(instructions_path, "w", encoding="utf-8") as f:
            f.write(challan_instructions_text())
        return

    import re
    from openpyxl import Workbook
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.formatting.rule import FormulaRule
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.workbook.defined_name import DefinedName
    from automation.challan_generator import PAYMENT_MODES, all_bank_options

    wb = Workbook()
    ws = wb.active
    ws.title = "Challans"
    ws.append(headers)
    for row in rows:
        ws.append(row)

    # Column positions are found by label text, not assumed fixed
    # indices — headers here always start "PAN", "Name", ... but the
    # rest can shift if CHALLAN_INPUT_COLUMNS changes shape later.
    def _col(label):
        return get_column_letter(headers.index(label) + 1)
    col_mode = _col("Payment Mode")
    col_bank = _col("Bank / Sub-Mode")
    col_drawee = _col("Drawn on Bank")
    last_row = max(len(rows), 1) + 200  # headroom for rows added later in Excel

    # ── Format the data itself as a real Excel Table ─────────────────
    # A plain grid of unstyled cells is what the user's screenshot
    # showed — banded rows, a styled header, filter arrows, and
    # sized-to-content columns make it read as an actual data sheet
    # rather than a raw CSV pasted into Excel.
    last_col_letter = get_column_letter(len(headers))
    # Confirmed live — a table sized to exactly the written rows (often
    # just the one sample row) forces the user through Table Design >
    # Resize Table before they can type a real batch in. Pad the table
    # itself out to a minimum row count so it's already sized for bulk
    # entry; the 200-row headroom on `last_row` above covers validation/
    # dropdowns further down still, past even this padded table.
    MIN_TEMPLATE_ROWS = 50
    table_last_row = max(len(rows), MIN_TEMPLATE_ROWS) + 1
    tab = Table(displayName="Challans", ref=f"A1:{last_col_letter}{table_last_row}")
    tab.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False,
        showRowStripes=True, showColumnStripes=False,
    )
    ws.add_table(tab)
    ws.freeze_panes = "A2"
    for i, label in enumerate(headers, start=1):
        width = max(12, min(30, len(label) + 4))
        ws.column_dimensions[get_column_letter(i)].width = width
    for row_cells in ws.iter_rows(min_row=2, max_row=max(table_last_row, 2), max_col=len(headers)):
        for cell in row_cells:
            cell.alignment = Alignment(vertical="center")

    # ── Hidden "Lists" sheet backing every dropdown ──────────────────
    # Excel's in-cell list validation (formula1='"A,B,C"') caps at 255
    # characters — nowhere near enough for a 33-bank list — so every
    # allowed-values set lives in real cells here, referenced by range.
    modes = list(PAYMENT_MODES.keys())

    def _sanitize(text: str) -> str:
        s = re.sub(r"[^A-Za-z0-9_]", "_", text)
        return s if s and (s[0].isalpha() or s[0] == "_") else f"_{s}"

    ws_lists = wb.create_sheet("Lists")
    ws_lists["A1"] = "Payment Mode"
    mode_range_names = {}
    for i, m in enumerate(modes, start=2):
        ws_lists.cell(row=i, column=1, value=m)
        mode_range_names[m] = f"Mode_{_sanitize(m)}"

    # One column per mode for its own Bank / Sub-Mode options — column 3
    # (C) onward, in the same order as `modes`. BUG FIX (2026-09-03):
    # confirmed live (a real template screenshot) — modes with no
    # picklist at all (RTGS/NEFT; see automation/challan_generator.py's
    # PAYMENT_MODES) used to get a single descriptive placeholder value
    # ("(not required for this mode)") as their only dropdown option,
    # which a user could still select and end up with misleading
    # non-blank text in a field the portal never shows a picklist for.
    # A single BLANK cell instead — Excel's list validation still needs
    # a real range to avoid an #REF! error, but a blank source means
    # blank is the only thing selectable from the dropdown, and (now
    # that showErrorMessage/errorStyle are wired below) any other typed
    # value is rejected outright rather than silently accepted.
    for m_idx, m in enumerate(modes):
        col_num = 3 + m_idx
        col_letter = get_column_letter(col_num)
        options = all_bank_options(m) or [""]
        ws_lists.cell(row=1, column=col_num, value=m)
        for i, opt in enumerate(options, start=2):
            ws_lists.cell(row=i, column=col_num, value=opt)
        range_ref = f"Lists!${col_letter}$2:${col_letter}${len(options) + 1}"
        wb.defined_names[mode_range_names[m]] = DefinedName(mode_range_names[m], attr_text=range_ref)

    # ── Drawn on Bank per-mode lists — mirrors the Bank / Sub-Mode ────
    # mechanism above, but keyed on whether the mode is "Pay at Bank
    # Counter" (the only mode this field ever applies to, confirmed
    # live against a real sample PDF's "Drawn on Bank" field) rather
    # than on that mode's own bank-tile list. Every other mode gets the
    # same single-blank-cell treatment as the no-picklist Bank /
    # Sub-Mode case above, for the same reason.
    drawee_full_options = all_bank_options("Net Banking")  # any real bank can plausibly issue a cheque/DD
    drawee_range_names = {}
    for m in modes:
        drawee_range_names[m] = f"Drawee_{_sanitize(m)}"
    drawee_col_start = 3 + len(modes)
    for m_idx, m in enumerate(modes):
        col_num = drawee_col_start + m_idx
        col_letter = get_column_letter(col_num)
        options = drawee_full_options if m == "Pay at Bank Counter" else [""]
        ws_lists.cell(row=1, column=col_num, value=f"Drawee for {m}")
        for i, opt in enumerate(options, start=2):
            ws_lists.cell(row=i, column=col_num, value=opt)
        range_ref = f"Lists!${col_letter}$2:${col_letter}${len(options) + 1}"
        wb.defined_names[drawee_range_names[m]] = DefinedName(drawee_range_names[m], attr_text=range_ref)

    ws_lists.sheet_state = "hidden"

    # ── Payment Mode (col B in the sheet, but located dynamically) ───
    dv_mode = DataValidation(
        type="list", formula1=f"Lists!$A$2:$A${len(modes) + 1}", allow_blank=True,
        showErrorMessage=True, errorStyle="stop",
    )
    dv_mode.errorTitle = "Invalid Payment Mode"
    dv_mode.error = "Please pick a Payment Mode from the dropdown."
    ws.add_data_validation(dv_mode)
    dv_mode.add(f"{col_mode}2:{col_mode}{last_row}")

    # ── Bank / Sub-Mode — CASCADING on Payment Mode ──────────────────
    # BUG FIX (2026-09-03): confirmed live (a real screenshot) — the
    # original approach used VLOOKUP against a small Lists!A:B helper
    # table to translate the mode name into its named-range name before
    # INDIRECT resolved it, and in real Excel that indirection somehow
    # produced the WRONG mode's list (Bank / Sub-Mode showing the
    # "Drawn on Bank" side's range names). Dropped the helper table
    # entirely — the range name is built directly from the mode cell's
    # own text with SUBSTITUTE (mirroring _sanitize() above: modes only
    # ever contain spaces and "/", both swapped for "_"), which is the
    # standard, more robust way to do a cascading Excel dropdown.
    # BUG FIX (2026-09-03, earlier): showErrorMessage/errorStyle were
    # never set, so this validation's errorTitle/error text was dead —
    # Excel accepted any typed value regardless of the dropdown's
    # contents. Now genuinely blocks ("stop") anything outside the
    # current mode's own list, including RTGS/NEFT's blank-only list.
    dv_bank = DataValidation(
        type="list",
        formula1=f'=INDIRECT("Mode_"&SUBSTITUTE(SUBSTITUTE(${col_mode}2," ","_"),"/","_"))',
        allow_blank=True, showErrorMessage=True, errorStyle="stop",
    )
    dv_bank.errorTitle = "Not Required / Invalid Bank"
    dv_bank.error = ("Please pick a Payment Mode first, then a matching Bank / Sub-Mode. "
                      "RTGS/NEFT has no Bank / Sub-Mode on the portal — leave this blank.")
    ws.add_data_validation(dv_bank)
    dv_bank.add(f"{col_bank}2:{col_bank}{last_row}")

    # ── Drawn on Bank — CASCADING the same way, now genuinely blank- ──
    # only outside "Pay at Bank Counter" instead of merely greyed out.
    # BUG FIX (2026-09-02, extended 2026-09-03): confirmed live — this
    # field is mandatory for EVERY Pay at Bank Counter sub-mode (Cash
    # included, not just Cheque/Demand Draft as originally assumed).
    dv_drawee = DataValidation(
        type="list",
        formula1=f'=INDIRECT("Drawee_"&SUBSTITUTE(SUBSTITUTE(${col_mode}2," ","_"),"/","_"))',
        allow_blank=True, showErrorMessage=True, errorStyle="stop",
    )
    dv_drawee.errorTitle = "Not Required / Invalid Bank"
    dv_drawee.error = "Drawn on Bank only applies to Pay at Bank Counter — leave this blank for other Payment Modes."
    ws.add_data_validation(dv_drawee)
    dv_drawee.add(f"{col_drawee}2:{col_drawee}{last_row}")

    # ── Visual hint to match the blocking above: grey out both fields ─
    # on rows where they're not applicable, so it reads as "disabled"
    # rather than just "happens to reject your input".
    grey_fill = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")
    grey_font = Font(color="AAAAAA")
    no_bank_modes = [m for m in modes if not all_bank_options(m)]
    if no_bank_modes:
        bank_not_applicable_formula = "OR(" + ",".join(
            f'${col_mode}2="{m}"' for m in no_bank_modes) + ")"
        ws.conditional_formatting.add(
            f"{col_bank}2:{col_bank}{last_row}",
            FormulaRule(formula=[bank_not_applicable_formula], fill=grey_fill, font=grey_font),
        )
    drawee_not_applicable_formula = f'${col_mode}2<>"Pay at Bank Counter"'
    ws.conditional_formatting.add(
        f"{col_drawee}2:{col_drawee}{last_row}",
        FormulaRule(formula=[drawee_not_applicable_formula], fill=grey_fill, font=grey_font),
    )

    # ── "Instructions" sheet — placed first so it's what opens by ────
    # default, explaining the column-by-column expectations and the
    # grey/blank-only behavior above (which otherwise looks like an
    # unexplained restriction to anyone who hasn't read this code).
    # Laid out as a two-column field/description table with a banner
    # and a highlighted closing note, rather than a single wall of
    # left-aligned text — confirmed live (a screenshot) that the
    # original plain stacked-paragraph version read as cramped and
    # hard to scan.
    from openpyxl.styles import Border, Side

    ws_help = wb.create_sheet("Instructions", 0)
    ws_help.sheet_view.showGridLines = False
    ws_help.column_dimensions["A"].width = 28
    ws_help.column_dimensions["B"].width = 95

    # Same brand banner (title + credit subtitle, navy/white/grey) used
    # on every generated report — automation/ais_converter.py's
    # "General Info" cover sheet is the reference this matches, so the
    # template reads as the same product rather than an unbranded file.
    NAVY = "0A1628"
    GREY = "94A3B8"
    BANNER_FILL = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
    HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    BAND_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    NOTE_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    thin = Side(style="thin", color="BFBFBF")
    box_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    r = 1
    ws_help.merge_cells(f"A{r}:B{r}")
    title_cell = ws_help.cell(row=r, column=1, value="Bulk Tax Challan — Import Template")
    title_cell.font = Font(bold=True, size=13, color="FFFFFF")
    title_cell.fill = BANNER_FILL
    title_cell.alignment = Alignment(vertical="center", horizontal="center", indent=1)
    ws_help.row_dimensions[r].height = 28
    r += 1

    ws_help.merge_cells(f"A{r}:B{r}")
    credit_cell = ws_help.cell(
        row=r, column=1,
        value="AayDoc Capio™  ·  © 2026  ·  Developed by CA. Deepak Bhholusaria  ·  "
              "linkedin.com/in/bhholusaria  ·  deepak@ailearrning.guru",
    )
    credit_cell.font = Font(size=8, color=GREY)
    credit_cell.fill = BANNER_FILL
    credit_cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
    ws_help.row_dimensions[r].height = 18
    r += 2

    ws_help.merge_cells(f"A{r}:B{r}")
    heading_cell = ws_help.cell(row=r, column=1, value="How to fill in this template")
    heading_cell.font = Font(bold=True, size=13, color=NAVY)
    heading_cell.alignment = Alignment(vertical="center", indent=1)
    ws_help.row_dimensions[r].height = 22
    r += 1

    ws_help.merge_cells(f"A{r}:B{r}")
    intro_cell = ws_help.cell(
        row=r, column=1,
        value='One row per client. Don\'t edit the hidden "Lists" sheet — it just '
              'powers the dropdowns on the "Challans" sheet.',
    )
    intro_cell.font = Font(italic=True, color="595959")
    intro_cell.alignment = Alignment(wrap_text=True, vertical="top", indent=1)
    ws_help.row_dimensions[r].height = 28
    r += 2

    header_row = r
    for col, text in ((1, "Column"), (2, "What to enter")):
        c = ws_help.cell(row=header_row, column=col, value=text)
        c.font = Font(bold=True, color="1F4E78")
        c.fill = HEADER_FILL
        c.border = box_border
        c.alignment = Alignment(vertical="center", indent=1)
    r += 1

    sections = [
        ("PAN",
         "Must already be saved in AayDocCapio's Client Master. If it isn't, you'll "
         "see a warning when you import this file — the row still comes in, but it "
         "won't run until you fix the PAN."),
        ("Name",
         "Just for you to see whose row is whose. The app ignores this column when "
         "importing — it only looks at the PAN."),
        ("Payment Mode",
         "Choose one from the dropdown: Net Banking, Debit Card, Pay at Bank "
         "Counter, RTGS/NEFT, or Payment Gateway including UPI and Credit Card."),
        ("Bank / Sub-Mode",
         "What you can pick here depends on the Payment Mode you chose on that row. "
         "RTGS/NEFT doesn't need a bank at all, so for RTGS/NEFT rows this cell "
         "greys out — leave it empty. If you change the Payment Mode after already "
         "picking a Bank / Sub-Mode, clear and re-pick this cell too — Excel doesn't "
         "do that for you automatically, and the app will flag the row on import if "
         "you forget."),
        ("Drawn on Bank",
         "Only needed for Pay at Bank Counter (Cash, Cheque, or Demand Draft) — the "
         "bank the payment is made at. Greys out for every other Payment Mode — "
         "leave it empty."),
        ("Tax / Surcharge / Cess /\nInterest / Penalty / Others",
         "Most of the time you'll only have one total amount — put it all in Tax "
         "and leave the rest as 0. Only split it across the other columns if you "
         "actually have a breakup, e.g. interest calculated separately."),
    ]
    for i, (field, desc) in enumerate(sections):
        field_cell = ws_help.cell(row=r, column=1, value=field)
        desc_cell = ws_help.cell(row=r, column=2, value=desc)
        fill = BAND_FILL if i % 2 == 1 else None
        for c in (field_cell, desc_cell):
            c.border = box_border
            if fill:
                c.fill = fill
        field_cell.font = Font(bold=True)
        field_cell.alignment = Alignment(wrap_text=True, vertical="top", indent=1)
        desc_cell.alignment = Alignment(wrap_text=True, vertical="top", indent=1)
        ws_help.row_dimensions[r].height = 42
        r += 1

    r += 1
    ws_help.merge_cells(f"A{r}:B{r}")
    note_cell = ws_help.cell(
        row=r, column=1,
        value='Note: If you type into a grey cell, you\'ll get a red error. That just '
              'means that field doesn\'t apply to the Payment Mode you picked on that '
              'row — leave it blank instead.',
    )
    note_cell.font = Font(bold=True, color="7F6000")
    note_cell.fill = NOTE_FILL
    note_cell.border = box_border
    note_cell.alignment = Alignment(wrap_text=True, vertical="center", indent=1)
    ws_help.row_dimensions[r].height = 40

    wb.active = wb["Instructions"]

    wb.save(path)


def download_challan_template(path: str):
    """Standalone equivalent of GenerateChallansDialog._download_template()
    — lets "E-Pay Tax > Download Import Template" generate the template
    without opening the full dialog first. No per-dialog state is needed:
    the sample row's PAN is fake and never matches a real vault entry, so
    there's nothing to look up."""
    from automation.challan_fields import CHALLAN_INPUT_COLUMNS
    from automation.challan_generator import DEFAULT_PAYMENT_MODE, DEFAULT_BANK
    headers = ["PAN", "Name"] + [label for key, label, _ in CHALLAN_INPUT_COLUMNS if key != "pan"]
    sample_values = {
        "payment_mode": DEFAULT_PAYMENT_MODE, "bank": DEFAULT_BANK, "drawee_bank": "",
        "tax": 15000, "surcharge": 0, "cess": 0, "interest": 0, "penalty": 0, "others": 0,
    }
    sample_row = ["AAAPT0001A", "Sample Client Name"] + [
        sample_values.get(key, "") for key, _, _ in CHALLAN_INPUT_COLUMNS if key != "pan"
    ]
    write_challan_table_file(path, headers, [sample_row])


class GenerateChallansDialog(QDialog):
    """
    Bulk-generate tax payment challans (Advance Tax / Self-Assessment Tax)
    on the ITD e-Pay Tax "New Payment" wizard, for many clients in one run.

    Financial Year is picked ONCE for the whole dialog (not per row) — the
    same year always applies to every client in a batch (a CA works through
    "Q2 advance tax for FY 2026-27, for these 15 clients", never a mix of
    years in one run). Tax Type (Advance/Self-Assessment) is never a field
    the user fills in — it's computed from the FY via
    automation.challan_generator.resolve_tax_type(), using the app's own
    assessment_years.json entries, and shown read-only next to the Year combo.

    The summary table only shows PAN, Name, Total Amount, and Payment Mode
    (per the user's request to keep the main view scannable) — double-
    clicking a row (or "+ Add Row") opens ChallanRowDetailDialog for full
    editing of PAN, Payment Mode, Bank/Sub-Mode, and the amount breakup.
    Row data lives in self._row_data (a plain list of dicts), not in the
    QTableWidget cells directly. Column labels come from
    automation.challan_fields.CHALLAN_INPUT_COLUMNS — the single source of
    truth also used by the Excel/CSV import, export, and template — see
    automation/doc_types.py's own docstring for why two independent column
    lists caused the Form 168 emailer bugs twice already.
    """
    _COL_PAN = 0
    _COL_NAME = 1
    _COL_TOTAL = 2
    _COL_MODE = 3
    _COL_BANK = 4

    def __init__(self, parent, vault, ay_entries):
        super().__init__(parent)
        from automation.challan_fields import CHALLAN_INPUT_COLUMNS, CHALLAN_AMOUNT_FIELDS
        from automation.challan_generator import resolve_tax_type, TAX_TYPES, cash_limit_exceeded

        self._vault = vault
        self._ay_entries = ay_entries
        self._resolve_tax_type = resolve_tax_type
        self._tax_types = TAX_TYPES
        self._cash_limit_exceeded = cash_limit_exceeded
        self._columns = CHALLAN_INPUT_COLUMNS  # [(key, label, kind), ...] — full per-client field shape
        self._amount_keys = CHALLAN_AMOUNT_FIELDS

        self.fy_value = None
        self.rows = []
        self._tax_type_valid = False
        self._row_data: list = []   # [{"pan", "payment_mode", "bank", "tax", ...}, ...]

        self.setWindowTitle("Generate Tax Challans")
        self.setMinimumSize(760, 520)
        self.resize(820, 560)
        self.setSizeGripEnabled(True)
        _bt = _t()
        self.setStyleSheet(f"QDialog{{background:{_bt.bg_window};}}")

        layout = QVBoxLayout(self)
        layout.setContentsMargins(16, 14, 16, 12)
        layout.setSpacing(8)

        # ── Year selector row ────────────────────────────────────────────
        # Same labels as the main screen's Assessment/Tax Year combo (e.g.
        # "TY 2026-27 (FY 2026-27)", "AY 2026-27 (FY 2025-26)") — single-
        # select here (not the main screen's checkable multi-select), since
        # exactly one Year/Tax Type applies to the whole batch (see Context
        # decision 5 in the plan). Each entry's own AY/TY key already tells
        # us the Tax Type directly.
        fy_row = QHBoxLayout()
        fy_row.setSpacing(10)
        fy_row.addWidget(_lbl("Assessment / Tax Year:"))
        self._fy_combo = QComboBox()
        for e in ay_entries:
            if not e.get("enabled", True):
                continue
            y = e.get("year", {})
            fy = y.get("FY")
            if not fy:
                continue
            self._fy_combo.addItem(e.get("label", fy), fy)
        self._fy_combo.currentIndexChanged.connect(self._on_fy_changed)
        fy_row.addWidget(self._fy_combo)
        # This is the single most consequential computed fact in the dialog
        # (it decides the Act and Type-of-Payment the whole run submits
        # under) — styled as a prominent accent-colored badge, not a small
        # muted caption, so it can't be missed.
        self._type_label = QLabel("")
        self._type_label.setStyleSheet(
            f"color:{_bt.accent_text};background:{_bt.accent};font-size:13px;"
            f"font-weight:bold;padding:4px 10px;border-radius:10px;")
        fy_row.addWidget(self._type_label)
        fy_row.addStretch(1)
        layout.addLayout(fy_row)

        # ── Toolbar ──────────────────────────────────────────────────────
        # Icon-only with tooltips, not icon+text — six full-text buttons
        # didn't fit the dialog width and truncated ("Import Excel/C...").
        toolbar = QHBoxLayout()
        toolbar.setSpacing(6)

        def _icon_btn(icon, tooltip, style="outline"):
            b = _btn(icon, style, min_width=40)
            b.setToolTip(tooltip)
            return b

        btn_add = _icon_btn("➕", "Add Row", "primary")
        btn_add.clicked.connect(self._add_row)
        toolbar.addWidget(btn_add)
        btn_remove = _icon_btn("🗑", "Remove Row", "danger")
        btn_remove.clicked.connect(self._remove_row)
        toolbar.addWidget(btn_remove)
        btn_import = _icon_btn("📥", "Import Excel/CSV")
        btn_import.clicked.connect(self._import_rows)
        toolbar.addWidget(btn_import)
        btn_export = _icon_btn("📤", "Export")
        btn_export.clicked.connect(self._export_rows)
        toolbar.addWidget(btn_export)
        btn_template = _icon_btn("📄", "Download Import Template")
        btn_template.clicked.connect(self._download_template)
        toolbar.addWidget(btn_template)
        toolbar.addStretch(1)
        # Lightweight in-app persistence — a single local "last saved"
        # slot the user can save to / reload from without going through a
        # file picker each time (Export/Import stay available for actually
        # sharing a file with someone or keeping dated copies).
        btn_save_draft = _icon_btn("💾", "Save Draft")
        btn_save_draft.clicked.connect(self._save_draft)
        toolbar.addWidget(btn_save_draft)
        self._btn_reload_draft = _icon_btn("📂", "Reload Last Saved")
        self._btn_reload_draft.clicked.connect(lambda: self._load_draft(confirm=True))
        self._btn_reload_draft.setEnabled(os.path.exists(self._draft_path()))
        toolbar.addWidget(self._btn_reload_draft)
        layout.addLayout(toolbar)

        # ── Summary table (PAN / Name / Total / Mode / Bank) ─────────────
        # Mode and Bank / Sub-Mode used to share one column as combined
        # text ("Payment Gateway including UPI and Credit Card / Cash"),
        # which wrapped to two lines and truncated the mode name — split
        # into their own columns so each reads cleanly at a glance.
        self._table = QTableWidget(0, 5)
        self._table.setHorizontalHeaderLabels(
            ["PAN", "Name", "Total Amount", "Payment Mode", "Bank / Sub-Mode"])
        hdr = self._table.horizontalHeader()
        hdr.setSectionResizeMode(self._COL_PAN, QHeaderView.ResizeMode.Interactive)
        hdr.setSectionResizeMode(self._COL_NAME, QHeaderView.ResizeMode.Stretch)
        hdr.setSectionResizeMode(self._COL_TOTAL, QHeaderView.ResizeMode.Interactive)
        hdr.setSectionResizeMode(self._COL_MODE, QHeaderView.ResizeMode.Interactive)
        hdr.setSectionResizeMode(self._COL_BANK, QHeaderView.ResizeMode.Interactive)
        self._table.setColumnWidth(self._COL_PAN, 110)
        self._table.setColumnWidth(self._COL_TOTAL, 110)
        self._table.setColumnWidth(self._COL_MODE, 200)
        self._table.setColumnWidth(self._COL_BANK, 150)
        self._table.verticalHeader().setVisible(False)
        self._table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self._table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self._table.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        self._table.cellDoubleClicked.connect(self._edit_row)
        # Same theme-aware table/header styling BatchProgressDialog already
        # uses — without it, QTableWidget falls back to native (light)
        # colors regardless of the app's dark/light theme, which is exactly
        # what made the header text unreadable in dark mode.
        self._table.setStyleSheet(
            f"QTableWidget{{border:1.5px solid {_bt.border};border-radius:8px;"
            f"background:{_bt.bg_table};outline:0;gridline-color:{_bt.grid};}}"
            f"QTableWidget::item{{border-bottom:1px solid {_bt.grid};padding:0 8px;}}")
        hdr.setStyleSheet(
            f"QHeaderView::section{{"
            f"background-color:{_bt.bg_header};"
            f"border:none;"
            f"border-right:1px solid {_bt.border};"
            f"border-bottom:1px solid {_bt.border};"
            f"font-weight:bold;color:{_bt.text_muted};"
            f"font-size:11px;height:34px;"
            f"padding:0 8px;}}")
        layout.addWidget(self._table, stretch=1)

        hint = QLabel("Double-click a row to view/edit its full details (bank, amount breakup).")
        hint.setStyleSheet(f"color:{_bt.text_muted};font-size:11px;background:transparent;")
        layout.addWidget(hint)

        # ── Footer ───────────────────────────────────────────────────────
        footer = QHBoxLayout()
        self._counts_label = QLabel("")
        self._counts_label.setStyleSheet(f"color:{_bt.text_muted};font-size:12px;background:transparent;")
        footer.addWidget(self._counts_label)
        footer.addStretch(1)
        btn_cancel = _btn("Cancel", "outline")
        btn_cancel.clicked.connect(self.reject)
        footer.addWidget(btn_cancel)
        self._btn_generate = _btn("Generate", "success")
        self._btn_generate.clicked.connect(self._on_generate_clicked)
        footer.addWidget(self._btn_generate)
        layout.addLayout(footer)

        self._on_fy_changed()
        self._refresh_table()

        if os.path.exists(self._draft_path()):
            import datetime as _dt
            saved_at = _dt.datetime.fromtimestamp(os.path.getmtime(self._draft_path()))
            if QMessageBox.question(
                self, "Resume Last Saved Draft?",
                f"A saved draft exists from {saved_at.strftime('%d-%b-%Y %H:%M')}. Load it?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.No,
            ) == QMessageBox.StandardButton.Yes:
                self._load_draft(confirm=False)

    # ── Save / reload draft (local, no file picker) ──────────────────────

    def _draft_path(self) -> str:
        return os.path.join(_app_dir(), "challan_draft.json")

    def _save_draft(self):
        data = {"fy_value": self.fy_value, "rows": self._row_data}
        try:
            with open(self._draft_path(), "w", encoding="utf-8") as f:
                json.dump(data, f, indent=2)
            self._btn_reload_draft.setEnabled(True)
            QMessageBox.information(self, "Draft Saved",
                f"{len(self._row_data)} row(s) saved as your last draft — "
                "use \"Reload Last Saved\" next time instead of Export/Import.")
        except Exception as e:
            QMessageBox.critical(self, "Save Failed", f"Could not save draft: {e}")

    def _load_draft(self, confirm: bool = True):
        path = self._draft_path()
        if not os.path.exists(path):
            QMessageBox.information(self, "No Saved Draft", "No saved draft was found.")
            return
        if confirm and self._row_data and QMessageBox.question(
            self, "Replace Current Rows?",
            f"This replaces the {len(self._row_data)} row(s) currently in the table with the "
            "last saved draft. Continue?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No,
        ) != QMessageBox.StandardButton.Yes:
            return
        try:
            with open(path, "r", encoding="utf-8") as f:
                data = json.load(f)
        except Exception as e:
            QMessageBox.critical(self, "Load Failed", f"Could not read saved draft: {e}")
            return

        fy_value = data.get("fy_value")
        if fy_value:
            idx = self._fy_combo.findData(fy_value)
            if idx >= 0:
                self._fy_combo.setCurrentIndex(idx)
        self._row_data = data.get("rows", [])
        self._refresh_table()

    # ── FY / Tax Type ────────────────────────────────────────────────────

    def _on_fy_changed(self):
        fy = self._fy_combo.currentData()
        self.fy_value = fy or None
        if not fy:
            self._type_label.setText("")
            self._tax_type_valid = False
            self._update_footer_counts()
            return
        try:
            tax_type, _portal_year_label = self._resolve_tax_type(fy, self._ay_entries)
            # portal_year_label is intentionally not shown here — it's
            # already visible as part of the selected combo item's own text
            # (e.g. "TY 2026-27" / "AY 2026-27"), so repeating it here would
            # just be noise.
            label = self._tax_types[tax_type]["label"]
            self._type_label.setText(f"→ {label}")
            _bt = _t()
            self._type_label.setStyleSheet(
                f"color:{_bt.accent_text};background:{_bt.accent};font-size:13px;"
                f"font-weight:bold;padding:4px 10px;border-radius:10px;")
            self._tax_type_valid = True
        except Exception as e:
            _bt = _t()
            warn = getattr(_bt, "warning", "#D97706")
            self._type_label.setText(f"→ {e}")
            self._type_label.setStyleSheet(
                f"color:{warn};background:transparent;font-size:12px;font-weight:normal;padding:0;")
            self._tax_type_valid = False
        self._update_footer_counts()

    # ── Row management ───────────────────────────────────────────────────

    def _row_total(self, row: dict) -> float:
        total = 0.0
        for key in self._amount_keys:
            try:
                total += float(row.get(key, 0) or 0)
            except (TypeError, ValueError):
                pass
        return total

    def _row_bank_problem(self, row: dict) -> str:
        # Delegates to automation.challan_generator.bank_problem() — the
        # single source of truth shared with ChallanRowDetailDialog (the
        # manual add/edit dialog) and the import flow below, so all three
        # agree on what counts as valid instead of drifting apart.
        from automation.challan_generator import bank_problem
        return bank_problem(row.get("payment_mode", ""), row.get("bank", ""))

    def _row_drawee_problem(self, row: dict) -> str:
        # Same idea, for Drawn on Bank — see automation.challan_generator's
        # drawee_bank_problem() docstring. BUG FIX (2026-09-03): confirmed
        # live — a row could be edited and saved with Drawn on Bank left
        # blank under Pay at Bank Counter, since ChallanRowDetailDialog's
        # OK button never checked it. Now checked here too, so such a row
        # can't slip into "Ready to generate" even if it somehow gets past
        # the edit dialog (e.g. an older draft file).
        from automation.challan_generator import drawee_bank_problem
        return drawee_bank_problem(row.get("payment_mode", ""), row.get("drawee_bank", ""))

    def _row_missing_bank(self, row: dict) -> bool:
        return bool(self._row_bank_problem(row)) or bool(self._row_drawee_problem(row))

    def _row_is_ready(self, row: dict) -> bool:
        pan = (row.get("pan") or "").upper()
        if not pan:
            return False
        matched = any(a.get("pan", "").upper() == pan for a in self._vault.get_all_assessees())
        if not matched:
            return False
        total = self._row_total(row)
        if total <= 0:
            return False
        if self._cash_limit_exceeded(row.get("payment_mode", ""), row.get("bank", ""), total):
            return False
        if self._row_missing_bank(row):
            return False
        return True

    def _refresh_table(self):
        _bt = _t()
        self._table.setRowCount(len(self._row_data))
        for i, row in enumerate(self._row_data):
            pan = (row.get("pan") or "").upper()
            matched = next((a for a in self._vault.get_all_assessees() if a.get("pan", "").upper() == pan), None)
            self._table.setRowHeight(i, 36)

            pan_item = QTableWidgetItem(pan)
            pan_item.setForeground(QColor(_bt.text_primary))
            self._table.setItem(i, self._COL_PAN, pan_item)

            name_item = QTableWidgetItem()
            if pan and not matched:
                name_item.setText("⚠ Unknown PAN")
                name_item.setForeground(QColor(getattr(_bt, "warning", "#D97706")))
            elif matched:
                name_item.setText(matched.get("name", ""))
                name_item.setForeground(QColor(_bt.text_primary))
            self._table.setItem(i, self._COL_NAME, name_item)

            total = self._row_total(row)
            total_item = QTableWidgetItem(f"{total:g}")
            total_item.setForeground(QColor(_bt.text_primary))
            if self._cash_limit_exceeded(row.get("payment_mode", ""), row.get("bank", ""), total):
                total_item.setForeground(QColor(getattr(_bt, "warning", "#D97706")))
            self._table.setItem(i, self._COL_TOTAL, total_item)

            mode = row.get("payment_mode", "")
            bank = row.get("bank", "")
            # Drawn on Bank has no dedicated table column (only shown in
            # the row detail dialog), so a problem with it surfaces here
            # too rather than staying invisible until the row is opened.
            problem = self._row_bank_problem(row) or self._row_drawee_problem(row)
            warn_color = QColor(getattr(_bt, "warning", "#D97706"))
            normal_color = QColor(_bt.text_primary)

            mode_item = QTableWidgetItem(mode)
            mode_item.setForeground(warn_color if problem else normal_color)
            self._table.setItem(i, self._COL_MODE, mode_item)

            bank_item = QTableWidgetItem(f"⚠ {problem}" if problem else bank)
            bank_item.setForeground(warn_color if problem else normal_color)
            if problem:
                bank_item.setToolTip(problem.capitalize() + ".")
            self._table.setItem(i, self._COL_BANK, bank_item)

        self._update_footer_counts()

    def _add_row(self):
        dlg = ChallanRowDetailDialog(self, self._vault)
        if dlg.exec() == dlg.DialogCode.Accepted and dlg.result_data:
            self._row_data.append(dlg.result_data)
            self._refresh_table()

    def _edit_row(self, row, _col=0):
        if row < 0 or row >= len(self._row_data):
            return
        dlg = ChallanRowDetailDialog(self, self._vault, self._row_data[row])
        if dlg.exec() == dlg.DialogCode.Accepted and dlg.result_data:
            self._row_data[row] = dlg.result_data
            self._refresh_table()

    def _remove_row(self):
        rows = sorted({idx.row() for idx in self._table.selectedIndexes()}, reverse=True)
        if not rows and self._row_data:
            rows = [len(self._row_data) - 1]
        for r in rows:
            if 0 <= r < len(self._row_data):
                del self._row_data[r]
        self._refresh_table()

    def _update_footer_counts(self):
        total_rows = len(self._row_data)
        ready = sum(1 for r in self._row_data if self._row_is_ready(r))
        flagged = total_rows - ready
        self._counts_label.setText(f"Rows: {total_rows}   Ready to generate: {ready}   Flagged: {flagged}")
        self._btn_generate.setEnabled(ready > 0 and self._tax_type_valid)

    # ── Import / Export / Template ───────────────────────────────────────

    def _import_rows(self):
        path, _ = QFileDialog.getOpenFileName(self, "Import Challan Rows", "",
            "Excel / CSV files (*.xlsx *.csv)")
        if not path:
            return
        try:
            headers, data_rows = self._read_table_file(path)
        except Exception as e:
            QMessageBox.critical(self, "Import Error", f"Failed to read file: {e}")
            return

        # Headers in the file are the human-readable LABELS ("Payment
        # Mode", "Drawn on Bank"), not the internal field keys
        # ("payment_mode", "drawee_bank") — map file headers to field keys
        # via CHALLAN_INPUT_COLUMNS' own label text, not by assuming they're
        # spelled the same. A raw "pan"/"tax" header still matches directly
        # since those particular labels happen to equal their key already.
        # An extra "Name" column (present in every file this dialog writes,
        # purely for the human filling it in — it never drives anything on
        # import, PAN is the only join key) is simply left unmapped here and
        # ignored, same as any other unrecognized column.
        label_to_key = {label.lower(): key for key, label, _ in self._columns}
        col = {label_to_key.get(h, h): i for i, h in enumerate(headers)}
        required = {"pan", "tax"}
        missing = required - set(col)
        if missing:
            QMessageBox.critical(self, "Import Error",
                f"Missing required columns: {', '.join(missing)}. Headers must include PAN and Tax.")
            return

        from automation.challan_generator import DEFAULT_PAYMENT_MODE, DEFAULT_BANK, all_bank_options

        added = 0
        errors = []
        warnings = []
        for idx, raw_row in enumerate(data_rows):
            row_num = idx + 2
            try:
                def _cell(key):
                    if key not in col or col[key] >= len(raw_row):
                        return None
                    v = raw_row[col[key]]
                    return str(v).strip() if v not in (None, "") else ""

                pan = (_cell("pan") or "").upper()
                if not pan:
                    errors.append(f"Row {row_num}: Missing PAN.")
                    continue

                row_payment_mode = _cell("payment_mode") or DEFAULT_PAYMENT_MODE
                row_bank_options = all_bank_options(row_payment_mode)
                # BUG FIX (2026-09-03): confirmed live, twice — the first
                # round only fixed RTGS/NEFT (no bank picklist at all, so
                # any default was wrong); this round confirmed live that
                # the same DEFAULT_BANK ("Cheque") was still being applied
                # to EVERY other mode with a blank Bank / Sub-Mode cell too,
                # including Net Banking — where "Cheque" isn't a valid
                # option at all (it's a Pay at Bank Counter sub-mode).
                # DEFAULT_BANK is only ever a sensible guess for Pay at
                # Bank Counter, so only default to it when it's actually
                # one of this mode's own options; every other mode with a
                # blank Bank / Sub-Mode is a genuinely missing mandatory
                # field, not something to guess at — leave it blank and
                # flag the row (see the missing-mandatory-bank check
                # below), same as a missing PAN is flagged rather than
                # silently guessed.
                default_bank = DEFAULT_BANK if DEFAULT_BANK in row_bank_options else ""
                row_bank = _cell("bank") or default_bank
                # BUG FIX (2026-09-03): confirmed live — a mode with NO
                # bank picklist at all (RTGS/NEFT) can still carry a stale
                # value left over from before the Payment Mode was changed
                # (Excel can't clear it automatically — see the note above
                # dv_bank). There's nothing ambiguous about this case: the
                # mode simply never uses this field, so the leftover value
                # is noise, not a problem to make the user go fix — silently
                # drop it instead of flagging the row over it.
                if not row_bank_options:
                    row_bank = ""
                row_drawee_bank = _cell("drawee_bank") or ""
                # Same leniency as the RTGS/NEFT case above — Drawn on Bank
                # only ever applies to Pay at Bank Counter, so a stray value
                # under any other mode is unambiguous noise, not something
                # to flag the row over.
                if row_payment_mode != "Pay at Bank Counter":
                    row_drawee_bank = ""
                new_row = {
                    "pan": pan,
                    "payment_mode": row_payment_mode,
                    "bank": row_bank,
                    "drawee_bank": row_drawee_bank,
                }
                # BUG FIX (2026-09-03): confirmed live — a row edited from
                # one Payment Mode to another that still HAS its own bank
                # picklist (e.g. Pay at Bank Counter to Net Banking) can
                # keep its OLD Bank / Sub-Mode value, since Excel's
                # cascading dropdown narrows future choices but can't clear
                # what's already typed in the cell. Only checking for a
                # *blank* Bank / Sub-Mode missed this — a stale,
                # no-longer-valid value slipped through unflagged. Unlike
                # the no-picklist-at-all case above, this one IS ambiguous
                # (which real bank did they mean?), so it's still flagged
                # rather than guessed at or silently dropped.
                bank_problem = self._row_bank_problem(new_row)
                if bank_problem:
                    warnings.append(
                        f"Row {row_num}: Payment Mode '{row_payment_mode}' {bank_problem} — "
                        f"row imported but flagged, won't run until fixed."
                    )
                # BUG FIX (2026-09-03): confirmed live — Drawn on Bank is
                # mandatory for every Pay at Bank Counter sub-mode (Cash
                # included), but a blank one wasn't checked on import at
                # all, only in the (also-unchecked-until-now) edit dialog.
                drawee_problem = self._row_drawee_problem(new_row)
                if drawee_problem:
                    warnings.append(
                        f"Row {row_num}: Payment Mode '{row_payment_mode}' {drawee_problem} — "
                        f"row imported but flagged, won't run until fixed."
                    )
                bad_amount = False
                for key, label, kind in self._columns:
                    if kind != "amount":
                        continue
                    raw_val = _cell(key)
                    if not raw_val:
                        new_row[key] = 0
                        continue
                    try:
                        new_row[key] = float(raw_val)
                    except ValueError:
                        errors.append(f"Row {row_num}: '{label}' is not a number.")
                        bad_amount = True
                if bad_amount:
                    continue

                self._row_data.append(new_row)
                added += 1
            except Exception as e:
                errors.append(f"Row {row_num}: Error importing entry: {e}")

        self._refresh_table()
        summary = f"{added} row(s) imported"
        if errors:
            summary += f", {len(errors)} skipped"
        if warnings:
            summary += f", {len(warnings)} flagged (Bank / Sub-Mode problem)"
        if errors or warnings:
            summary += " — see details"
            QMessageBox.warning(self, "Import Complete", summary + "\n\n" + "\n".join(errors + warnings))
        else:
            QMessageBox.information(self, "Import Complete", summary + ".")

    def _read_table_file(self, path):
        if path.endswith(".xlsx"):
            from openpyxl import load_workbook
            wb = load_workbook(path, data_only=True)
            # Files this dialog writes carry a hidden "Lists" sheet
            # alongside the data — pick the "Challans" sheet explicitly if
            # present rather than trusting wb.active (which is usually
            # right, but is a saved-state property, not a guarantee).
            ws = wb["Challans"] if "Challans" in wb.sheetnames else wb.active
            raw_rows = list(ws.iter_rows(values_only=True))
        elif path.endswith(".csv"):
            import csv
            with open(path, newline="", encoding="utf-8-sig") as f:
                raw_rows = list(csv.reader(f))
        else:
            raise ValueError("Unsupported file format. Please use Excel (.xlsx) or CSV (.csv).")
        if not raw_rows:
            raise ValueError("File is empty.")
        headers = [str(c).strip().lower() if c is not None else "" for c in raw_rows[0]]
        # BUG FIX (2026-09-03): confirmed live — padding the Challans table
        # out to a 50-row minimum (so it's ready for bulk entry without a
        # manual Resize Table) makes openpyxl report those blank padded
        # rows as real rows on read-back, each triggering a "missing PAN"
        # error when the still-mostly-blank template is re-imported as-is.
        # A completely empty row was never real data to begin with.
        data_rows = [
            row for row in raw_rows[1:]
            if any(c is not None and str(c).strip() != "" for c in row)
        ]
        return headers, data_rows

    def _offer_open_file(self, title: str, message: str, path: str):
        """Success dialog with a one-click way to open the file it just
        wrote, instead of leaving the user to go find it themselves."""
        box = QMessageBox(self)
        box.setWindowTitle(title)
        box.setText(message)
        open_btn = box.addButton("Open File", QMessageBox.ButtonRole.ActionRole)
        box.addButton(QMessageBox.StandardButton.Ok)
        box.exec()
        if box.clickedButton() is open_btn:
            _open_path(path)

    def _export_headers(self) -> list:
        """Column layout for both Export and the Import Template — a
        read-only "Name" column sits right after PAN purely so the person
        filling the sheet can see whose row is whose; it's never read back
        on import (PAN is the only join key — see _import_rows)."""
        return ["PAN", "Name"] + [label for key, label, _ in self._columns if key != "pan"]

    def _row_to_export_values(self, row: dict) -> list:
        pan = (row.get("pan") or "").upper()
        matched = next((a for a in self._vault.get_all_assessees() if a.get("pan", "").upper() == pan), None)
        name = matched.get("name", "") if matched else ""
        return [pan, name] + [row.get(key, "") for key, _, _ in self._columns if key != "pan"]

    def _export_rows(self):
        path, _ = QFileDialog.getSaveFileName(self, "Export Challan Rows",
            "Challan_Rows", "Excel Workbook (*.xlsx);;CSV (*.csv)")
        if not path:
            return
        if not (path.endswith(".xlsx") or path.endswith(".csv")):
            path += ".xlsx"
        headers = self._export_headers()
        rows = [self._row_to_export_values(row) for row in self._row_data]
        try:
            write_challan_table_file(path, headers, rows)
            self._offer_open_file("Export Complete", f"{len(rows)} row(s) exported to:\n{path}", path)
        except Exception as e:
            QMessageBox.critical(self, "Export Error", f"Failed: {e}")

    def _download_template(self):
        path, _ = QFileDialog.getSaveFileName(self, "Save Template",
            "Challan_Rows_Template", "Excel Workbook (*.xlsx);;CSV (*.csv)")
        if not path:
            return
        try:
            download_challan_template(path)
            self._offer_open_file("Success", f"Template generated at:\n{path}", path)
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed: {e}")

    # ── Accept ───────────────────────────────────────────────────────────

    def _on_generate_clicked(self):
        if not self.fy_value:
            QMessageBox.warning(self, "No Year Selected", "Please select an Assessment / Tax Year.")
            return
        try:
            self._resolve_tax_type(self.fy_value, self._ay_entries)
        except Exception as e:
            QMessageBox.critical(self, "Year Not Ready", str(e))
            return

        rows = [dict(row) for row in self._row_data if self._row_is_ready(row)]
        if not rows:
            QMessageBox.warning(self, "Nothing to Generate", "No rows are ready to generate.")
            return

        self.rows = rows
        self.accept()
class ChallanGenerationProgressDialog(QDialog):
    """
    Live progress popup for a Generate Tax Challans run (F-64). Deliberately
    NOT a reuse of BatchProgressDialog — that class's rows are a hardcoded
    `targets × year_specs` cross-product built for multiple years per run;
    this feature has exactly one FY/Tax Type for the whole run (see
    GenerateChallansDialog). Rows are keyed by their position in `targets`,
    not by PAN — a client can appear more than once in the same run (e.g.
    a Cash challan up to the ₹10,000 cap plus a Cheque challan for the
    remainder), and keying by PAN alone would collide both rows' status
    updates onto whichever one was registered last.
    """
    _update_signal = pyqtSignal(int, str)   # row_index, status
    _path_signal = pyqtSignal(int, str)     # row_index, path
    _done_signal = pyqtSignal()

    _COL_NAME = 0
    _COL_PAN = 1
    _COL_STATUS = 2
    _COL_PATH = 3

    def __init__(self, targets: list, year_display: str, type_label: str,
                 stop_callback=None, tray_callback=None, output_dir: str = "", parent=None):
        super().__init__(parent)
        self._stop_callback = stop_callback
        self._tray_callback = tray_callback
        self._output_dir = output_dir
        self._targets = targets
        self._last_report_path = ""

        self.setWindowTitle(f"Generate Tax Challans — {year_display} — Batch Progress")
        self.setMinimumSize(820, 460)
        self.resize(960, min(160 + len(targets) * 42, 680))
        self.setSizeGripEnabled(True)
        _bt = _t()
        self.setStyleSheet(f"QDialog{{background:{_bt.bg_window};}}")

        self._counted_rows = set()
        self._done_count = 0
        self._total = len(targets)

        layout = QVBoxLayout(self)
        layout.setContentsMargins(16, 14, 16, 12)
        layout.setSpacing(8)

        title = QLabel(f"<b>Generating Tax Challans</b> — {len(targets)} client(s) "
                        f"&nbsp;·&nbsp; <span style='color:{_bt.accent}'>{year_display} → {type_label}</span>")
        title.setStyleSheet(f"font-size:14px; color:{_bt.text_primary}; background:transparent;")
        layout.addWidget(title)

        self._table = QTableWidget(len(targets), 4)
        self._table.setHorizontalHeaderLabels(["Name", "PAN", "Status", "Artifact"])
        hdr = self._table.horizontalHeader()
        hdr.setSectionResizeMode(self._COL_NAME, QHeaderView.ResizeMode.Interactive)
        hdr.setSectionResizeMode(self._COL_PAN, QHeaderView.ResizeMode.Interactive)
        hdr.setSectionResizeMode(self._COL_STATUS, QHeaderView.ResizeMode.Interactive)
        hdr.setSectionResizeMode(self._COL_PATH, QHeaderView.ResizeMode.Stretch)
        self._table.setColumnWidth(self._COL_NAME, 180)
        self._table.setColumnWidth(self._COL_PAN, 120)
        self._table.setColumnWidth(self._COL_STATUS, 260)
        self._table.verticalHeader().setVisible(False)
        self._table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self._table.setSelectionMode(QAbstractItemView.SelectionMode.NoSelection)
        self._table.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        # Same theme-aware table/header styling BatchProgressDialog already
        # uses — without it, QTableWidget (and the QProgressBar below)
        # fall back to native light-mode colors regardless of the app's
        # theme, same bug already fixed once for GenerateChallansDialog's
        # own table but missed here.
        self._table.setStyleSheet(
            f"QTableWidget{{border:1.5px solid {_bt.border};border-radius:8px;"
            f"background:{_bt.bg_table};outline:0;gridline-color:{_bt.grid};}}"
            f"QTableWidget::item{{border-bottom:1px solid {_bt.grid};padding:0 8px;}}")
        hdr.setStyleSheet(
            f"QHeaderView::section{{"
            f"background-color:{_bt.bg_header};"
            f"border:none;"
            f"border-right:1px solid {_bt.border};"
            f"border-bottom:1px solid {_bt.border};"
            f"font-weight:bold;color:{_bt.text_muted};"
            f"font-size:11px;height:34px;"
            f"padding:0 8px;}}")

        for row, tgt in enumerate(targets):
            pan = tgt.get("pan", "")
            self._table.setRowHeight(row, 40)
            name_item = QTableWidgetItem(tgt.get("name", "—"))
            name_item.setForeground(QColor(_bt.text_primary))
            self._table.setItem(row, self._COL_NAME, name_item)
            pan_item = QTableWidgetItem(pan)
            pan_item.setFont(QFont(_MONO_FONT, 10))
            pan_item.setForeground(QColor(_bt.text_muted))
            self._table.setItem(row, self._COL_PAN, pan_item)
            status_item = QTableWidgetItem("⬜ Waiting")
            status_item.setForeground(QColor(_bt.text_primary))
            self._table.setItem(row, self._COL_STATUS, status_item)
            path_lbl = QLabel("—")
            path_lbl.setStyleSheet(f"color:{_bt.text_muted};font-size:11px;padding:0 8px;background:transparent;")
            path_lbl.setAlignment(Qt.AlignmentFlag.AlignVCenter | Qt.AlignmentFlag.AlignLeft)
            path_lbl.setWordWrap(False)
            path_lbl.setTextInteractionFlags(Qt.TextInteractionFlag.TextBrowserInteraction)
            path_lbl.setOpenExternalLinks(False)
            path_lbl.linkActivated.connect(self._open_row_path)
            self._table.setCellWidget(row, self._COL_PATH, path_lbl)

        layout.addWidget(self._table, stretch=1)

        self._progress_bar = QProgressBar()
        self._progress_bar.setRange(0, self._total)
        self._progress_bar.setFixedHeight(18)
        self._progress_bar.setTextVisible(True)
        self._progress_bar.setFormat(f"0 / {self._total} done")
        self._progress_bar.setStyleSheet(
            f"QProgressBar{{border:1px solid {_bt.border};border-radius:9px;"
            f"background:{_bt.scrollbar_handle};text-align:center;font-size:11px;"
            f"font-weight:600;color:{_bt.accent_text};}}"
            f"QProgressBar::chunk{{background:#16A34A;border-radius:9px;}}")
        layout.addWidget(self._progress_bar)

        footer = QHBoxLayout()
        self._loc_val = QLabel(output_dir or "—")
        self._loc_val.setStyleSheet(f"color:{_bt.text_muted};font-size:11px;background:transparent;")
        self._loc_val.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Preferred)
        footer.addWidget(self._loc_val, stretch=1)

        self._open_folder_btn = _btn("📂  Open Folder", "outline")
        self._open_folder_btn.clicked.connect(self._open_output_dir)
        footer.addWidget(self._open_folder_btn)

        self._report_btn = _btn("📊  Open Summary", "outline")
        self._report_btn.setEnabled(False)
        self._report_btn.clicked.connect(self._open_report)
        footer.addWidget(self._report_btn)

        self._tray_btn = _btn("⬇  Tray", "outline")
        self._tray_btn.setToolTip("Hide to system tray — click the tray icon to restore")
        self._tray_btn.setVisible(bool(self._tray_callback))
        self._tray_btn.clicked.connect(self._on_tray_clicked)
        footer.addWidget(self._tray_btn)

        self._stop_btn = _btn("⏹  Stop", "danger")
        self._stop_btn.clicked.connect(self._on_stop_clicked)
        footer.addWidget(self._stop_btn)
        layout.addLayout(footer)

        self._update_signal.connect(self._on_update)
        self._path_signal.connect(self._on_path)

    def set_status(self, row_index, text):
        self._update_signal.emit(row_index, text)

    def set_artifact_path(self, row_index, path):
        self._path_signal.emit(row_index, path)

    def set_report_path(self, path):
        self._last_report_path = path
        self._report_btn.setEnabled(bool(path))

    def _on_update(self, row_index, text):
        if not (0 <= row_index < self._total):
            return
        self._table.item(row_index, self._COL_STATUS).setText(text)
        terminal = ("✅", "❌", "🕐", "⏹", "⬜", "⚠")
        if row_index not in self._counted_rows and any(text.startswith(p) for p in terminal) and text != "⬜ Waiting":
            self._counted_rows.add(row_index)
            self._done_count += 1
            self._progress_bar.setValue(self._done_count)
            self._progress_bar.setFormat(f"{self._done_count} / {self._total} done")

    def _on_path(self, row_index, path):
        if not (0 <= row_index < self._total):
            return
        lbl = self._table.cellWidget(row_index, self._COL_PATH)
        if lbl:
            # BUG FIX (2026-09-02): this was plain setText(path) — no <a
            # href> markup and the label was never given
            # TextBrowserInteraction/linkActivated wiring at row-creation
            # time either, so the path rendered as inert text that only
            # looked clickable. Match BatchProgressDialog's row-link
            # pattern (_on_path_update) so clicking it actually opens the
            # file/folder via _open_row_path.
            lbl.setText(
                f'<a href="{path}" style="color:#2563EB;text-decoration:underline;">'
                f'{path}</a>')
            lbl.setToolTip(path)

    def batch_finished(self):
        # A disabled "Stop" button once nothing is left to stop reads as
        # broken, not finished — turn it into a real "Close" action instead.
        _bt = _t()
        self._stop_btn.setText("Close")
        self._stop_btn.setStyleSheet(
            f"QPushButton{{background:{_bt.bg_table_alt};color:{_bt.text_primary};"
            f"border:1px solid {_bt.border};border-radius:6px;padding:6px 14px;"
            f"font-weight:bold;font-size:12px;}}"
            f"QPushButton:hover{{background:{_bt.bg_input};}}")
        self._stop_btn.clicked.disconnect()
        self._stop_btn.clicked.connect(self.accept)
        self._stop_btn.setEnabled(True)
        self._tray_btn.setVisible(False)  # nothing left to send to tray for

    def _on_stop_clicked(self):
        if self._stop_callback:
            self._stop_callback()
        self._stop_btn.setEnabled(False)

    def _on_tray_clicked(self):
        if self._tray_callback:
            self._tray_callback()

    def _open_output_dir(self):
        if self._output_dir:
            _open_path(self._output_dir)

    def _open_report(self):
        if self._last_report_path:
            _open_path(self._last_report_path)

    def _open_row_path(self, url: str):
        _open_path(url)


# ── Download Picker Dialog (F-56 Phase 3) ─────────────────────────────────────

class DownloadPickerDialog(QDialog):
    """
    Replaces the old Run button's dropdown menu — a single checkbox picker
    lets the user select any combination of document types for one batch
    run instead of running each mode separately (F-56 Phase 2/3).

    Note: "ITR Return" and "Intimation Orders" are one combined checkbox,
    not two — automation.downloader_filed_returns.download_filed_returns()
    always fetches Form/Receipt/JSON and any Intimation Orders together in
    the same pass per filing, so offering them as independently toggleable
    would be misleading (unchecking one wouldn't actually skip it).
    """

    def __init__(self, parent, vault):
        super().__init__(parent)
        self._vault = vault
        self.selected_docs: set = set()
        self.filing_scope = "all"

        self.setWindowTitle("Download Documents")
        self.setMinimumWidth(440)
        self.resize(460, 460)
        self.setWindowFlags(
            Qt.WindowType.Dialog |
            Qt.WindowType.WindowTitleHint |
            Qt.WindowType.WindowCloseButtonHint)
        self._build_ui()

    def _build_ui(self):
        t = _t()
        self.setStyleSheet(
            f"QDialog{{background:{t.bg_window};}}"
            f"QLabel{{color:{t.text_primary};background:transparent;}}"
            f"QCheckBox{{color:{t.text_primary};background:transparent;"
            f"font-size:13px;font-weight:600;spacing:9px;}}"
            f"QCheckBox::indicator{{width:16px;height:16px;border:1.5px solid {t.border};"
            f"border-radius:4px;background:{t.bg_checkbox};}}"
            f"QCheckBox::indicator:hover{{border-color:{t.border_focus};}}"
            f"QCheckBox::indicator:checked{{background:{t.accent};border-color:{t.accent};}}"
            f"QRadioButton{{color:{t.text_primary};background:transparent;font-size:12px;spacing:6px;}}"
            f"QRadioButton::indicator{{width:14px;height:14px;border:1.5px solid {t.border};"
            f"border-radius:7px;background:{t.bg_checkbox};}}"
            f"QRadioButton::indicator:checked{{background:{t.accent};border-color:{t.accent};}}"
        )

        main = QVBoxLayout(self)
        main.setContentsMargins(22, 20, 22, 16)
        main.setSpacing(2)

        self._cbs = {}

        def add_option(key: str, label: str, sub: str = "", checked: bool = False):
            row = QVBoxLayout()
            row.setSpacing(2)
            cb = QCheckBox(label)
            cb.setChecked(checked)
            self._cbs[key] = cb
            row.addWidget(cb)
            if sub:
                sub_lbl = QLabel(sub)
                sub_lbl.setStyleSheet(f"font-size:11px;color:{t.text_muted};margin-left:25px;")
                row.addWidget(sub_lbl)
            main.addLayout(row)
            line = QFrame()
            line.setFrameShape(QFrame.Shape.HLine)
            line.setStyleSheet(f"color:{t.border};background:{t.border};max-height:1px;")
            main.addWidget(line)
            main.addSpacing(4)
            return cb

        add_option("26as", "26AS / Form 168",
                    "PDF + Excel/TXT — form picked automatically by year")
        add_option("request_ais", "AIS + TIS",
                    "Requests generation if not ready yet, downloads instantly if it is")
        add_option("ais_tis", "Download Previously Requested AIS",
                    "For clients whose AIS was requested earlier and should be ready now")

        itr_cb = add_option("filed_returns", "ITR Return + Intimation Orders",
                             "Form, Receipt (or ITR-V), JSON, and any Intimation Orders")

        add_option("challans", "Tax Payment Challans",
                    "Downloads challan PDFs from e-Pay Tax Payment History for the selected year")

        self._scope_panel = QFrame()
        self._scope_panel.setStyleSheet(
            f"QFrame{{background:{t.bg_table_alt};border:1px solid {t.border};"
            f"border-left:3px solid {t.accent};border-radius:6px;}}"
        )
        scope_v = QVBoxLayout(self._scope_panel)
        scope_v.setContentsMargins(12, 10, 12, 10)
        scope_v.setSpacing(4)
        scope_title = QLabel("FILING SCOPE")
        scope_title.setStyleSheet(f"font-size:10.5px;font-weight:700;color:{t.text_muted};letter-spacing:0.04em;")
        scope_v.addWidget(scope_title)

        _scope_sub_ss = f"font-size:11px;color:{t.text_muted};margin-left:23px;"

        self._rb_all = QRadioButton("All filings for the year")
        self._rb_all.setChecked(True)
        scope_v.addWidget(self._rb_all)
        all_sub = QLabel("e.g. Original, then a later Revised, Rectification, or Updated return — downloads every one")
        all_sub.setStyleSheet(_scope_sub_ss)
        all_sub.setWordWrap(True)
        scope_v.addWidget(all_sub)

        self._rb_latest = QRadioButton("Latest filing only")
        scope_v.addWidget(self._rb_latest)
        latest_sub = QLabel("same example — downloads whichever was filed most recently, by date, regardless of type")
        latest_sub.setStyleSheet(_scope_sub_ss)
        latest_sub.setWordWrap(True)
        scope_v.addWidget(latest_sub)

        main.addWidget(self._scope_panel)
        main.addSpacing(4)

        saved_scope = self._vault.get_setting("filed_returns_scope", "all")
        if saved_scope == "latest":
            self._rb_latest.setChecked(True)

        self._scope_panel.setVisible(itr_cb.isChecked())
        itr_cb.toggled.connect(self._scope_panel.setVisible)

        main.addStretch()

        footer = QHBoxLayout()
        footer.addStretch()
        cancel_btn = _btn("Cancel", "outline", height=32)
        cancel_btn.clicked.connect(self.reject)
        download_btn = _btn("Download", "primary", height=32)
        download_btn.clicked.connect(self._on_download)
        footer.addWidget(cancel_btn)
        footer.addWidget(download_btn)
        main.addLayout(footer)

    def _on_download(self):
        self.selected_docs = {key for key, cb in self._cbs.items() if cb.isChecked()}
        if not self.selected_docs:
            QMessageBox.warning(self, "Nothing Selected", "Please select at least one document type.")
            return
        self.filing_scope = "latest" if self._rb_latest.isChecked() else "all"
        try:
            self._vault.update_setting("filed_returns_scope", self.filing_scope)
        except Exception:
            pass
        self.accept()


# ── Mail Docs to Clients Dialog ───────────────────────────────────────────────

class MailDocsDialog(QDialog):
    """
    Tools → Mail Docs to Clients
    User picks a root folder, app scans for {PAN}-* sub-folders,
    matches against vault, filters by current AY, then lets user
    select clients and send emails with attachments.
    """

    _status_signal = pyqtSignal(str, str)   # pan, status text

    _COL_CHK   = 0
    _COL_NAME  = 1
    _COL_PAN   = 2
    _COL_EMAIL = 3
    _COL_CC    = 4
    _COL_FILES = 5

    def __init__(self, parent, vault, ay_label: str):
        super().__init__(parent)
        self._vault = vault
        self._ay_label = ay_label
        self._clients = []          # list of dicts from emailer.scan_for_clients
        self._sending = False
        self._client_status = {}    # pan → status string (✅ Sent / ❌ Failed / ⏳ …)
        self._scan_summary = ""     # base text from last scan

        self.setWindowTitle("Mail Docs to Clients")
        self.setMinimumSize(900, 560)
        self.resize(1020, 620)
        self.setSizeGripEnabled(True)
        self.setWindowFlags(
            Qt.WindowType.Dialog |
            Qt.WindowType.WindowTitleHint |
            Qt.WindowType.WindowCloseButtonHint |
            Qt.WindowType.WindowMaximizeButtonHint)
        self._build_ui()
        self._status_signal.connect(self._on_status)

    def _build_ui(self):
        t = _t()
        self.setStyleSheet(
            f"QDialog{{background:{t.bg_window};}}"
            f"QLabel{{color:{t.text_primary};background:transparent;}}"
            f"QLineEdit{{border:1px solid {t.border};border-radius:6px;padding:4px 8px;"
            f"font-size:11px;background:{t.bg_input};color:{t.text_primary};}}"
            f"QLineEdit:focus{{border-color:{t.border_focus};}}"
            f"QTableWidget{{border:1.5px solid {t.border};border-radius:8px;"
            f"background:{t.bg_table};outline:0;gridline-color:{t.grid};}}"
            f"QTableWidget::item{{border-bottom:1px solid {t.grid};padding:0 6px;}}"
        )

        main = QVBoxLayout(self)
        main.setContentsMargins(16, 14, 16, 12)
        main.setSpacing(8)

        # ── Title ─────────────────────────────────────────────────────────────
        title = QLabel("<b>Mail Docs to Clients</b>")
        title.setStyleSheet(f"font-size:14px;color:{t.text_primary};background:transparent;")
        main.addWidget(title)

        hint = QLabel(
            f"Select a folder, click <b>Scan</b> to find clients, then choose who to email."
            f"&nbsp;&nbsp;AY: <span style='color:{t.accent}'>{self._ay_label}</span>")
        hint.setStyleSheet(f"font-size:11px;color:{t.text_muted};background:transparent;")
        main.addWidget(hint)
        # ── Template / Folder / Filter — unified form rows ────────────────────
        from PyQt6.QtWidgets import QFormLayout
        _ROW_H = 30
        _lbl_ss  = f"font-size:11px;color:{t.text_muted};background:transparent;"
        _field_ss = (
            f"border:1px solid {t.border};border-radius:5px;padding:0 8px;"
            f"font-size:11px;background:{t.bg_input};color:{t.text_primary};"
        )
        _combo_ss = (
            f"QComboBox{{border:1px solid {t.border};border-radius:5px;padding:0 8px;"
            f"font-size:11px;background:{t.bg_input};color:{t.text_primary};}}"
            f"QComboBox::drop-down{{border:none;width:18px;}}"
            f"QComboBox QAbstractItemView{{background:{t.bg_input};color:{t.text_primary};"
            f"selection-background-color:{t.accent};}}"
        )
        _btn_ss_outline = (
            f"QPushButton{{background:transparent;color:{t.text_primary};"
            f"border:1px solid {t.border};border-radius:5px;padding:0 8px;"
            f"font-size:11px;font-weight:600;}}"
            f"QPushButton:hover{{background:{t.bg_table_alt};}}"
        )
        _btn_ss_primary = (
            f"QPushButton{{background:{t.accent};color:white;border:none;"
            f"border-radius:5px;padding:0 8px;font-size:11px;font-weight:600;}}"
            f"QPushButton:hover{{background:{t.accent_hover};}}"
            f"QPushButton:disabled{{background:{t.border};color:{t.text_muted};}}"
        )

        form = QFormLayout()
        form.setSpacing(8)
        form.setContentsMargins(0, 0, 0, 0)
        form.setLabelAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        form.setFormAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignTop)

        def _form_lbl(text):
            l = QLabel(text)
            l.setFixedHeight(_ROW_H)
            l.setStyleSheet(_lbl_ss)
            return l

        # Template row
        self._tpl_combo = QComboBox()
        self._tpl_combo.setFixedHeight(_ROW_H)
        self._tpl_combo.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        self._tpl_combo.setStyleSheet(_combo_ss)
        templates = self._vault.get_email_templates()
        active_name = self._vault.get_active_template_name()
        for tpl in templates:
            self._tpl_combo.addItem(tpl["name"])
        idx = next((i for i, t_ in enumerate(templates) if t_["name"] == active_name), 0)
        self._tpl_combo.setCurrentIndex(idx)
        form.addRow(_form_lbl("Template:"), self._tpl_combo)

        # Folder row
        folder_widget = QWidget()
        folder_widget.setStyleSheet("QWidget{background:transparent;}")
        folder_h = QHBoxLayout(folder_widget)
        folder_h.setContentsMargins(0, 0, 0, 0)
        folder_h.setSpacing(6)
        self._folder_edit = QLineEdit()
        self._folder_edit.setFixedHeight(_ROW_H)
        self._folder_edit.setPlaceholderText("Browse to the folder containing client sub-folders…")
        self._folder_edit.setStyleSheet(f"QLineEdit{{{_field_ss}}}QLineEdit:focus{{border-color:{t.border_focus};}}")
        default_dir = self._vault.get_setting("download_root_dir", "")
        if default_dir and os.path.isdir(default_dir):
            self._folder_edit.setText(default_dir)
        folder_h.addWidget(self._folder_edit, stretch=1)
        def _small_icon(name):
            p = _icon_path(name)
            if not p:
                return QIcon()
            return QIcon(QPixmap(p).scaled(16, 16, Qt.AspectRatioMode.KeepAspectRatio,
                                           Qt.TransformationMode.SmoothTransformation))

        browse_btn = QPushButton("Browse…")
        browse_btn.setIcon(_small_icon("btn_browse_folder.png"))
        browse_btn.setFixedSize(90, _ROW_H)
        browse_btn.setStyleSheet(_btn_ss_outline)
        browse_btn.clicked.connect(self._browse)
        folder_h.addWidget(browse_btn)
        self._scan_btn = QPushButton("Scan Folder")
        self._scan_btn.setIcon(_small_icon("btn_scan.png"))
        self._scan_btn.setFixedSize(110, _ROW_H)
        self._scan_btn.setStyleSheet(_btn_ss_primary)
        self._scan_btn.clicked.connect(self._scan)
        folder_h.addWidget(self._scan_btn)
        form.addRow(_form_lbl("Folder:"), folder_widget)

        # Filter row
        self._filter_edit = QLineEdit()
        self._filter_edit.setFixedHeight(_ROW_H)
        self._filter_edit.setPlaceholderText("Search by name, PAN or email…")
        self._filter_edit.setStyleSheet(f"QLineEdit{{{_field_ss}}}QLineEdit:focus{{border-color:{t.border_focus};}}")
        self._filter_edit.setClearButtonEnabled(True)
        self._filter_edit.textChanged.connect(self._apply_filter)
        form.addRow(_form_lbl("Filter:"), self._filter_edit)

        main.addLayout(form)

        # ── Table ─────────────────────────────────────────────────────────────
        self._table = QTableWidget(0, 6)
        self._table.setHorizontalHeaderLabels(
            ["", "Name  ⇅", "PAN  ⇅", "Email  ⇅", "CC  ⇅", "Files  ⇅"])

        hdr = self._table.horizontalHeader()
        hdr.setSectionResizeMode(self._COL_CHK,   QHeaderView.ResizeMode.Fixed)
        hdr.setSectionResizeMode(self._COL_NAME,  QHeaderView.ResizeMode.Interactive)
        hdr.setSectionResizeMode(self._COL_PAN,   QHeaderView.ResizeMode.Interactive)
        hdr.setSectionResizeMode(self._COL_EMAIL, QHeaderView.ResizeMode.Stretch)
        hdr.setSectionResizeMode(self._COL_CC,    QHeaderView.ResizeMode.Interactive)
        hdr.setSectionResizeMode(self._COL_FILES, QHeaderView.ResizeMode.Fixed)
        hdr.setStretchLastSection(False)
        self._table.setColumnWidth(self._COL_CHK,   36)
        self._table.setColumnWidth(self._COL_NAME, 180)
        self._table.setColumnWidth(self._COL_PAN,  110)
        self._table.setColumnWidth(self._COL_CC,   180)
        self._table.setColumnWidth(self._COL_FILES, 80)

        t2 = _t()
        hdr.setStyleSheet(
            f"QHeaderView::section{{background-color:{t2.bg_header};border:none;"
            f"border-right:1px solid {t2.border};border-bottom:1px solid {t2.border};"
            f"font-weight:bold;color:{t2.text_muted};font-size:11px;height:32px;padding:0 6px;}}")
        self._table.verticalHeader().setVisible(False)
        self._table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self._table.setSelectionMode(QAbstractItemView.SelectionMode.NoSelection)
        self._table.setShowGrid(True)
        self._table.setAlternatingRowColors(False)
        self._table.setWordWrap(False)
        self._table.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)

        self._sort_col = -1
        self._sort_order = Qt.SortOrder.AscendingOrder
        hdr.setSortIndicatorShown(False)
        hdr.sectionClicked.connect(self._on_header_clicked)

        main.addWidget(self._table, stretch=1)

        # ── Status bar ────────────────────────────────────────────────────────
        self._status_lbl = QLabel("")
        self._status_lbl.setStyleSheet(
            f"font-size:11px;color:{t.text_muted};background:transparent;")
        main.addWidget(self._status_lbl)

        # ── Footer ────────────────────────────────────────────────────────────
        footer = QHBoxLayout(); footer.setSpacing(8)

        self._sel_all_btn = _btn("Select All", "outline", height=32, icon="btn_select_all.png")
        self._sel_all_btn.clicked.connect(self._select_all)
        footer.addWidget(self._sel_all_btn)

        self._sel_none_btn = _btn("Select None", "outline", height=32, icon="btn_select_none.png")
        self._sel_none_btn.clicked.connect(self._select_none)
        footer.addWidget(self._sel_none_btn)

        email_settings_btn = _btn("Email Settings", "outline", height=32, icon="icon_email.png")
        email_settings_btn.clicked.connect(self._open_email_settings)
        footer.addWidget(email_settings_btn)

        footer.addStretch()

        self._send_btn = _btn("Send to Selected", "success", height=34, icon="btn_send.png")
        self._send_btn.setMinimumWidth(150)
        self._send_btn.setEnabled(False)
        self._send_btn.clicked.connect(self._send)
        footer.addWidget(self._send_btn)

        self._close_btn = _btn("Close", "secondary", height=34, icon="btn_close.png")
        self._close_btn.clicked.connect(self.accept)
        footer.addWidget(self._close_btn)

        main.addLayout(footer)

    def _open_email_settings(self):
        SmtpSettingsDialog(self, self._vault).exec()
        self._refresh_tpl_combo()

    def _refresh_tpl_combo(self):
        templates = self._vault.get_email_templates()
        active_name = self._vault.get_active_template_name()
        current = self._tpl_combo.currentText()
        self._tpl_combo.blockSignals(True)
        self._tpl_combo.clear()
        for tpl in templates:
            self._tpl_combo.addItem(tpl["name"])
        # Restore previous selection if still exists, else use active
        names = [t["name"] for t in templates]
        if current in names:
            self._tpl_combo.setCurrentText(current)
        elif active_name in names:
            self._tpl_combo.setCurrentText(active_name)
        self._tpl_combo.blockSignals(False)

    # ── folder browse ─────────────────────────────────────────────────────────

    def _browse(self):
        start = self._folder_edit.text().strip() or os.path.expanduser("~")
        chosen = QFileDialog.getExistingDirectory(self, "Select Folder Containing Client Sub-Folders", start)
        if chosen:
            self._folder_edit.setText(chosen)

    # ── scan ──────────────────────────────────────────────────────────────────

    def _scan(self):
        from automation.emailer import scan_for_clients
        root = self._folder_edit.text().strip()
        if not root:
            QMessageBox.warning(self, "No Folder", "Please select a folder first.")
            return
        if not os.path.isdir(root):
            QMessageBox.warning(self, "Invalid Folder", f"Folder not found:\n{root}")
            return
        if not self._ay_label or self._ay_label == "Select AY/TY":
            QMessageBox.warning(self, "No AY Selected",
                                "Please select an Assessment Year in the main window first.")
            return

        assessees = self._vault.get_all_assessees()
        self._clients = scan_for_clients(root, self._ay_label, assessees)
        self._client_status = {}
        self._filter_edit.clear()
        self._populate_table()

        n = len(self._clients)
        if n == 0:
            self._status_lbl.setText(
                "No matching clients found. Check that the folder contains {PAN}-Name sub-folders.")
            self._send_btn.setEnabled(False)
        else:
            n_files = sum(1 for c in self._clients if c["attachments"])
            self._scan_summary = f"Found {n} client(s) — {n_files} with files for {self._ay_label}."
            self._update_status_label()
            self._send_btn.setEnabled(True)

    def _update_status_label(self):
        selected_n = sum(1 for chk in self._checkboxes.values() if chk.isChecked())
        base = getattr(self, "_scan_summary", "")
        if base:
            self._status_lbl.setText(f"{base}  ·  {selected_n} selected")
        else:
            self._status_lbl.setText(f"{selected_n} selected")

    def _apply_filter(self, text: str):
        q = text.strip().lower()
        for row in range(self._table.rowCount()):
            name  = (self._table.item(row, self._COL_NAME)  or QTableWidgetItem()).text().lower()
            pan   = (self._table.item(row, self._COL_PAN)   or QTableWidgetItem()).text().lower()
            email_w = self._table.cellWidget(row, self._COL_EMAIL)
            email = email_w.text().lower() if email_w else ""
            visible = not q or q in name or q in pan or q in email
            self._table.setRowHidden(row, not visible)

    def _populate_table(self, checked_pans: set = None, status_map: dict = None):
        t = _t()
        self._table.setRowCount(0)
        self._email_edits = {}
        self._cc_edits = {}
        self._checkboxes = {}

        for row, client in enumerate(self._clients):
            self._table.insertRow(row)
            self._table.setRowHeight(row, 38)

            pan = client["pan"]
            has_files = bool(client["attachments"])  # used for Files column display
            can_select = True  # always allow selection; doc filtering happens at send time

            # Checkbox cell
            row_bg = t.bg_table_alt if row % 2 else t.bg_table
            chk_widget = QWidget()
            chk_widget.setAttribute(Qt.WidgetAttribute.WA_StyledBackground, True)
            chk_widget.setStyleSheet(f"QWidget{{background:{row_bg};}}")
            chk_layout = QHBoxLayout(chk_widget)
            chk_layout.setContentsMargins(0, 0, 0, 0)
            chk_layout.setAlignment(Qt.AlignmentFlag.AlignCenter)
            chk = QCheckBox()
            chk.setChecked(checked_pans is not None and pan in checked_pans)
            chk.setEnabled(can_select)
            chk.setStyleSheet(
                f"QCheckBox{{background:transparent;}}"
                f"QCheckBox::indicator{{width:15px;height:15px;border:1.5px solid {t.border};"
                f"border-radius:3px;background:{t.bg_checkbox};}}"
                f"QCheckBox::indicator:checked{{background:{t.accent};border-color:{t.accent};}}"
                f"QCheckBox::indicator:disabled{{background:{t.border};}}")
            chk_layout.addWidget(chk)
            self._table.setCellWidget(row, self._COL_CHK, chk_widget)
            self._checkboxes[pan] = chk
            chk.stateChanged.connect(self._update_status_label)

            # Name
            name_item = QTableWidgetItem(client["name"])
            name_item.setForeground(QColor(t.text_primary if has_files else t.text_muted))
            self._table.setItem(row, self._COL_NAME, name_item)

            # PAN
            pan_item = QTableWidgetItem(pan)
            pan_item.setForeground(QColor(t.text_muted))
            pan_item.setFont(QFont(_MONO_FONT, 10))
            pan_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter | Qt.AlignmentFlag.AlignVCenter)
            self._table.setItem(row, self._COL_PAN, pan_item)

            # Email (editable inline + hidden sort item)
            email_val = client.get("email", "")
            email_edit = QLineEdit(email_val)
            email_edit.setPlaceholderText("type email…")
            email_edit.setStyleSheet(
                f"QLineEdit{{border:none;background:{row_bg};color:{t.text_primary};"
                f"font-size:11px;padding:0 4px;}}"
                f"QLineEdit:focus{{border-bottom:1px solid {t.accent};background:{t.bg_input_focus};}}")
            self._table.setCellWidget(row, self._COL_EMAIL, email_edit)
            email_sort = QTableWidgetItem(email_val)
            self._table.setItem(row, self._COL_EMAIL, email_sort)
            self._email_edits[pan] = email_edit

            # CC (editable inline + hidden sort item)
            cc_val = client.get("cc", "")
            cc_edit = QLineEdit(cc_val)
            cc_edit.setPlaceholderText("optional; separate with ;")
            cc_edit.setStyleSheet(
                f"QLineEdit{{border:none;background:{row_bg};color:{t.text_primary};"
                f"font-size:11px;padding:0 4px;}}"
                f"QLineEdit:focus{{border-bottom:1px solid {t.accent};background:{t.bg_input_focus};}}")
            self._table.setCellWidget(row, self._COL_CC, cc_edit)
            cc_sort = QTableWidgetItem(cc_val)
            self._table.setItem(row, self._COL_CC, cc_sort)
            self._cc_edits[pan] = cc_edit

            # Files — restore live status (e.g. ✅ Sent) if available, else show count
            n_files = len(client["attachments"])
            restored = status_map.get(pan) if status_map else None
            if restored and restored not in (f"{n_files} file{'s' if n_files != 1 else ''}", "⚠ No files"):
                files_item = QTableWidgetItem(restored)
                if "✅" in restored:
                    files_item.setForeground(QColor("#15803D"))
                elif "❌" in restored:
                    files_item.setForeground(QColor("#EF4444"))
                elif "⚠" in restored:
                    files_item.setForeground(QColor("#D97706"))
                else:
                    files_item.setForeground(QColor(t.text_primary))
            elif has_files:
                tip = "\n".join(os.path.basename(f) for f in client["attachments"])
                files_item = QTableWidgetItem(f"{n_files} file{'s' if n_files != 1 else ''}")
                files_item.setForeground(QColor(t.text_primary))
                files_item.setToolTip(tip)
            else:
                files_item = QTableWidgetItem("⚠ No files")
                files_item.setForeground(QColor("#EF4444"))
            files_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter | Qt.AlignmentFlag.AlignVCenter)
            self._table.setItem(row, self._COL_FILES, files_item)

    # ── sort ──────────────────────────────────────────────────────────────────

    def _on_header_clicked(self, col: int):
        if col == self._COL_CHK:
            return
        hdr = self._table.horizontalHeader()
        if self._sort_col == col:
            self._sort_order = (
                Qt.SortOrder.DescendingOrder
                if self._sort_order == Qt.SortOrder.AscendingOrder
                else Qt.SortOrder.AscendingOrder
            )
        else:
            self._sort_col = col
            self._sort_order = Qt.SortOrder.AscendingOrder
        hdr.setSortIndicatorShown(True)
        hdr.setSortIndicator(self._sort_col, self._sort_order)
        self._sort_and_repopulate()

    def _sort_and_repopulate(self):
        # Capture current state keyed by PAN (stable, not row-index dependent)
        checked_pans = {pan for pan, chk in self._checkboxes.items() if chk.isChecked()}
        live_emails  = {pan: ed.text() for pan, ed in self._email_edits.items()}
        live_cc      = {pan: ed.text() for pan, ed in self._cc_edits.items()}
        live_status  = {pan: self._client_status.get(pan, "") for pan in self._checkboxes}

        # Merge typed email/cc back into clients list
        for c in self._clients:
            if c["pan"] in live_emails:
                c["email"] = live_emails[c["pan"]]
            if c["pan"] in live_cc:
                c["cc"] = live_cc[c["pan"]]

        # Sort clients list by chosen column
        rev = (self._sort_order == Qt.SortOrder.DescendingOrder)
        key_map = {
            self._COL_NAME:  lambda c: c["name"].lower(),
            self._COL_PAN:   lambda c: c["pan"].lower(),
            self._COL_EMAIL: lambda c: c.get("email", "").lower(),
            self._COL_CC:    lambda c: c.get("cc", "").lower(),
            self._COL_FILES: lambda c: len(c["attachments"]),
        }
        key_fn = key_map.get(self._sort_col)
        if key_fn:
            self._clients.sort(key=key_fn, reverse=rev)

        # Repopulate with sorted order, restoring checked state and status
        self._populate_table(checked_pans=checked_pans, status_map=live_status)
        # Re-apply active filter
        self._apply_filter(self._filter_edit.text())

    # ── select all / none ─────────────────────────────────────────────────────

    def _select_all(self):
        for row in range(self._table.rowCount()):
            if self._table.isRowHidden(row):
                continue
            item = self._table.item(row, self._COL_NAME)
            if not item:
                continue
            pan = self._table.item(row, self._COL_PAN)
            pan_text = pan.text() if pan else ""
            chk = self._checkboxes.get(pan_text)
            if chk and chk.isEnabled():
                chk.setChecked(True)

    def _select_none(self):
        for row in range(self._table.rowCount()):
            if self._table.isRowHidden(row):
                continue
            pan = self._table.item(row, self._COL_PAN)
            pan_text = pan.text() if pan else ""
            chk = self._checkboxes.get(pan_text)
            if chk:
                chk.setChecked(False)

    # ── send ──────────────────────────────────────────────────────────────────

    def _send(self):
        if self._sending:
            return

        cfg = self._vault.get_email_settings()
        if not cfg.get("smtp_host") or not cfg.get("smtp_user"):
            QMessageBox.warning(self, "SMTP Not Configured",
                                "Please configure email settings first.\n"
                                "Go to Settings → Email Settings.")
            return

        # Load selected template — use its subject, body and doc filter
        all_templates = self._vault.get_email_templates()
        tpl_name = self._tpl_combo.currentText()
        active_tpl = next((t for t in all_templates if t["name"] == tpl_name),
                          all_templates[0] if all_templates else None)
        if active_tpl:
            self._vault.set_active_template(active_tpl["name"])
            cfg["email_subject_tpl"] = active_tpl.get("subject", cfg.get("email_subject_tpl", ""))
            cfg["email_body_tpl"]    = active_tpl.get("body",    cfg.get("email_body_tpl", ""))
            doc_filter = active_tpl.get("docs", {})
        else:
            doc_filter = {}

        def _keep(path: str) -> bool:
            if not doc_filter:
                return True
            n = os.path.basename(path).upper()
            entry = match_doc_type(n)
            if entry and entry["template_key"]:
                return doc_filter.get(entry["template_key"], True)
            return True

        # Collect selected clients and validate emails
        selected = []
        missing_email = []
        skipped_no_docs = []   # (name, pan) of clients skipped due to no matching docs
        for i, client in enumerate(self._clients):
            pan = client["pan"]
            chk = self._checkboxes.get(pan)
            if not chk or not chk.isChecked():
                continue
            email = self._email_edits[pan].text().strip()
            if not email:
                missing_email.append(client["name"])
                self._email_edits[pan].setStyleSheet(
                    "QLineEdit{border:1.5px solid #EF4444;border-radius:4px;"
                    "background:#FEF2F2;color:#B91C1C;font-size:11px;padding:0 4px;}")
                continue
            cc = self._cc_edits[pan].text().strip()
            attachments = [a for a in client["attachments"] if _keep(a)]
            if not attachments:
                skipped_no_docs.append((client["name"], pan))
                continue
            selected.append({**client, "email": email, "cc": cc,
                             "ay_label": self._ay_label, "attachments": attachments})

        if missing_email:
            QMessageBox.warning(self, "Missing Email",
                                "Please enter email addresses for:\n" +
                                "\n".join(f"  • {n}" for n in missing_email))
            return

        if not selected and not skipped_no_docs:
            QMessageBox.information(self, "Nothing Selected", "No clients selected to email.")
            return

        if not selected:
            # All checked clients had no matching docs — mark rows and bail
            t = _t()
            for name, pan in skipped_no_docs:
                row = self._pan_to_row(pan)
                if row >= 0:
                    item = QTableWidgetItem("❌ Docs not found")
                    item.setForeground(QColor("#EF4444"))
                    item.setTextAlignment(Qt.AlignmentFlag.AlignCenter | Qt.AlignmentFlag.AlignVCenter)
                    self._table.setItem(row, self._COL_FILES, item)
            self._status_lbl.setText("No matching documents found for the selected template.")
            return

        # Save any inline-typed emails back to vault
        for client in selected:
            self._vault.update_assessee_email(client["pan"], client["email"], client.get("cc", ""))

        # Confirm
        reply = QMessageBox.question(self, "Confirm Send", "Send emails to selected clients?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.Cancel)
        if reply != QMessageBox.StandardButton.Yes:
            return

        # Update UI to sending state
        self._sending = True
        self._send_btn.setEnabled(False)
        self._send_btn.setText("Sending…")
        self._scan_btn.setEnabled(False)

        t = _t()

        # Mark skipped clients with inline error status
        for name, pan in skipped_no_docs:
            row = self._pan_to_row(pan)
            if row >= 0:
                item = QTableWidgetItem("❌ Docs not found")
                item.setForeground(QColor("#EF4444"))
                item.setTextAlignment(Qt.AlignmentFlag.AlignCenter | Qt.AlignmentFlag.AlignVCenter)
                self._table.setItem(row, self._COL_FILES, item)

        # Mark rows as pending
        for client in selected:
            pan = client["pan"]
            row = self._pan_to_row(pan)
            if row >= 0:
                item = QTableWidgetItem("⏳ Sending…")
                item.setForeground(QColor("#92400E" if t.name != "light" else "#D97706"))
                item.setTextAlignment(Qt.AlignmentFlag.AlignCenter | Qt.AlignmentFlag.AlignVCenter)
                self._table.setItem(row, self._COL_FILES, item)

        # Run send in background thread
        from ui.log_history import LogStore
        _log_store = LogStore()
        _pan_to_ay   = {c["pan"]: c.get("ay_label", self._ay_label) for c in selected}
        def _doc_label(path: str) -> str:
            n = os.path.basename(path).upper()
            entry = match_doc_type(n)
            return entry["short_label"] if entry else os.path.basename(path)

        _pan_to_docs = {
            c["pan"]: ", ".join(_doc_label(f) for f in c.get("attachments", []))
            for c in selected
        }

        def _on_progress(pan, status):
            self._status_signal.emit(pan, status)
            if status == "✅ Sent":
                try:
                    docs = _pan_to_docs.get(pan, "")
                    _log_store.record(pan, _pan_to_ay.get(pan, self._ay_label),
                                      f"[Email] {docs}" if docs else "[Email] Sent")
                except Exception:
                    pass

        def _worker():
            from automation.emailer import send_batch
            send_batch(
                cfg, selected,
                subject_tpl=cfg.get("email_subject_tpl", "Your Tax Documents — {ay}"),
                body_tpl=cfg.get("email_body_tpl", "Dear {client_name},\n\nPlease find attached your documents for {ay}.\n\nRegards,\n{firm_name}"),
                bcc_addresses=cfg.get("bcc_addresses", ""),
                progress_cb=_on_progress,
            )
            self._status_signal.emit("__done__", "")

        threading.Thread(target=_worker, daemon=True).start()

    def _pan_to_row(self, pan: str) -> int:
        for i, c in enumerate(self._clients):
            if c["pan"] == pan:
                return i
        return -1

    def _on_status(self, pan: str, status: str):
        if pan == "__done__":
            self._sending = False
            self._send_btn.setEnabled(True)
            self._send_btn.setText("Send to Selected")
            self._scan_btn.setEnabled(True)
            sent = sum(
                1 for c in self._clients
                if self._email_edits.get(c["pan"]) and
                   self._table.item(self._pan_to_row(c["pan"]), self._COL_FILES) and
                   (self._table.item(self._pan_to_row(c["pan"]), self._COL_FILES).text() or "").startswith("✅")
            )
            self._status_lbl.setText(f"Done — {sent} email(s) sent successfully.")
            return

        self._client_status[pan] = status
        row = self._pan_to_row(pan)
        if row < 0:
            return
        t = _t()
        item = QTableWidgetItem(status)
        if status.startswith("✅"):
            item.setForeground(QColor("#16A34A"))
        elif status.startswith("❌"):
            item.setForeground(QColor("#EF4444"))
        else:
            item.setForeground(QColor(t.text_muted))
        item.setTextAlignment(Qt.AlignmentFlag.AlignCenter | Qt.AlignmentFlag.AlignVCenter)
        self._table.setItem(row, self._COL_FILES, item)


# ── Return Status Dialog (F-67) ────────────────────────────────────────────

class ReturnStatusProgressDialog(QDialog):
    """
    Live progress popup for an ITR Processing Status check run (F-67).
    Deliberately a lighter sibling of ChallanGenerationProgressDialog — no
    artifact/report/output-dir/tray plumbing, since this flow never produces
    a file, just an updated status stored in the vault. Rows are keyed by
    position in `targets`, same reasoning as the challan dialog: simpler to
    reason about than re-deriving a row from PAN, and a PAN could in theory
    appear twice if the same client got queued from two different views.
    """
    _update_signal = pyqtSignal(int, str)   # row_index, status

    _COL_NAME = 0
    _COL_PAN = 1
    _COL_STATUS = 2

    def __init__(self, targets: list, ay_value: str, stop_callback=None, parent=None):
        super().__init__(parent)
        self._stop_callback = stop_callback
        self._targets = targets

        self.setWindowTitle(f"Checking ITR Processing Status — AY {ay_value} — Batch Progress")
        self.setMinimumSize(680, 380)
        self.resize(760, min(160 + len(targets) * 42, 640))
        self.setSizeGripEnabled(True)
        _bt = _t()
        self.setStyleSheet(f"QDialog{{background:{_bt.bg_window};}}")

        self._counted_rows = set()
        self._done_count = 0
        self._total = len(targets)

        layout = QVBoxLayout(self)
        layout.setContentsMargins(16, 14, 16, 12)
        layout.setSpacing(8)

        title = QLabel(f"<b>Checking ITR Processing Status</b> — {len(targets)} client(s) "
                        f"&nbsp;·&nbsp; <span style='color:{_bt.accent}'>AY {ay_value}</span>")
        title.setStyleSheet(f"font-size:14px; color:{_bt.text_primary}; background:transparent;")
        layout.addWidget(title)

        self._table = QTableWidget(len(targets), 3)
        self._table.setHorizontalHeaderLabels(["Name", "PAN", "Status"])
        hdr = self._table.horizontalHeader()
        hdr.setSectionResizeMode(self._COL_NAME, QHeaderView.ResizeMode.Interactive)
        hdr.setSectionResizeMode(self._COL_PAN, QHeaderView.ResizeMode.Interactive)
        hdr.setSectionResizeMode(self._COL_STATUS, QHeaderView.ResizeMode.Stretch)
        self._table.setColumnWidth(self._COL_NAME, 200)
        self._table.setColumnWidth(self._COL_PAN, 130)
        self._table.verticalHeader().setVisible(False)
        self._table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self._table.setSelectionMode(QAbstractItemView.SelectionMode.NoSelection)
        self._table.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        self._table.setStyleSheet(
            f"QTableWidget{{border:1.5px solid {_bt.border};border-radius:8px;"
            f"background:{_bt.bg_table};outline:0;gridline-color:{_bt.grid};}}"
            f"QTableWidget::item{{border-bottom:1px solid {_bt.grid};padding:0 8px;}}")
        hdr.setStyleSheet(
            f"QHeaderView::section{{"
            f"background-color:{_bt.bg_header};"
            f"border:none;"
            f"border-right:1px solid {_bt.border};"
            f"border-bottom:1px solid {_bt.border};"
            f"font-weight:bold;color:{_bt.text_muted};"
            f"font-size:11px;height:34px;"
            f"padding:0 8px;}}")

        for row, tgt in enumerate(targets):
            self._table.setRowHeight(row, 36)
            name_item = QTableWidgetItem(tgt.get("name", "—"))
            name_item.setForeground(QColor(_bt.text_primary))
            self._table.setItem(row, self._COL_NAME, name_item)
            pan_item = QTableWidgetItem(tgt.get("pan", ""))
            pan_item.setFont(QFont(_MONO_FONT, 10))
            pan_item.setForeground(QColor(_bt.text_muted))
            self._table.setItem(row, self._COL_PAN, pan_item)
            status_item = QTableWidgetItem("⬜ Waiting")
            status_item.setForeground(QColor(_bt.text_primary))
            self._table.setItem(row, self._COL_STATUS, status_item)

        layout.addWidget(self._table, stretch=1)

        self._progress_bar = QProgressBar()
        self._progress_bar.setRange(0, self._total)
        self._progress_bar.setFixedHeight(18)
        self._progress_bar.setTextVisible(True)
        self._progress_bar.setFormat(f"0 / {self._total} done")
        self._progress_bar.setStyleSheet(
            f"QProgressBar{{border:1px solid {_bt.border};border-radius:9px;"
            f"background:{_bt.scrollbar_handle};text-align:center;font-size:11px;"
            f"font-weight:600;color:{_bt.accent_text};}}"
            f"QProgressBar::chunk{{background:#16A34A;border-radius:9px;}}")
        layout.addWidget(self._progress_bar)

        footer = QHBoxLayout()
        footer.addStretch(1)
        self._stop_btn = _btn("⏹  Stop", "danger")
        self._stop_btn.clicked.connect(self._on_stop_clicked)
        footer.addWidget(self._stop_btn)
        layout.addLayout(footer)

        self._update_signal.connect(self._on_update)

    def set_status(self, row_index, text):
        self._update_signal.emit(row_index, text)

    def _on_update(self, row_index, text):
        if not (0 <= row_index < self._total):
            return
        self._table.item(row_index, self._COL_STATUS).setText(text)
        terminal = ("✅", "❌", "🕐", "⏹", "⬜", "⚠")
        if row_index not in self._counted_rows and any(text.startswith(p) for p in terminal) and text != "⬜ Waiting":
            self._counted_rows.add(row_index)
            self._done_count += 1
            self._progress_bar.setValue(self._done_count)
            self._progress_bar.setFormat(f"{self._done_count} / {self._total} done")

    def batch_finished(self):
        _bt = _t()
        self._stop_btn.setText("Close")
        self._stop_btn.setStyleSheet(
            f"QPushButton{{background:{_bt.bg_table_alt};color:{_bt.text_primary};"
            f"border:1px solid {_bt.border};border-radius:6px;padding:6px 14px;"
            f"font-weight:bold;font-size:12px;}}"
            f"QPushButton:hover{{background:{_bt.bg_input};}}")
        self._stop_btn.clicked.disconnect()
        self._stop_btn.clicked.connect(self.accept)
        self._stop_btn.setEnabled(True)

    def _on_stop_clicked(self):
        if self._stop_callback:
            self._stop_callback()
        self._stop_btn.setEnabled(False)


class ReturnStatusDialog(QDialog):
    """
    Menu entry point: "Return Status > Check Processing Status...". Browse,
    filter, and select clients/AY, then hand off to the parent window's
    start_return_status_check() to actually run the live portal check (same
    QThread + fresh-asyncio-loop worker pattern every other automation flow
    in this app already uses) — this dialog itself never touches Playwright.

    Persisted status (vault.get_return_status/_all) is what's shown; "Update
    Selected" is the only thing that ever changes it, per the user's own
    answer that this is a portal re-check, not a manual status edit.
    """
    _COL_CHECK = 0
    _COL_PAN = 1
    _COL_NAME = 2
    _COL_AY = 3
    _COL_STATUS = 4
    _COL_LAST_CHECKED = 5

    def __init__(self, parent, vault, ay_entries):
        super().__init__(parent)
        self._vault = vault
        # Only AY-bearing entries make sense here — the "View Filed Returns"
        # AY filter (and so the whole check_return_status() flow) works off
        # the real AY string, not a not-yet-open TY-only current year.
        self._ay_entries = [e for e in ay_entries if e.get("year", {}).get("AY")]
        self._row_data: list = []
        # (pan, ay) pairs the user has checked — kept across _refresh_table()
        # rebuilds (search text changing, AY switching back and forth) so
        # ticking a few rows then narrowing the search further doesn't
        # silently lose the earlier selections.
        self._checked_keys: set = set()

        self.setWindowTitle("ITR Processing Status")
        self.setMinimumSize(860, 520)
        self.resize(940, 580)
        self.setSizeGripEnabled(True)
        _bt = _t()
        self.setStyleSheet(f"QDialog{{background:{_bt.bg_window};}}")

        self._build_ui()
        self._refresh_table()

    def _build_ui(self):
        _bt = _t()
        layout = QVBoxLayout(self)
        layout.setContentsMargins(16, 14, 16, 12)
        layout.setSpacing(8)

        filter_row = QHBoxLayout()
        filter_row.setSpacing(10)
        self._search = QLineEdit()
        self._search.setPlaceholderText("🔍  Search by name or PAN...")
        self._search.setClearButtonEnabled(True)
        self._search.textChanged.connect(self._refresh_table)
        filter_row.addWidget(self._search, 1)

        filter_row.addWidget(_lbl("Year:"))
        self._ay_combo = QComboBox()
        self._ay_combo.addItem("All Years", "")
        for e in self._ay_entries:
            y = e.get("year", {})
            self._ay_combo.addItem(e.get("label", y.get("AY", "")), y.get("AY", ""))
        self._ay_combo.currentIndexChanged.connect(self._refresh_table)
        filter_row.addWidget(self._ay_combo)

        # F-11: Group filter, same "All X + each real value" shape as the
        # Year combo above, ANDed with it and the search box below.
        filter_row.addWidget(_lbl("Group:"))
        self._group_combo = QComboBox()
        self._group_combo.addItem("All Groups", "")
        for g in self._vault.get_all_groups():
            self._group_combo.addItem(g, g)
        self._group_combo.currentIndexChanged.connect(self._refresh_table)
        filter_row.addWidget(self._group_combo)

        self._btn_update = _btn("🔄  Update Selected", "primary")
        self._btn_update.clicked.connect(self._on_update_clicked)
        filter_row.addWidget(self._btn_update)
        layout.addLayout(filter_row)

        self._table = QTableWidget(0, 6)
        self._table.setHorizontalHeaderLabels(
            ["", "PAN", "Name", "AY", "Status", "Last Checked"])
        hdr = self._table.horizontalHeader()
        hdr.setSectionResizeMode(self._COL_CHECK, QHeaderView.ResizeMode.Fixed)
        hdr.setSectionResizeMode(self._COL_PAN, QHeaderView.ResizeMode.Interactive)
        hdr.setSectionResizeMode(self._COL_NAME, QHeaderView.ResizeMode.Stretch)
        hdr.setSectionResizeMode(self._COL_AY, QHeaderView.ResizeMode.Interactive)
        hdr.setSectionResizeMode(self._COL_STATUS, QHeaderView.ResizeMode.Interactive)
        hdr.setSectionResizeMode(self._COL_LAST_CHECKED, QHeaderView.ResizeMode.Interactive)
        self._table.setColumnWidth(self._COL_CHECK, 32)
        self._table.setColumnWidth(self._COL_PAN, 110)
        self._table.setColumnWidth(self._COL_AY, 80)
        self._table.setColumnWidth(self._COL_STATUS, 320)
        self._table.setColumnWidth(self._COL_LAST_CHECKED, 140)
        self._table.verticalHeader().setVisible(False)
        self._table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self._table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self._table.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        self._table.setStyleSheet(
            f"QTableWidget{{border:1.5px solid {_bt.border};border-radius:8px;"
            f"background:{_bt.bg_table};outline:0;gridline-color:{_bt.grid};}}"
            f"QTableWidget::item{{border-bottom:1px solid {_bt.grid};padding:0 8px;}}")
        hdr.setStyleSheet(
            f"QHeaderView::section{{"
            f"background-color:{_bt.bg_header};"
            f"border:none;"
            f"border-right:1px solid {_bt.border};"
            f"border-bottom:1px solid {_bt.border};"
            f"font-weight:bold;color:{_bt.text_muted};"
            f"font-size:11px;height:34px;"
            f"padding:0 8px;}}")
        layout.addWidget(self._table, stretch=1)

        hint = QLabel(
            "Pick a specific Year to select clients and check their current status "
            "live from the portal. \"All Years\" is browse-only.")
        hint.setStyleSheet(f"color:{_bt.text_muted};font-size:11px;background:transparent;")
        layout.addWidget(hint)

        footer = QHBoxLayout()
        self._counts_label = QLabel("")
        self._counts_label.setStyleSheet(f"color:{_bt.text_muted};font-size:12px;background:transparent;")
        footer.addWidget(self._counts_label, 1)
        self._sel_all_btn = _btn("Select All", "outline")
        self._sel_all_btn.clicked.connect(self._select_all)
        footer.addWidget(self._sel_all_btn)
        self._sel_none_btn = _btn("Select None", "outline")
        self._sel_none_btn.clicked.connect(self._select_none)
        footer.addWidget(self._sel_none_btn)
        btn_close = _btn("Close", "outline")
        btn_close.clicked.connect(self.accept)
        footer.addWidget(btn_close)
        layout.addLayout(footer)

    def _select_all(self):
        # No-op in the "All Years" view — every checkbox there is disabled
        # (isEnabled() reflects allow_select from _refresh_table), matching
        # the existing rule that only a specific Year's rows can be picked.
        for i in range(self._table.rowCount()):
            chk = self._table.cellWidget(i, self._COL_CHECK)
            if chk and chk.isEnabled():
                chk.setChecked(True)

    def _select_none(self):
        for i in range(self._table.rowCount()):
            chk = self._table.cellWidget(i, self._COL_CHECK)
            if chk:
                chk.setChecked(False)

    def _refresh_table(self, *_args):
        _bt = _t()
        ay_value = self._ay_combo.currentData()
        search = self._search.text().strip().lower()
        clients = self._vault.get_all_assessees()
        by_pan = {c.get("pan", "").upper(): c for c in clients}

        rows = []
        if ay_value:
            # One row per known client, whether or not they've been checked
            # yet — "⚠ Not yet checked" is itself useful information, not
            # something to hide until a first check happens to have run.
            status_map = self._vault.get_return_status(ay_value)
            for c in clients:
                pan = c.get("pan", "").upper()
                entry = status_map.get(pan)
                rows.append({
                    "pan": pan, "name": c.get("name", ""), "ay": ay_value,
                    "status": entry.get("status", "") if entry else "",
                    "status_date": entry.get("status_date", "") if entry else "",
                    "ts": entry.get("ts", "") if entry else "",
                })
        else:
            # "All Years" is a browse of everything ever checked — not every
            # client × every enabled year, which would mostly be noise for
            # years nobody has asked about yet.
            for pan, ay_map in self._vault.get_return_status_all().items():
                client = by_pan.get(pan)
                name = client.get("name", "") if client else ""
                for ay_label, entry in ay_map.items():
                    rows.append({
                        "pan": pan, "name": name, "ay": ay_label,
                        "status": entry.get("status", ""),
                        "status_date": entry.get("status_date", ""),
                        "ts": entry.get("ts", ""),
                    })
            rows.sort(key=lambda r: (r["name"].lower(), r["ay"]))

        if search:
            rows = [r for r in rows if search in r["name"].lower() or search in r["pan"].lower()]

        group_value = self._group_combo.currentData()
        if group_value:
            rows = [r for r in rows
                    if by_pan.get(r["pan"], {}).get("group", "").strip() == group_value]

        self._row_data = rows
        allow_select = bool(ay_value)
        self._table.setRowCount(len(rows))
        for i, r in enumerate(rows):
            self._table.setRowHeight(i, 34)

            key = (r["pan"], r["ay"])
            chk = QCheckBox()
            chk.setEnabled(allow_select)
            chk.setChecked(key in self._checked_keys)
            chk.setStyleSheet(
                f"QCheckBox{{background:transparent;}}"
                f"QCheckBox::indicator{{width:15px;height:15px;border:1.5px solid {_bt.border};"
                f"border-radius:3px;background:{_bt.bg_checkbox};}}"
                f"QCheckBox::indicator:checked{{background:{_bt.accent};border-color:{_bt.accent};}}"
                f"QCheckBox::indicator:disabled{{background:{_bt.border};}}")
            chk.stateChanged.connect(
                lambda state, k=key: self._checked_keys.add(k) if state
                else self._checked_keys.discard(k))
            self._table.setCellWidget(i, self._COL_CHECK, chk)

            pan_item = QTableWidgetItem(r["pan"])
            pan_item.setFont(QFont(_MONO_FONT, 10))
            pan_item.setForeground(QColor(_bt.text_muted))
            self._table.setItem(i, self._COL_PAN, pan_item)

            name_item = QTableWidgetItem(r["name"] or "⚠ Unknown PAN")
            name_item.setForeground(QColor(_bt.text_primary if r["name"] else
                                            getattr(_bt, "warning", "#D97706")))
            self._table.setItem(i, self._COL_NAME, name_item)

            ay_item = QTableWidgetItem(r["ay"])
            ay_item.setForeground(QColor(_bt.text_primary))
            self._table.setItem(i, self._COL_AY, ay_item)

            if r["status"]:
                status_text = f"{r['status']} ({r['status_date']})" if r["status_date"] else r["status"]
            else:
                status_text = "⚠ Not yet checked"
            status_item = QTableWidgetItem(status_text)
            status_item.setForeground(QColor(_bt.text_primary if r["status"] else
                                              getattr(_bt, "warning", "#D97706")))
            self._table.setItem(i, self._COL_STATUS, status_item)

            ts_item = QTableWidgetItem(r["ts"] or "—")
            ts_item.setForeground(QColor(_bt.text_muted))
            self._table.setItem(i, self._COL_LAST_CHECKED, ts_item)

        self._btn_update.setEnabled(allow_select)
        self._counts_label.setText(f"Rows: {len(rows)}")

    def _on_update_clicked(self):
        ay_value = self._ay_combo.currentData()
        if not ay_value:
            return
        # Scoped to the CURRENT AY only — _checked_keys can still hold
        # selections made while a different Year was picked (persisted
        # across _refresh_table() rebuilds deliberately, see __init__), but
        # one Update run must never mix years across its rows.
        selected_pans = [pan for (pan, ay) in self._checked_keys if ay == ay_value]
        if not selected_pans:
            QMessageBox.information(self, "Nothing Selected",
                                     "Select at least one client to update.")
            return

        by_pan = {c.get("pan", "").upper(): c for c in self._vault.get_all_assessees()}
        targets = [by_pan[p] for p in selected_pans if p in by_pan]
        if not targets:
            QMessageBox.warning(self, "Nothing to Update",
                                 "None of the selected clients could be matched in Client Master.")
            return

        parent = self.parent()
        if parent is not None and hasattr(parent, "start_return_status_check"):
            parent.start_return_status_check(ay_value, targets, on_done=self._refresh_table)
